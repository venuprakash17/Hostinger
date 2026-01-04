"""Job Round Management API endpoints"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Body
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import csv
import io
from openpyxl import load_workbook

from app.core.database import get_db
from app.models.job import Job, JobApplication
from app.models.job_round import JobRound, JobApplicationRound
from app.models.user import User, UserRole, RoleEnum
from app.models.profile import Profile
from app.schemas.job_round import (
    JobRoundCreate, JobRoundUpdate, JobRoundResponse,
    JobApplicationRoundCreate, JobApplicationRoundUpdate, JobApplicationRoundResponse,
    BulkRoundUpdateRequest
)
from app.api.auth import get_current_user, get_current_admin

router = APIRouter(prefix="/job-rounds", tags=["job-rounds"])


@router.post("/jobs/{job_id}/rounds", response_model=JobRoundResponse)
async def create_job_round(
    job_id: int,
    round_data: JobRoundCreate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Create a round for a job (College Admin only)"""
    # Verify job exists and belongs to admin's college
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Verify admin has access to this job
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        # College Admin - verify job belongs to their college
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or not admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin must be associated with a college"
            )
        
        if job.college_id != admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only manage rounds for jobs in your college"
            )
    
    # Check for duplicate round name
    existing = db.query(JobRound).filter(
        JobRound.job_id == job_id,
        JobRound.name == round_data.name
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Round '{round_data.name}' already exists for this job"
        )
    
    # Ensure Round 0 exists (Step 0 - Applied)
    round_0 = db.query(JobRound).filter(
        JobRound.job_id == job_id,
        JobRound.order == 0
    ).first()
    
    if not round_0:
        # Create Round 0 if it doesn't exist
        round_0 = JobRound(
            job_id=job_id,
            name="Applied",
            order=0,
            description="Default round for tracking all students who applied for this job",
            is_active=True
        )
        db.add(round_0)
        db.flush()
    
    # Prevent creating another Round 0
    if round_data.order == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Round 0 (Applied) already exists and cannot be recreated. It is automatically created for every job."
        )
    
    new_round = JobRound(
        job_id=job_id,
        name=round_data.name,
        order=round_data.order,
        description=round_data.description,
        is_active=round_data.is_active
    )
    
    db.add(new_round)
    db.commit()
    db.refresh(new_round)
    
    return new_round


@router.get("/jobs/{job_id}/rounds", response_model=List[JobRoundResponse])
async def list_job_rounds(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List all rounds for a job. Automatically creates Round 0 (Applied) if it doesn't exist."""
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Ensure Round 0 (Applied) exists for this job
    round_0 = db.query(JobRound).filter(
        JobRound.job_id == job_id,
        JobRound.order == 0
    ).first()
    
    if not round_0:
        # Create Round 0 if it doesn't exist (for existing jobs created before this feature)
        round_0 = JobRound(
            job_id=job_id,
            name="Applied",
            order=0,
            description="Default round for tracking all students who applied for this job",
            is_active=True
        )
        db.add(round_0)
        db.commit()
        db.refresh(round_0)
    
    rounds = db.query(JobRound).filter(
        JobRound.job_id == job_id,
        JobRound.is_active == True
    ).order_by(JobRound.order).all()
    
    return rounds


@router.put("/rounds/{round_id}", response_model=JobRoundResponse)
async def update_job_round(
    round_id: int,
    round_data: JobRoundUpdate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update a job round (College Admin only)"""
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Prevent modifying Round 0 (Applied - default round) order or name
    if round_obj.order == 0:
        update_data = round_data.model_dump(exclude_unset=True)
        # Allow only description and is_active updates for Round 0
        if 'order' in update_data and update_data['order'] != 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Round 0 (Applied) order cannot be changed"
            )
        if 'name' in update_data and update_data['name'] != round_obj.name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Round 0 (Applied) name cannot be changed"
            )
    
    # Verify admin has access
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or job.college_id != admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only manage rounds for jobs in your college"
            )
    
    update_data = round_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(round_obj, field, value)
    
    db.commit()
    db.refresh(round_obj)
    
    return round_obj


@router.delete("/rounds/{round_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_job_round(
    round_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Delete a job round (College Admin only)"""
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Prevent deleting Round 0 (Applied - default round)
    if round_obj.order == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Round 0 (Applied) cannot be deleted. It is the default round for tracking all applicants."
        )
    
    # Verify admin has access
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or job.college_id != admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only manage rounds for jobs in your college"
            )
    
    db.delete(round_obj)
    db.commit()
    
    return None


@router.get("/rounds/{round_id}/students", response_model=List[dict])
async def get_round_students(
    round_id: int,
    status_filter: Optional[str] = None,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Get students in a round with filtering logic:
    - For current round: Show only students who haven't been promoted to next round (left behind)
    - For next round: Show only students who were promoted from previous round (are in that round)
    """
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Verify admin has access
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or job.college_id != admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only view rounds for jobs in your college"
            )
    
    # Find next round (if exists) - used to filter out promoted students from current round
    next_round = db.query(JobRound).filter(
        JobRound.job_id == round_obj.job_id,
        JobRound.order == round_obj.order + 1
    ).first()
    
    # Get all students in current round
    round_statuses = db.query(JobApplicationRound).filter(
        JobApplicationRound.round_id == round_id
    ).all()
    
    result = []
    for round_status in round_statuses:
        if status_filter and round_status.status != status_filter:
            continue
        
        # Get application
        app = db.query(JobApplication).filter(
            JobApplication.id == round_status.job_application_id
        ).first()
        
        if not app:
            continue
        
        # Filtering logic:
        # When viewing a round, show only students who are "left behind" (not promoted to next round)
        # This means: if student exists in next round, exclude them from current round view
        if next_round:
            # Check if this student exists in the next round (was promoted)
            next_round_status = db.query(JobApplicationRound).filter(
                JobApplicationRound.job_application_id == app.id,
                JobApplicationRound.round_id == next_round.id
            ).first()
            
            if next_round_status:
                # Student was promoted to next round, exclude from current round view
                # (They will appear in the next round's student list instead)
                continue
        
        # Get student profile
        profile = db.query(Profile).filter(Profile.user_id == app.user_id).first()
        
        result.append({
            "application_id": app.id,
            "student_id": app.user_id,
            "student_name": profile.full_name if profile else None,
            "email": profile.email if profile else None,
            "roll_number": profile.roll_number if profile else None,
            "status": round_status.status if round_status else "PENDING",
            "remarks": round_status.remarks if round_status else None,
            "updated_at": round_status.updated_at if round_status else None
        })
    
    return result


@router.post("/rounds/{round_id}/bulk-upload")
async def bulk_upload_round_results(
    round_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Bulk upload round results from Excel/CSV (College Admin only)"""
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Verify admin has access
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or job.college_id != admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only upload results for rounds in your college"
            )
    
    # Parse file
    file_ext = file.filename.split('.')[-1].lower() if file.filename else ''
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    data = []
    
    if file_ext == 'csv':
        csv_content = contents.decode('utf-8')
        reader = csv.DictReader(io.StringIO(csv_content))
        data = list(reader)
    else:
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    value = row[idx]
                    row_dict[header] = str(value).strip() if value is not None else ""
            if row_dict:
                data.append(row_dict)
    
    # Validate and process data
    results = {
        "success": [],
        "failed": [],
        "total": len(data)
    }
    
    # Get all rounds for this job to determine next round
    all_rounds = db.query(JobRound).filter(
        JobRound.job_id == job.id,
        JobRound.is_active == True
    ).order_by(JobRound.order).all()
    
    current_round_index = next((i for i, r in enumerate(all_rounds) if r.id == round_id), -1)
    next_round = all_rounds[current_round_index + 1] if current_round_index >= 0 and current_round_index + 1 < len(all_rounds) else None
    
    for idx, row in enumerate(data, start=2):  # Start at 2 (row 1 is header)
        try:
            # Normalize keys
            normalized_row = {k.lower().strip(): v for k, v in row.items()}
            
            # Validate Round ID if provided (for consistency check)
            round_id_str = normalized_row.get('round_id', '').strip()
            if round_id_str:
                try:
                    provided_round_id = int(round_id_str)
                    if provided_round_id != round_id:
                        results["failed"].append({
                            "row": idx,
                            "error": f"Round ID mismatch: Template has Round ID {provided_round_id}, but uploading to Round ID {round_id}. Please use the correct template for this round.",
                            "data": row
                        })
                        continue
                except ValueError:
                    # Round ID is optional, so invalid format is not critical
                    pass
            
            # Extract required fields - Roll Number is the primary identifier
            roll_number = normalized_row.get('roll_number', '').strip()
            if not roll_number:
                # Try legacy 'student_id' for backward compatibility
                roll_number = normalized_row.get('student_id', '').strip()
                if not roll_number:
                    results["failed"].append({
                        "row": idx,
                        "error": "Missing roll_number (required - this is the primary identifier for tracking)",
                        "data": row
                    })
                    continue
            
            # Find student by roll number
            profile = db.query(Profile).filter(Profile.roll_number == roll_number).first()
            if not profile:
                results["failed"].append({
                    "row": idx,
                    "error": f"Student not found with roll number: {roll_number}",
                    "data": row
                })
                continue
            
            student_id = profile.user_id
            
            status_value = normalized_row.get('status', '').strip().upper()
            if status_value not in ['QUALIFIED', 'REJECTED', 'ABSENT']:
                results["failed"].append({
                    "row": idx,
                    "error": f"Invalid status: {status_value}. Must be QUALIFIED, REJECTED, or ABSENT",
                    "data": row
                })
                continue
            
            remarks = normalized_row.get('remarks', '').strip() or None
            
            # Get or create application (for external applications, create if doesn't exist)
            application = db.query(JobApplication).filter(
                JobApplication.job_id == job.id,
                JobApplication.user_id == student_id
            ).first()
            
            if not application:
                # Auto-create application for students who applied externally
                # This allows tracking progress even if they didn't apply through the system
                application = JobApplication(
                    job_id=job.id,
                    user_id=student_id,
                    status="Applied",
                    notes="Application created automatically - student applied externally"
                )
                db.add(application)
                db.flush()  # Flush to get the application ID
            
            # Verify student eligibility (year/branch)
            profile = db.query(Profile).filter(Profile.user_id == student_id).first()
            if profile:
                # Check year eligibility
                if job.eligible_years:
                    user_year = profile.present_year
                    if user_year:
                        from app.core.year_utils import parse_year, format_year
                        user_year_normalized = parse_year(user_year)
                        user_year_formatted = format_year(user_year_normalized) if user_year_normalized else None
                        eligible_years_list = job.eligible_years if isinstance(job.eligible_years, list) else [job.eligible_years]
                        
                        year_match = False
                        for eligible_year in eligible_years_list:
                            if eligible_year is None:
                                continue
                            eligible_year_normalized = parse_year(str(eligible_year))
                            if user_year_normalized and eligible_year_normalized and user_year_normalized == eligible_year_normalized:
                                year_match = True
                                break
                            if str(user_year).upper().strip() == str(eligible_year).upper().strip():
                                year_match = True
                                break
                        
                        if not year_match:
                            results["failed"].append({
                                "row": idx,
                                "error": f"Student {student_id} is not eligible for this job (year mismatch)",
                                "data": row
                            })
                            continue
                
                # Check branch eligibility
                if job.eligibility_type == "branch" and job.eligible_branches:
                    user_dept_name = profile.department
                    user_dept_code = None
                    if profile.department_id:
                        from app.models.academic import Department
                        dept = db.query(Department).filter(Department.id == profile.department_id).first()
                        if dept:
                            user_dept_code = dept.code
                    
                    eligible_branches_list = job.eligible_branches if isinstance(job.eligible_branches, list) else [job.eligible_branches]
                    branch_match = False
                    
                    for branch in eligible_branches_list:
                        if branch is None:
                            continue
                        branch_normalized = str(branch).upper().strip()
                        if user_dept_name and user_dept_name.upper().strip() == branch_normalized:
                            branch_match = True
                            break
                        if user_dept_code and user_dept_code.upper().strip() == branch_normalized:
                            branch_match = True
                            break
                    
                    if not branch_match:
                        results["failed"].append({
                            "row": idx,
                            "error": f"Student {student_id} is not eligible for this job (branch mismatch)",
                            "data": row
                        })
                        continue
            
            # Check if student was in previous round (if not first round)
            if current_round_index > 0:
                prev_round = all_rounds[current_round_index - 1]
                prev_round_status = db.query(JobApplicationRound).filter(
                    JobApplicationRound.job_application_id == application.id,
                    JobApplicationRound.round_id == prev_round.id
                ).first()
                
                if not prev_round_status or prev_round_status.status != "QUALIFIED":
                    results["failed"].append({
                        "row": idx,
                        "error": f"Student {student_id} did not qualify in previous round",
                        "data": row
                    })
                    continue
            
            # Create or update round status
            round_status = db.query(JobApplicationRound).filter(
                JobApplicationRound.job_application_id == application.id,
                JobApplicationRound.round_id == round_id
            ).first()
            
            if round_status:
                round_status.status = status_value
                round_status.remarks = remarks
                round_status.updated_by = current_user.id
                round_status.updated_at = datetime.utcnow()
            else:
                round_status = JobApplicationRound(
                    job_application_id=application.id,
                    round_id=round_id,
                    status=status_value,
                    remarks=remarks,
                    updated_by=current_user.id
                )
                db.add(round_status)
            
            # If qualified and there's a next round, auto-create entry for next round
            if status_value == "QUALIFIED" and next_round:
                next_round_status = db.query(JobApplicationRound).filter(
                    JobApplicationRound.job_application_id == application.id,
                    JobApplicationRound.round_id == next_round.id
                ).first()
                
                if not next_round_status:
                    next_round_status = JobApplicationRound(
                        job_application_id=application.id,
                        round_id=next_round.id,
                        status="PENDING",
                        updated_by=current_user.id
                    )
                    db.add(next_round_status)
            
            # Update application status
            if status_value == "QUALIFIED" and next_round:
                application.current_round = next_round.name
            elif status_value in ["REJECTED", "ABSENT"]:
                application.status = "Rejected"
                application.current_round = round_obj.name
            
            results["success"].append({
                "row": idx,
                "student_id": student_id,
                "status": status_value
            })
            
        except Exception as e:
            results["failed"].append({
                "row": idx,
                "error": str(e),
                "data": row
            })
    
    db.commit()
    
    return {
        "message": f"Processed {results['total']} rows. {len(results['success'])} successful, {len(results['failed'])} failed.",
        "results": results
    }


@router.post("/rounds/{round_id}/add-students")
async def add_students_to_round(
    round_id: int,
    roll_numbers: List[str] = Body(..., description="List of student roll numbers to add to the round"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Add students to a round using roll numbers (College Admin only)
    
    This endpoint allows adding students directly to a round by providing their roll numbers.
    The system will:
    1. Verify students exist (by roll number) and have applied for the job
    2. Create JobApplicationRound entries with PENDING status
    3. Fetch student data from backend automatically
    """
    
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Verify admin has access
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or job.college_id != admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only add students to rounds in your college"
            )
    
    if not roll_numbers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one roll number is required"
        )
    
    results = {
        "success": [],
        "failed": [],
        "total": len(roll_numbers)
    }
    
    for roll_number in roll_numbers:
        roll_number = roll_number.strip()  # Clean whitespace
        if not roll_number:
            results["failed"].append({
                "roll_number": roll_number,
                "error": "Empty roll number"
            })
            continue
            
        try:
            # Find student by roll number
            profile = db.query(Profile).filter(Profile.roll_number == roll_number).first()
            if not profile:
                results["failed"].append({
                    "roll_number": roll_number,
                    "error": "Student not found with this roll number"
                })
                continue
            
            user_id = profile.user_id
            
            # Get or create application (for external applications, create if doesn't exist)
            application = db.query(JobApplication).filter(
                JobApplication.job_id == job.id,
                JobApplication.user_id == user_id
            ).first()
            
            application_created = False
            if not application:
                # Auto-create application for students who applied externally
                # This allows tracking progress even if they didn't apply through the system
                application = JobApplication(
                    job_id=job.id,
                    user_id=user_id,
                    status="Applied",
                    notes="Application created automatically - student applied externally"
                )
                db.add(application)
                db.flush()  # Flush to get the application ID
                application_created = True
            
            # Check if entry already exists
            existing = db.query(JobApplicationRound).filter(
                JobApplicationRound.job_application_id == application.id,
                JobApplicationRound.round_id == round_id
            ).first()
            
            if existing:
                results["failed"].append({
                    "roll_number": roll_number,
                    "error": "Student already exists in this round"
                })
                continue
            
            # Create new round entry
            round_status = JobApplicationRound(
                job_application_id=application.id,
                round_id=round_id,
                status="PENDING",
                updated_by=current_user.id
            )
            db.add(round_status)
            
            # Update application current_round and status
            application.current_round = round_obj.name
            
            # If added to "Selected" round, mark as job cracked
            if round_obj.name.lower() == "selected":
                application.status = "Selected"
            
            results["success"].append({
                "roll_number": roll_number,
                "student_id": user_id,
                "student_name": profile.full_name if profile else None,
                "email": profile.email if profile else None,
                "application_created": application_created
            })
            
        except Exception as e:
            results["failed"].append({
                "roll_number": roll_number,
                "error": str(e)
            })
    
    db.commit()
    
    return {
        "message": f"Added {len(results['success'])} students. {len(results['failed'])} failed.",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.put("/rounds/{round_id}/students/{student_id}")
async def update_student_round_status(
    round_id: int,
    student_id: int,
    status: str = Body(..., description="Status: QUALIFIED, REJECTED, ABSENT, or PENDING"),
    remarks: Optional[str] = Body(None, description="Optional remarks"),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update individual student's status in a round (College Admin only)"""
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Verify admin has access
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or job.college_id != admin_role.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only update students in rounds from your college"
            )
    
    # Validate status
    status_upper = status.upper().strip()
    if status_upper not in ['QUALIFIED', 'REJECTED', 'ABSENT', 'PENDING']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Status must be QUALIFIED, REJECTED, ABSENT, or PENDING"
        )
    
    # Find application
    application = db.query(JobApplication).filter(
        JobApplication.job_id == job.id,
        JobApplication.user_id == student_id
    ).first()
    
    if not application:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student has not applied for this job"
        )
    
    # Get or create round status
    round_status = db.query(JobApplicationRound).filter(
        JobApplicationRound.job_application_id == application.id,
        JobApplicationRound.round_id == round_id
    ).first()
    
    if not round_status:
        # Create if doesn't exist
        round_status = JobApplicationRound(
            job_application_id=application.id,
            round_id=round_id,
            status=status_upper,
            remarks=remarks,
            updated_by=current_user.id
        )
        db.add(round_status)
    else:
        # Update existing
        round_status.status = status_upper
        if remarks is not None:
            round_status.remarks = remarks
        round_status.updated_by = current_user.id
        round_status.updated_at = datetime.utcnow()
    
    # If rejected/absent in next round, they stay in current round (don't remove from current)
    # Only update the status in the current round
    if status_upper in ["REJECTED", "ABSENT"]:
        # Check if this is a next round rejection - if so, student stays in current round
        all_rounds = db.query(JobRound).filter(
            JobRound.job_id == job.id,
            JobRound.is_active == True
        ).order_by(JobRound.order).all()
        
        current_round_index = next((i for i, r in enumerate(all_rounds) if r.id == round_id), -1)
        prev_round = all_rounds[current_round_index - 1] if current_round_index > 0 else None
        
        # If rejected in next round, they remain in previous round
        # Don't update application status to "Rejected" - keep them in current round
        if prev_round:
            # Student stays in previous round, just mark as rejected in current round
            application.current_round = prev_round.name
        else:
            # This is the first round or Round 0 - update application status
            application.status = "Rejected"
            application.current_round = round_obj.name
    
    db.commit()
    db.refresh(round_status)
    
    return {
        "message": "Student status updated successfully",
        "round_status": {
            "id": round_status.id,
            "status": round_status.status,
            "remarks": round_status.remarks,
            "updated_at": round_status.updated_at
        }
    }


@router.post("/rounds/{round_id}/bulk-promote")
async def bulk_promote_students(
    round_id: int,
    request_data: dict = Body(default_factory=dict, description="Optional dict with 'student_ids' list. If not provided, promotes all students."),
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Bulk promote students from current round to next round.
    
    If request_data contains 'student_ids', only promotes those specific students.
    If request_data is empty or doesn't contain 'student_ids', promotes all students in the round.
    """
    student_ids = request_data.get('student_ids') if request_data and isinstance(request_data, dict) else None
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Get job
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Verify admin owns this job
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        if not admin_role or admin_role.college_id != job.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only manage rounds for jobs from your own college"
            )
    
    # Get all rounds for this job
    all_rounds = db.query(JobRound).filter(
        JobRound.job_id == job.id,
        JobRound.is_active == True
    ).order_by(JobRound.order).all()
    
    current_round_index = next((i for i, r in enumerate(all_rounds) if r.id == round_id), -1)
    if current_round_index < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Round not found in job's round list"
        )
    
    next_round = all_rounds[current_round_index + 1] if current_round_index + 1 < len(all_rounds) else None
    if not next_round:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No next round found. This is the final round."
        )
    
    # Get students in current round (any status - no qualification required)
    query = db.query(JobApplicationRound).filter(
        JobApplicationRound.round_id == round_id
    )
    
    # If specific student IDs provided, filter by them
    if student_ids and len(student_ids) > 0:
        # Get applications for these student IDs
        applications = db.query(JobApplication).filter(
            JobApplication.job_id == job.id,
            JobApplication.user_id.in_(student_ids)
        ).all()
        application_ids = [app.id for app in applications]
        if application_ids:
            query = query.filter(JobApplicationRound.job_application_id.in_(application_ids))
        else:
            # No applications found for these student IDs
            return {
                "message": "No students found for the selected student IDs",
                "promoted_count": 0,
                "already_promoted": 0
            }
    
    round_statuses = query.all()
    
    promoted_count = 0
    already_promoted = 0
    
    for round_status in round_statuses:
        # Check if already in next round
        next_round_status = db.query(JobApplicationRound).filter(
            JobApplicationRound.job_application_id == round_status.job_application_id,
            JobApplicationRound.round_id == next_round.id
        ).first()
        
        if not next_round_status:
            # Create entry in next round
            next_round_status = JobApplicationRound(
                job_application_id=round_status.job_application_id,
                round_id=next_round.id,
                status="PENDING",
                updated_by=current_user.id
            )
            db.add(next_round_status)
            promoted_count += 1
            
            # Update application current_round and status
            application = db.query(JobApplication).filter(
                JobApplication.id == round_status.job_application_id
            ).first()
            if application:
                application.current_round = next_round.name
                
                # If promoted to "Selected" round, mark as job cracked
                if next_round.name.lower() == "selected":
                    application.status = "Selected"
        else:
            already_promoted += 1
    
    db.commit()
    
    return {
        "message": f"Promoted {promoted_count} students to {next_round.name}",
        "promoted_count": promoted_count,
        "already_promoted": already_promoted,
        "next_round_id": next_round.id,
        "next_round_name": next_round.name
    }


@router.get("/rounds/{round_id}/template")
async def download_round_template(
    round_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Download Excel template for round results bulk upload"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    round_obj = db.query(JobRound).filter(JobRound.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Round not found"
        )
    
    # Get all students who should be in this round
    job = db.query(Job).filter(Job.id == round_obj.job_id).first()
    
    # Get students from previous round (if not first round)
    all_rounds = db.query(JobRound).filter(
        JobRound.job_id == job.id,
        JobRound.is_active == True
    ).order_by(JobRound.order).all()
    
    current_round_index = next((i for i, r in enumerate(all_rounds) if r.id == round_id), -1)
    
    student_ids = []
    if current_round_index == 0:
        # First round - get all applicants
        applications = db.query(JobApplication).filter(JobApplication.job_id == job.id).all()
        student_ids = [app.user_id for app in applications]
    else:
        # Get students who qualified in previous round
        prev_round = all_rounds[current_round_index - 1]
        prev_round_statuses = db.query(JobApplicationRound).filter(
            JobApplicationRound.round_id == prev_round.id,
            JobApplicationRound.status == "QUALIFIED"
        ).all()
        student_ids = [status.application.user_id for status in prev_round_statuses]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Round {round_obj.name}"
    
    # Headers - Roll Number is the primary identifier
    headers = ["Round ID", "Roll Number", "Student Name", "Email", "Round Name", "Status", "Remarks"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add student data
    for student_id in student_ids:
        profile = db.query(Profile).filter(Profile.user_id == student_id).first()
        ws.append([
            round_obj.id,  # Round ID - pre-filled, stable identifier
            profile.roll_number if profile else "",  # Roll Number - primary identifier
            profile.full_name if profile else "",
            profile.email if profile else "",
            round_obj.name,
            "PENDING",  # Default status
            ""  # Remarks
        ])
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ["ROUND RESULTS BULK UPLOAD TEMPLATE"],
        [""],
        ["ROUND INFORMATION:"],
        [f"• Round ID: {round_obj.id} (This is the stable identifier used for tracking)"],
        [f"• Round Name: {round_obj.name} (Informational - can change without affecting tracking)"],
        [f"• Round Order: {round_obj.order}"],
        [""],
        ["INSTRUCTIONS:"],
        ["1. Roll Number is the PRIMARY identifier - all tracking uses Roll Number"],
        ["2. Round ID is pre-filled and should NOT be modified"],
        ["3. Round Name is pre-filled for reference only (name changes don't affect tracking)"],
        ["4. Fill in the Status column with one of: QUALIFIED, REJECTED, or ABSENT"],
        ["5. Optionally add remarks in the Remarks column"],
        ["6. Do NOT modify Round ID, Roll Number, Student Name, Email, or Round Name columns"],
        ["7. Save the file and upload it using the Bulk Upload button"],
        [""],
        ["STATUS VALUES:"],
        ["• QUALIFIED - Student passed this round and moves to next round"],
        ["• REJECTED - Student failed this round and is removed from process"],
        ["• ABSENT - Student was absent and is removed from process"],
        [""],
        ["HOW TRACKING WORKS:"],
        ["• Student associations are tracked using Round ID (stable) and Roll Number (primary)"],
        ["• Round Name is only for display - changing it won't affect student tracking"],
        ["• The system fetches student data from backend using Roll Number"],
        ["• All students in this template are already associated with Round ID: " + str(round_obj.id)],
        [""],
        ["EXTERNAL APPLICATIONS:"],
        ["• Students who applied externally (using apply link) can be added to rounds"],
        ["• The system will automatically create application records when needed"],
        ["• This allows tracking progress even if students didn't apply through the system"],
        [""],
        ["IMPORTANT:"],
        ["• Only students who qualified in previous rounds will appear in this template"],
        ["• For Round 1, you can add all students who applied externally"],
        ["• Students marked REJECTED or ABSENT will not appear in future round templates"],
        ["• The system will automatically create entries for qualified students in the next round"],
        ["• Round ID ensures consistent tracking even if round names change"]
    ]
    
    for row in instructions:
        ws2.append(row)
    
    ws2.column_dimensions['A'].width = 80
    
    # Set column widths for better readability
    column_widths = {
        'A': 12,  # Round ID
        'B': 18,  # Roll Number
        'C': 25,  # Student Name
        'D': 30,  # Email
        'E': 20,  # Round Name
        'F': 15,  # Status
        'G': 40   # Remarks
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=round_{round_obj.id}_{round_obj.name}_template.xlsx"
        }
    )
