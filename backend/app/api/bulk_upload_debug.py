"""
Debug endpoint for bulk upload staff to diagnose Excel parsing issues
"""
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status
from sqlalchemy.orm import Session
from app.core.database import get_db
from openpyxl import load_workbook
import io
import csv

router = APIRouter()

@router.post("/debug/staff-excel")
async def debug_staff_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Debug endpoint to analyze Excel file structure without processing"""
    
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    
    debug_info = {
        "filename": file.filename,
        "file_size": len(contents),
        "file_type": file_ext,
        "sheets": [],
        "headers": [],
        "row_count": 0,
        "sample_rows": [],
        "parsed_rows": []
    }
    
    if file_ext == 'csv':
        csv_content = contents.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        headers = csv_reader.fieldnames or []
        debug_info["headers"] = headers
        
        rows = list(csv_reader)
        debug_info["row_count"] = len(rows)
        
        # Sample first 5 rows
        for i, row in enumerate(rows[:5]):
            debug_info["sample_rows"].append({
                "row_number": i + 2,
                "data": dict(row)
            })
        
        # Check parsed rows
        for i, row in enumerate(rows):
            row_dict = {}
            for k, v in row.items():
                if k:
                    key_normalized = str(k).lower().strip()
                    value_normalized = str(v).strip() if v is not None else ""
                    row_dict[key_normalized] = value_normalized
            
            has_data = any(str(v).strip() for v in row_dict.values() if v)
            debug_info["parsed_rows"].append({
                "row_number": i + 2,
                "has_data": has_data,
                "email": row_dict.get('email', ''),
                "role": row_dict.get('role', ''),
                "branch_id": row_dict.get('branch_id', ''),
                "all_keys": list(row_dict.keys()),
                "all_values": list(row_dict.values())
            })
    else:
        # Excel file
        excel_file = io.BytesIO(contents)
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        
        debug_info["sheets"] = workbook.sheetnames
        
        # Try to find the "Staff" sheet first
        sheet = None
        if "Staff" in workbook.sheetnames:
            sheet = workbook["Staff"]
            debug_info["active_sheet"] = "Staff"
        else:
            sheet = workbook.active
            debug_info["active_sheet"] = sheet.title
        
        # Get headers
        headers_raw = []
        for cell in sheet[1]:
            headers_raw.append(cell.value)
        
        # Normalize headers
        headers = []
        for header in headers_raw:
            if header is None:
                headers.append("")
            else:
                headers.append(str(header).strip().lower())
        
        debug_info["headers"] = headers
        debug_info["headers_raw"] = [str(h) if h else "" for h in headers_raw]
        
        # Parse rows
        row_count = 0
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for idx, header in enumerate(headers):
                if header:
                    if idx < len(row):
                        cell_value = row[idx]
                        if cell_value is None:
                            row_dict[header] = ""
                        else:
                            row_dict[header] = str(cell_value).strip()
                    else:
                        row_dict[header] = ""
            
            # Check if row has data
            has_data = any(str(v).strip() for v in row_dict.values() if v)
            
            if row_idx <= 5 or row_idx % 10 == 0 or row_idx >= 35:
                debug_info["sample_rows"].append({
                    "row_number": row_idx,
                    "raw_values": [str(v) if v is not None else None for v in row],
                    "parsed_dict": row_dict,
                    "has_data": has_data
                })
            
            debug_info["parsed_rows"].append({
                "row_number": row_idx,
                "has_data": has_data,
                "email": row_dict.get('email', ''),
                "role": row_dict.get('role', ''),
                "branch_id": row_dict.get('branch_id', ''),
                "all_keys": list(row_dict.keys()),
                "all_values": list(row_dict.values())[:3]  # First 3 values only
            })
            
            if has_data:
                row_count += 1
        
        debug_info["row_count"] = row_count
        debug_info["total_excel_rows"] = sheet.max_row - 1  # Exclude header
    
    return debug_info

