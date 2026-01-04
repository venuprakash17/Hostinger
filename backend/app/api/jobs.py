"""Job/Placement API endpoints"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Body
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import csv
import io
from openpyxl import load_workbook
from pydantic import BaseModel
from app.core.database import get_db
from app.config import get_settings
from app.models.job import Job, JobApplication
from app.models.user import User, UserRole, RoleEnum
from app.models.profile import Profile
from app.schemas.job import (
    JobCreate, JobUpdate, JobResponse,
    JobApplicationCreate, JobApplicationUpdate, JobApplicationResponse
)
from app.api.auth import get_current_user, get_optional_user

settings = get_settings()

def get_current_super_admin(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> User:
    """Verify user is super admin"""
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only super admins can perform this action"
        )
    
    return current_user

router = APIRouter(prefix="/jobs", tags=["jobs"])


def get_current_admin_or_faculty(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> User:
    """Verify user is admin, faculty, or super admin"""
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names and RoleEnum.ADMIN not in role_names and RoleEnum.FACULTY not in role_names:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admins, faculty, and super admins can perform this action"
        )
    
    return current_user


def get_current_admin(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> User:
    """Verify user is admin or super admin"""
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.ADMIN not in role_names and RoleEnum.SUPER_ADMIN not in role_names:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only college admins or super admins can perform this action"
        )
    
    return current_user


@router.post("/", response_model=JobResponse, status_code=status.HTTP_201_CREATED)
async def create_job(
    job_data: JobCreate,
    current_user: User = Depends(get_current_admin_or_faculty),
    db: Session = Depends(get_db)
):
    """Create a new job posting (Admin, Faculty, or Super Admin) - Jobs are global by default (available to all students)"""
    # Jobs are now global - college_id is optional (NULL = available to all students)
    job_dict = job_data.model_dump()
    college_id = job_dict.get('college_id')  # Optional - NULL means global
    
    # Get user roles to determine college_id and if super admin
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    # If not super admin and no college_id provided, get from user's role
    if RoleEnum.SUPER_ADMIN not in role_names and not college_id:
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role.in_([RoleEnum.ADMIN, RoleEnum.FACULTY])
        ).first()
        if admin_role and admin_role.college_id:
            college_id = admin_role.college_id
    
    # Validate eligibility data
    if job_data.eligibility_type == "branch" and not job_data.eligible_branches:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="eligible_branches is required when eligibility_type is 'branch'"
        )
    
    if job_data.eligibility_type == "specific_students" and not job_data.eligible_user_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="eligible_user_ids is required when eligibility_type is 'specific_students'"
        )
    
    # Create job
    job_dict = job_data.model_dump()
    if 'college_id' in job_dict:
        del job_dict['college_id']  # Remove from dict, set separately
    
    new_job = Job(
        **job_dict,
        college_id=college_id,
        created_by=current_user.id
    )
    
    db.add(new_job)
    db.flush()
    
    # Automatically create Round 0 (Step 0) for tracking all applicants
    # This is the default round where all students who applied are added
    from app.models.job_round import JobRound
    round_0 = JobRound(
        job_id=new_job.id,
        name="Applied",
        order=0,
        description="Default round for tracking all students who applied for this job",
        is_active=True
    )
    db.add(round_0)
    
    db.commit()
    db.refresh(new_job)
    
    return new_job


@router.get("/", response_model=List[JobResponse])
async def list_jobs(
    skip: int = 0,
    limit: int = 100,
    is_active: Optional[bool] = None,
    job_type: Optional[str] = None,
    search: Optional[str] = None,
    current_user: Optional[User] = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """List jobs - All jobs are global (available to all registered students)"""
    query = db.query(Job)
    
    # Filter by active status
    if is_active is not None:
        query = query.filter(Job.is_active == is_active)
    
    # Filter by job type
    if job_type:
        query = query.filter(Job.job_type == job_type)
    
    # Search filter (title, company, role)
    if search:
        search_term = f"%{search.lower()}%"
        query = query.filter(
            db.or_(
                Job.title.ilike(search_term),
                Job.company.ilike(search_term),
                Job.role.ilike(search_term)
            )
        )
    
    # IMPORTANT: All jobs are accessible to all students by default
    # Filter by eligibility criteria only (not by college_id)
    user_profile = None
    if current_user:
        try:
            user_profile = db.query(Profile).filter(Profile.user_id == current_user.id).first()
        except Exception as e:
            # If profile query fails, continue without profile
            user_profile = None
        
        # Check if user is admin or super admin (they see all jobs)
        try:
            user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
            role_names = [role.role for role in user_roles]
        except Exception as e:
            # If roles query fails, treat as regular user
            role_names = []
        
        if RoleEnum.SUPER_ADMIN in role_names:
            # Super Admin sees all jobs - no filtering needed
            jobs = query.order_by(Job.posted_date.desc()).offset(skip).limit(limit).all()
            return jobs
        elif RoleEnum.ADMIN in role_names:
            # College Admin: Only see jobs from their college (exclude Super Admin jobs with college_id = NULL)
            admin_role = db.query(UserRole).filter(
                UserRole.user_id == current_user.id,
                UserRole.role == RoleEnum.ADMIN
            ).first()
            if admin_role and admin_role.college_id:
                # Only show jobs where college_id matches admin's college
                query = query.filter(Job.college_id == admin_role.college_id)
            else:
                # Admin without college_id - return empty (shouldn't happen but safe fallback)
                return []
            jobs = query.order_by(Job.posted_date.desc()).offset(skip).limit(limit).all()
            return jobs
        else:
            # Students - filter by eligibility only (NOT by college_id - all jobs are global)
            # Get all jobs first, then filter by eligibility
            all_jobs = query.all()
            eligible_jobs = []
            
            for job in all_jobs:
                is_eligible = False
                
                # Default to "all_students" if eligibility_type is None or empty
                eligibility_type = job.eligibility_type or "all_students"
                
                # All students can see jobs with eligibility_type = "all_students" or None
                if eligibility_type == "all_students" or not job.eligibility_type:
                    is_eligible = True
                elif eligibility_type == "branch" and user_profile:
                    # Check if user's department matches eligible branches
                    if job.eligible_branches:
                        # Get department name and code from profile or department relationship
                        user_department_name = None
                        user_department_code = None
                        
                        if user_profile.department:
                            user_department_name = user_profile.department.strip()
                        
                        if user_profile.department_id:
                            # Get department name and code from department_id
                            from app.models.academic import Department
                            dept_obj = db.query(Department).filter(Department.id == user_profile.department_id).first()
                            if dept_obj:
                                if dept_obj.name:
                                    user_department_name = dept_obj.name.strip()
                                if dept_obj.code:
                                    user_department_code = dept_obj.code.strip()
                        
                        if user_department_name or user_department_code:
                            # Normalize for case-insensitive comparison
                            user_dept_name_normalized = user_department_name.upper().strip() if user_department_name else None
                            user_dept_code_normalized = user_department_code.upper().strip() if user_department_code else None
                            
                            # Check if any eligible branch matches (case-insensitive)
                            # eligible_branches can be a list or a single value
                            eligible_branches_list = job.eligible_branches if isinstance(job.eligible_branches, list) else [job.eligible_branches]
                            
                            # Debug logging (can be removed in production)
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.debug(f"Job {job.id}: Checking eligibility - user_dept_name={user_department_name}, user_dept_code={user_department_code}, eligible_branches={eligible_branches_list}")
                            
                            for branch in eligible_branches_list:
                                if branch is None:
                                    continue
                                branch_normalized = str(branch).upper().strip()
                                
                                # Match against both department name and code
                                if user_dept_name_normalized and user_dept_name_normalized == branch_normalized:
                                    is_eligible = True
                                    logger.debug(f"Job {job.id}: Matched by department name: {user_dept_name_normalized} == {branch_normalized}")
                                    break
                                if user_dept_code_normalized and user_dept_code_normalized == branch_normalized:
                                    is_eligible = True
                                    logger.debug(f"Job {job.id}: Matched by department code: {user_dept_code_normalized} == {branch_normalized}")
                                    break
                            
                            if not is_eligible:
                                logger.debug(f"Job {job.id}: Not eligible - no match found for user department")
                elif eligibility_type == "specific_students":
                    # Check if user ID is in eligible_user_ids
                    if job.eligible_user_ids and current_user.id in job.eligible_user_ids:
                        is_eligible = True
                
                # Additional filter: Check year eligibility if eligible_years is specified
                if is_eligible and job.eligible_years and user_profile:
                    # Job has year restrictions - check if student's year matches
                    user_year = user_profile.present_year
                    if user_year:
                        # Normalize year for comparison (handle "1st", "1", etc.)
                        from app.core.year_utils import parse_year, format_year
                        user_year_normalized = parse_year(user_year)  # Convert to numeric
                        user_year_formatted = format_year(user_year_normalized) if user_year_normalized else None
                        
                        # eligible_years can be a list or a single value
                        eligible_years_list = job.eligible_years if isinstance(job.eligible_years, list) else [job.eligible_years]
                        
                        year_match = False
                        for eligible_year in eligible_years_list:
                            if eligible_year is None:
                                continue
                            # Normalize eligible year
                            eligible_year_normalized = parse_year(str(eligible_year))
                            eligible_year_formatted = format_year(eligible_year_normalized) if eligible_year_normalized else None
                            
                            # Match if normalized years match OR formatted years match
                            if (user_year_normalized and eligible_year_normalized and 
                                user_year_normalized == eligible_year_normalized):
                                year_match = True
                                break
                            if (user_year_formatted and eligible_year_formatted and 
                                user_year_formatted.upper() == eligible_year_formatted.upper()):
                                year_match = True
                                break
                            # Also check direct string match (case-insensitive)
                            if str(user_year).upper().strip() == str(eligible_year).upper().strip():
                                year_match = True
                                break
                        
                        # If job has year restrictions, student must match
                        if not year_match:
                            is_eligible = False
                
                if is_eligible:
                    eligible_jobs.append(job)
            
            # Return only eligible jobs for students (sorted by posted_date)
            eligible_jobs.sort(key=lambda j: j.posted_date if j.posted_date else j.id, reverse=True)
            return eligible_jobs[skip:skip+limit]
    
    # For non-logged-in users, return all active jobs
    jobs = query.filter(Job.is_active == True).order_by(Job.posted_date.desc()).offset(skip).limit(limit).all()
    return jobs


@router.get("/template")
async def download_job_template(
    current_user: User = Depends(get_current_admin_or_faculty),
    db: Session = Depends(get_db)
):
    """Download Excel template for job bulk upload (Admin, Faculty, and Super Admin)"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs Template"
    
    # Headers with better organization - job_id is first for easy reference
    headers = [
        "job_id", "title", "company", "role", "description", "location", "ctc",
        "job_type", "eligibility_type", "eligible_branches", "eligible_years",
        "requirements", "rounds", "deadline", "apply_link", "is_active"
    ]
    ws.append(headers)
    
    # Style headers with better colors
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Example rows with comprehensive examples
    # Note: job_id is optional - leave empty to create new job, or provide ID to update existing job
    example_rows = [
        # Example 1: All students eligible (new job - no job_id)
        [
            "",  # job_id - empty means create new job
            "Software Engineer",  # title
            "Google",  # company
            "Software Engineer",  # role
            "Looking for experienced software engineer with 2-5 years of experience in web development. Must have strong problem-solving skills.",  # description
            "Hyderabad",  # location
            "₹15 LPA",  # ctc
            "On-Campus",  # job_type
            "all_students",  # eligibility_type
            "",  # eligible_branches - empty means all branches
            "",  # eligible_years - empty means all years
            "Bachelor's degree in Computer Science;2-5 years experience;Strong problem-solving skills",  # requirements
            "Online Test;Technical Interview;HR Round",  # rounds
            "2024-12-31",  # deadline
            "https://careers.google.com/jobs/apply",  # apply_link
            "true"  # is_active
        ],
        # Example 2: Branch-specific (update existing job - with job_id)
        [
            "1",  # job_id - if provided, will update existing job with this ID
            "Data Scientist",  # title
            "Microsoft",  # company
            "Data Scientist",  # role
            "Looking for data scientist with ML experience. Must have knowledge of Python, TensorFlow, and data analysis.",  # description
            "Bangalore",  # location
            "₹20 LPA",  # ctc
            "Off-Campus",  # job_type
            "branch",  # eligibility_type
            "CSE,ECE,IT",  # eligible_branches
            "3rd,4th",  # eligible_years
            "Master's degree in Data Science or related field;ML experience;Python and TensorFlow knowledge",  # requirements
            "Technical Round;Managerial Round;Final HR Round",  # rounds
            "2024-11-30",  # deadline
            "",  # apply_link - empty means no external link
            "true"  # is_active
        ],
        # Example 3: Branch-specific with year filter (new job)
        [
            "",  # job_id - empty means create new job
            "Product Manager",  # title
            "Amazon",  # company
            "Product Manager",  # role
            "Looking for product manager with 3+ years experience. Must have strong analytical and communication skills.",  # description
            "Mumbai",  # location
            "₹25 LPA",  # ctc
            "On-Campus",  # job_type
            "branch",  # eligibility_type
            "CSE,IT",  # eligible_branches - comma-separated branch names
            "4th",  # eligible_years - only 4th year students
            "Bachelor's degree;3+ years experience;Strong analytical skills;Excellent communication",  # requirements
            "Aptitude Test;Case Study Round;Product Round;HR Round",  # rounds
            "2024-10-15",  # deadline
            "",  # apply_link - empty means no external link
            "true"  # is_active
        ],
        # Example 4: Internship (new job)
        [
            "",  # job_id - empty means create new job
            "Software Development Intern",  # title
            "Flipkart",  # company
            "Software Development Intern",  # role
            "Summer internship for students interested in e-commerce and software development.",  # description
            "Bangalore",  # location
            "₹30,000/month",  # ctc
            "Internship",  # job_type
            "all_students",  # eligibility_type
            "",  # eligible_branches - empty means all branches
            "2nd,3rd",  # eligible_years
            "Currently pursuing Bachelor's or Master's degree;Basic programming knowledge",  # requirements
            "Online Test;Technical Interview",  # rounds
            "2024-09-30",  # deadline
            "",  # apply_link - empty means no external link
            "true"  # is_active
        ]
    ]
    
    for row in example_rows:
        ws.append(row)
    
    # Style example rows
    for row_idx in range(2, len(example_rows) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Set column widths for better readability - ensure headers are fully visible
    column_widths = {
        'A': 12,  # job_id
        'B': 25,  # title
        'C': 20,  # company
        'D': 20,  # role
        'E': 50,  # description
        'F': 15,  # location
        'G': 15,  # ctc
        'H': 15,  # job_type
        'I': 20,  # eligibility_type (increased from 18)
        'J': 25,  # eligible_branches
        'K': 20,  # eligible_years
        'L': 40,  # requirements
        'M': 35,  # rounds
        'N': 15,  # deadline
        'O': 35,  # apply_link (increased from 30)
        'P': 12   # is_active
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Add comprehensive instructions in a separate sheet
    ws2 = wb.create_sheet("Instructions", 0)  # Insert at beginning
    instructions = [
        ["JOB BULK UPLOAD TEMPLATE - INSTRUCTIONS"],
        [""],
        ["REQUIRED FIELDS (Must be filled):"],
        ["• title: Job title (e.g., 'Software Engineer', 'Data Scientist')"],
        ["• company: Company name (e.g., 'Google', 'Microsoft')"],
        ["• role: Job role/position (e.g., 'Software Engineer', 'Product Manager')"],
        [""],
        ["JOB ID FIELD (Optional but recommended):"],
        ["• job_id: Leave EMPTY to create a new job, or provide existing job ID to UPDATE that job"],
        ["  - If job_id is provided and exists, the job will be UPDATED with new data"],
        ["  - If job_id is provided but doesn't exist, a new job will be created with that ID"],
        ["  - If job_id is empty, a new job will be created with auto-generated ID"],
        ["  - Example: To update job #5, set job_id = '5' and provide all other fields"],
        [""],
        ["OPTIONAL FIELDS:"],
        ["• description: Detailed job description"],
        ["• location: Job location (e.g., 'Hyderabad', 'Bangalore', 'Remote')"],
        ["• ctc: Compensation package (e.g., '₹15 LPA', '₹20-25 LPA', '₹30,000/month')"],
        ["• job_type: Must be one of: 'On-Campus', 'Off-Campus', or 'Internship'"],
        ["• eligibility_type: Must be one of: 'all_students', 'branch', or 'specific_students'"],
        ["• eligible_branches: Comma-separated branch names (e.g., 'CSE,ECE,IT')"],
        ["  - REQUIRED if eligibility_type is 'branch'"],
        ["  - Leave empty if eligibility_type is 'all_students'"],
        ["• eligible_years: Comma-separated years (e.g., '1st,2nd,3rd' or '1,2,3')"],
        ["  - Optional: Filter by student year (works with any eligibility_type)"],
        ["  - Leave empty to allow all years"],
        ["• requirements: Semicolon-separated requirements (e.g., 'Bachelor's degree;2-5 years')"],
        ["• rounds: Semicolon-separated selection rounds (e.g., 'Online Test;Technical Interview;HR Round')"],
        ["• deadline: Application deadline in YYYY-MM-DD format (e.g., '2024-12-31')"],
        ["• is_active: 'true' or 'false' (default: 'true')"],
        [""],
        ["ELIGIBILITY TYPES EXPLAINED:"],
        ["1. all_students: Job is visible to ALL registered students (can be filtered by eligible_years)"],
        ["2. branch: Job is visible ONLY to students from specified branches"],
        ["   - Example: If eligible_branches = 'CSE,ECE', only CSE and ECE students can see this job"],
        ["   - Can be combined with eligible_years for additional filtering"],
        [""],
        ["IMPORTANT NOTES:"],
        ["• Jobs uploaded here are GLOBAL - available to ALL registered students (not college-specific)"],
        ["• Use eligibility_type and eligible_branches to restrict visibility to specific branches"],
        ["• All dates should be in YYYY-MM-DD format"],
        ["• Multiple values (branches, requirements, rounds) should be separated by commas or semicolons as specified"],
        ["• Do NOT delete the header row"],
        ["• Empty rows will be skipped"],
        [""],
        ["EXAMPLES:"],
        ["• For a job open to all students: eligibility_type='all_students', eligible_branches='', eligible_years=''"],
        ["• For a job only for CSE and ECE: eligibility_type='branch', eligible_branches='CSE,ECE', eligible_years=''"],
        ["• For a job only for 3rd and 4th year students: eligible_years='3rd,4th' (works with any eligibility_type)"],
        ["• For a job for CSE/ECE students in 3rd/4th year: eligibility_type='branch', eligible_branches='CSE,ECE', eligible_years='3rd,4th'"],
        [""],
        ["TIPS:"],
        ["• Copy the example rows and modify them for your jobs"],
        ["• Make sure to match the exact format for eligibility_type values"],
        ["• Branch names must match exactly with student department names in the system"],
    ]
    
    for row in instructions:
        ws2.append(row)
    
    # Style instructions sheet
    title_cell = ws2['A1']
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Set column width for instructions
    ws2.column_dimensions['A'].width = 100
    
    # Add data validation sheet with common values
    ws3 = wb.create_sheet("Reference Data")
    reference_data = [
        ["Common Job Types"],
        ["On-Campus"],
        ["Off-Campus"],
        ["Internship"],
        [""],
        ["Eligibility Types"],
        ["all_students"],
        ["branch"],
        [""],
        ["Common Branches (Examples)"],
        ["CSE"],
        ["ECE"],
        ["IT"],
        ["EEE"],
        ["ME"],
        ["CE"],
        ["AIML"],
        ["DS"],
        [""],
        ["Date Format"],
        ["YYYY-MM-DD"],
        ["Example: 2024-12-31"],
        [""],
        ["Boolean Values"],
        ["true"],
        ["false"],
    ]
    
    for row in reference_data:
        ws3.append(row)
    
    # Style reference data
    ws3['A1'].font = Font(bold=True, size=12, color="1F4E78")
    ws3['A6'].font = Font(bold=True, size=12, color="1F4E78")
    ws3['A11'].font = Font(bold=True, size=12, color="1F4E78")
    ws3['A20'].font = Font(bold=True, size=12, color="1F4E78")
    ws3['A24'].font = Font(bold=True, size=12, color="1F4E78")
    ws3.column_dimensions['A'].width = 25
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=job_upload_template.xlsx",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
    )


@router.get("/{job_id}", response_model=JobResponse)
async def get_job(
    job_id: int,
    current_user: Optional[User] = Depends(get_optional_user),
    db: Session = Depends(get_db)
):
    """Get a specific job by ID"""
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Check eligibility if user is logged in as student
    if current_user:
        user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
        role_names = [role.role for role in user_roles]
        
        if RoleEnum.ADMIN not in role_names:
            # Student - check eligibility
            user_profile = db.query(Profile).filter(Profile.user_id == current_user.id).first()
            is_eligible = False
            
            # Default to "all_students" if eligibility_type is None or empty
            eligibility_type = job.eligibility_type or "all_students"
            
            if eligibility_type == "all_students":
                is_eligible = True
            elif eligibility_type == "branch" and user_profile:
                # Check if user's department matches eligible branches
                if job.eligible_branches:
                    # Get department name and code from profile or department relationship
                    user_department_name = None
                    user_department_code = None
                    
                    if user_profile.department:
                        user_department_name = user_profile.department.strip()
                    
                    if user_profile.department_id:
                        # Get department name and code from department_id
                        from app.models.academic import Department
                        dept_obj = db.query(Department).filter(Department.id == user_profile.department_id).first()
                        if dept_obj:
                            if dept_obj.name:
                                user_department_name = dept_obj.name.strip()
                            if dept_obj.code:
                                user_department_code = dept_obj.code.strip()
                    
                    if user_department_name or user_department_code:
                        # Normalize for case-insensitive comparison
                        user_dept_name_normalized = user_department_name.upper().strip() if user_department_name else None
                        user_dept_code_normalized = user_department_code.upper().strip() if user_department_code else None
                        
                        # Check if any eligible branch matches (case-insensitive)
                        # eligible_branches can be a list or a single value
                        eligible_branches_list = job.eligible_branches if isinstance(job.eligible_branches, list) else [job.eligible_branches]
                        
                        # Debug logging (can be removed in production)
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.debug(f"Job {job.id}: Checking eligibility - user_dept_name={user_department_name}, user_dept_code={user_department_code}, eligible_branches={eligible_branches_list}")
                        
                        for branch in eligible_branches_list:
                            if branch is None:
                                continue
                            branch_normalized = str(branch).upper().strip()
                            
                            # Match against both department name and code
                            if user_dept_name_normalized and user_dept_name_normalized == branch_normalized:
                                is_eligible = True
                                logger.debug(f"Job {job.id}: Matched by department name: {user_dept_name_normalized} == {branch_normalized}")
                                break
                            if user_dept_code_normalized and user_dept_code_normalized == branch_normalized:
                                is_eligible = True
                                logger.debug(f"Job {job.id}: Matched by department code: {user_dept_code_normalized} == {branch_normalized}")
                                break
                        
                        if not is_eligible:
                            logger.debug(f"Job {job.id}: Not eligible - no match found for user department")
            elif eligibility_type == "specific_students":
                if job.eligible_user_ids and current_user.id in job.eligible_user_ids:
                    is_eligible = True
            
            # Check year eligibility if eligible_years is specified (applies to all eligibility types)
            if is_eligible and job.eligible_years and user_profile:
                user_year = user_profile.present_year
                if user_year:
                    from app.core.year_utils import parse_year, format_year
                    user_year_normalized = parse_year(user_year)
                    user_year_formatted = format_year(user_year_normalized) if user_year_normalized else None
                    
                    eligible_years_list = job.eligible_years if isinstance(job.eligible_years, list) else [job.eligible_years]
                    
                    year_match = False
                    for eligible_year in eligible_years_list:
                        if eligible_year is None:
                            continue
                        eligible_year_normalized = parse_year(str(eligible_year))
                        eligible_year_formatted = format_year(eligible_year_normalized) if eligible_year_normalized else None
                        
                        if (user_year_normalized and eligible_year_normalized and 
                            user_year_normalized == eligible_year_normalized):
                            year_match = True
                            break
                        if (user_year_formatted and eligible_year_formatted and 
                            user_year_formatted.upper() == eligible_year_formatted.upper()):
                            year_match = True
                            break
                        if str(user_year).upper().strip() == str(eligible_year).upper().strip():
                            year_match = True
                            break
                    
                    if not year_match:
                        raise HTTPException(
                            status_code=status.HTTP_403_FORBIDDEN,
                            detail="You are not eligible for this job based on your year"
                        )
            
            if not is_eligible:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You are not eligible for this job"
                )
    
    return job


@router.put("/{job_id}", response_model=JobResponse)
async def update_job(
    job_id: int,
    job_data: JobUpdate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update a job posting (Admin or Super Admin)"""
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Jobs are global - any admin can update any job
    # Update job fields
    update_data = job_data.model_dump(exclude_unset=True)
    
    for field, value in update_data.items():
        setattr(job, field, value)
    
    db.commit()
    db.refresh(job)
    
    return job


@router.delete("/{job_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_job(
    job_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Delete a job posting (Admin or Super Admin)"""
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Jobs are global - any admin can delete any job
    
    db.delete(job)
    db.commit()
    
    return None


@router.post("/bulk-delete", response_model=dict)
async def bulk_delete_jobs(
    job_ids: List[int] = Body(...),
    current_user: User = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Bulk delete jobs (Super Admin only)"""
    if not job_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No job IDs provided"
        )
    
    deleted_count = 0
    not_found_ids = []
    
    for job_id in job_ids:
        job = db.query(Job).filter(Job.id == job_id).first()
        if job:
            db.delete(job)
            deleted_count += 1
        else:
            not_found_ids.append(job_id)
    
    db.commit()
    
    return {
        "message": f"Deleted {deleted_count} job(s)",
        "deleted_count": deleted_count,
        "not_found_ids": not_found_ids
    }


class BulkUpdateStatusRequest(BaseModel):
    job_ids: List[int]
    is_active: bool

@router.post("/bulk-update-status", response_model=dict)
async def bulk_update_job_status(
    request: BulkUpdateStatusRequest,
    current_user: User = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Bulk update job status (active/inactive) (Super Admin only)"""
    if not request.job_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No job IDs provided"
        )
    
    updated_count = 0
    not_found_ids = []
    
    for job_id in request.job_ids:
        job = db.query(Job).filter(Job.id == job_id).first()
        if job:
            job.is_active = request.is_active
            updated_count += 1
        else:
            not_found_ids.append(job_id)
    
    db.commit()
    
    status_text = "activated" if request.is_active else "deactivated"
    return {
        "message": f"{updated_count} job(s) {status_text}",
        "updated_count": updated_count,
        "not_found_ids": not_found_ids
    }


@router.post("/{job_id}/apply", response_model=JobApplicationResponse, status_code=status.HTTP_201_CREATED)
async def apply_for_job(
    job_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Apply for a job (Student only)"""
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Check if already applied
    existing_application = db.query(JobApplication).filter(
        JobApplication.job_id == job_id,
        JobApplication.user_id == current_user.id
    ).first()
    
    if existing_application:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You have already applied for this job"
        )
    
    # Create application
    new_application = JobApplication(
        job_id=job_id,
        user_id=current_user.id,
        status="Applied"
    )
    
    db.add(new_application)
    db.commit()
    db.refresh(new_application)
    
    return new_application


@router.get("/{job_id}/applications", response_model=List[JobApplicationResponse])
async def get_job_applications(
    job_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Get all applications for a job (Admin or Super Admin)"""
    job = db.query(Job).filter(Job.id == job_id).first()
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Check if user is super admin - they can view all applications
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        # For regular admins, verify they own this job's college
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or admin_role.college_id != job.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only view applications for jobs from your own college"
            )
    
    applications = db.query(JobApplication).filter(JobApplication.job_id == job_id).all()
    return applications


@router.put("/applications/{application_id}", response_model=JobApplicationResponse)
async def update_application(
    application_id: int,
    application_data: JobApplicationUpdate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update an application status (Admin or Super Admin)"""
    application = db.query(JobApplication).filter(JobApplication.id == application_id).first()
    
    if not application:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Application not found"
        )
    
    # Verify admin owns this application's job college
    job = db.query(Job).filter(Job.id == application.job_id).first()
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job not found"
        )
    
    # Check if user is super admin - they can update all applications
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        # For regular admins, verify they own this job's college
        admin_role = db.query(UserRole).filter(
            UserRole.user_id == current_user.id,
            UserRole.role == RoleEnum.ADMIN
        ).first()
        
        if not admin_role or admin_role.college_id != job.college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only update applications for jobs from your own college"
            )
    
    # Update application fields
    update_data = application_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(application, field, value)
    
    db.commit()
    db.refresh(application)
    
    return application


@router.get("/applications/my", response_model=List[JobApplicationResponse])
async def get_my_applications(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all applications for the current user (Student)"""
    applications = db.query(JobApplication).filter(
        JobApplication.user_id == current_user.id
    ).order_by(JobApplication.applied_at.desc()).all()
    
    return applications


@router.post("/bulk-upload", response_model=dict)
async def bulk_upload_jobs(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_admin_or_faculty),
    db: Session = Depends(get_db)
):
    """Bulk upload jobs from Excel or CSV file (Admin, Faculty, and Super Admin)
    
    Supported formats: .xlsx, .xls, .csv
    
    Excel/CSV Format (headers required):
    title,company,role,description,location,ctc,job_type,eligibility_type,eligible_branches,eligible_years,requirements,rounds,deadline,apply_link,is_active
    
    Required fields: title, company, role
    Optional fields: description, location, ctc, job_type, eligibility_type, eligible_branches (comma-separated), eligible_years (comma-separated, e.g., "1st,2nd,3rd"), requirements (semicolon-separated), rounds (semicolon-separated), deadline, apply_link (external application URL), is_active
    
    Eligibility types: "all_students", "branch"
    For "branch" type, provide eligible_branches as comma-separated (e.g., "CSE,ECE,IT")
    For year filtering, provide eligible_years as comma-separated (e.g., "1st,2nd,3rd" or "1,2,3") - works independently of eligibility_type
    
    Example:
    title,company,role,description,location,ctc,job_type,eligibility_type,eligible_branches,eligible_years,requirements,rounds,deadline,apply_link,is_active
    Software Engineer,Google,Software Engineer,Looking for experienced developer,Hyderabad,₹15 LPA,On-Campus,branch,"CSE,ECE,IT","3rd,4th","Bachelor's degree;2-5 years experience","Online Test;Technical Interview;HR Round",2024-12-31,https://careers.google.com/jobs/apply,true
    """
    try:
        print(f"[Bulk Upload Jobs] Starting upload for user {current_user.id}")
        
        if not file.filename:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No file provided"
            )
        
        file_ext = file.filename.lower().split('.')[-1]
        print(f"[Bulk Upload Jobs] File extension: {file_ext}")
        
        if file_ext not in ['csv', 'xlsx', 'xls']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
            )
        
        print(f"[Bulk Upload Jobs] Reading file contents...")
        contents = await file.read()
        print(f"[Bulk Upload Jobs] File size: {len(contents)} bytes")
        data = []
        
        # Parse based on file type
        if file_ext == 'csv':
            csv_content = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_content))
            data = list(reader)
            print(f"[Bulk Upload Jobs] CSV: Found {len(data)} rows")
            if data:
                print(f"[Bulk Upload Jobs] CSV headers: {list(data[0].keys())}")
                print(f"[Bulk Upload Jobs] CSV first row sample: {dict(list(data[0].items())[:3])}")
        else:
            # Excel file
            wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            
            # Find the correct sheet - skip "Instructions" and "Reference Data" sheets
            # Look for a sheet that has job data headers (title, company, role)
            target_sheet = None
            sheet_names = wb.sheetnames
            print(f"[Bulk Upload Jobs] Excel sheets found: {sheet_names}")
            
            # Expected data sheet names (case-insensitive)
            data_sheet_names = ['jobs template', 'jobs', 'job template', 'template', 'data', 'sheet1']
            
            # First, try to find a sheet with a known data sheet name
            for sheet_name in sheet_names:
                if sheet_name.lower() in data_sheet_names:
                    target_sheet = wb[sheet_name]
                    print(f"[Bulk Upload Jobs] Found data sheet by name: {sheet_name}")
                    break
            
            # If not found by name, check each sheet for job data headers
            if not target_sheet:
                for sheet_name in sheet_names:
                    # Skip instruction/reference sheets
                    if 'instruction' in sheet_name.lower() or 'reference' in sheet_name.lower():
                        continue
                    
                    ws = wb[sheet_name]
                    # Check first row for job headers
                    first_row = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
                    # Look for key job fields
                    if any(key in first_row for key in ['title', 'company', 'role']):
                        target_sheet = ws
                        print(f"[Bulk Upload Jobs] Found data sheet by headers: {sheet_name}")
                        break
            
            # If still not found, use the first non-instruction sheet
            if not target_sheet:
                for sheet_name in sheet_names:
                    if 'instruction' not in sheet_name.lower() and 'reference' not in sheet_name.lower():
                        target_sheet = wb[sheet_name]
                        print(f"[Bulk Upload Jobs] Using first non-instruction sheet: {sheet_name}")
                        break
            
            # Fallback to active sheet if nothing found
            if not target_sheet:
                target_sheet = wb.active
                print(f"[Bulk Upload Jobs] Using active sheet: {target_sheet.title}")
            
            ws = target_sheet
            print(f"[Bulk Upload Jobs] Excel: Using sheet '{ws.title}', Max row = {ws.max_row}, Max col = {ws.max_column}")
            
            # Get headers from first row - handle None values
            headers = []
            for cell in ws[1]:
                if cell.value is not None:
                    header = str(cell.value).strip().lower()  # Normalize to lowercase
                    headers.append(header)
                else:
                    headers.append("")
            
            print(f"[Bulk Upload Jobs] Excel headers (normalized): {headers}")
            
            # Validate that we have the expected headers
            expected_headers = ['title', 'company', 'role']
            has_required_headers = any(h in headers for h in expected_headers)
            if not has_required_headers:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Excel file does not contain required headers (title, company, role). Found headers: {headers}. Please ensure you're using the correct sheet with job data."
                )
            
            # Process data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip completely empty rows
                    continue
                
                job_dict = {}
                for idx, header in enumerate(headers):
                    if header and idx < len(row):
                        value = row[idx]
                        if value is not None:
                            job_dict[header] = str(value).strip()
                        else:
                            job_dict[header] = ""
                
                # Only add if dict has at least some data
                if job_dict:
                    data.append(job_dict)
                    if len(data) == 1:
                        print(f"[Bulk Upload Jobs] First data row sample: {dict(list(job_dict.items())[:3])}")
            
            print(f"[Bulk Upload Jobs] Excel: Parsed {len(data)} data rows")
        
        # Determine college_id based on user role
        user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
        role_names = [role.role for role in user_roles]
        
        # Super admin can create global jobs (college_id = None)
        # Admin and Faculty create jobs for their college
        college_id = None
        if RoleEnum.SUPER_ADMIN not in role_names:
            # For admin/faculty, get their college_id
            admin_role = db.query(UserRole).filter(
                UserRole.user_id == current_user.id,
                UserRole.role.in_([RoleEnum.ADMIN, RoleEnum.FACULTY])
            ).first()
            if admin_role and admin_role.college_id:
                college_id = admin_role.college_id
        
        results = {
            "success": [],
            "failed": [],
            "total": 0
        }
        
        for idx, job_data in enumerate(data, start=1):
            try:
                # Normalize keys to lowercase for case-insensitive matching
                normalized_data = {k.lower().strip(): v for k, v in job_data.items()}
                
                # Check for job_id - if provided, we'll update existing job
                job_id_str = normalized_data.get('job_id', '').strip() if normalized_data.get('job_id') else ''
                job_id = None
                existing_job = None
                
                if job_id_str:
                    try:
                        job_id = int(job_id_str)
                        # Check if job exists
                        existing_job = db.query(Job).filter(Job.id == job_id).first()
                        if existing_job:
                            print(f"[Bulk Upload Jobs] Row {idx}: Found existing job with ID {job_id}, will update")
                    except ValueError:
                        # Invalid job_id format, treat as new job
                        print(f"[Bulk Upload Jobs] Row {idx}: Invalid job_id '{job_id_str}', will create new job")
                        job_id = None
                
                # Validate required fields (case-insensitive)
                title = normalized_data.get('title', '').strip() if normalized_data.get('title') else ''
                company = normalized_data.get('company', '').strip() if normalized_data.get('company') else ''
                role = normalized_data.get('role', '').strip() if normalized_data.get('role') else ''
                
                if not title or not company or not role:
                    available_keys = list(normalized_data.keys())
                    raise ValueError(f"Missing required fields: title, company, role. Available keys: {available_keys}")
                
                # Use normalized data for the rest of processing
                job_data = normalized_data
            
                # Parse eligibility
                eligibility_type = job_data.get('eligibility_type', 'all_students').strip().lower()
                if eligibility_type not in ['all_students', 'branch']:
                    eligibility_type = 'all_students'
                
                eligible_branches = None
                eligible_user_ids = None  # Not used anymore, but keep for backward compatibility
                eligible_years = None
                
                if eligibility_type == 'branch':
                    branches_str = job_data.get('eligible_branches', '').strip()
                    if branches_str:
                        # Remove quotes if present and split by comma
                        branches_str = branches_str.strip('"\'')
                        eligible_branches = [b.strip() for b in branches_str.split(',') if b.strip()]
                        if not eligible_branches:
                            raise ValueError("eligible_branches is required when eligibility_type is 'branch'")
                
                # Note: eligible_user_ids feature removed - not needed for College Admin jobs
                # Only branch-based and year-based filtering is supported
                # If eligible_user_ids is provided in upload, it will be ignored
                
                # Parse eligible_years (comma-separated, e.g., "1st,2nd,3rd" or "1,2,3")
                years_str = job_data.get('eligible_years', '').strip()
                if years_str:
                    years_str = years_str.strip('"\'')
                    # Normalize years to formatted format (1st, 2nd, etc.)
                    from app.core.year_utils import format_year
                    years_list = [y.strip() for y in years_str.split(',') if y.strip()]
                    eligible_years = [format_year(y) or y for y in years_list]  # Format to "1st", "2nd", etc.
                
                # Parse requirements (semicolon-separated)
                requirements = None
                if job_data.get('requirements'):
                    req_str = job_data.get('requirements').strip().strip('"\'')
                    requirements = [r.strip() for r in req_str.split(';') if r.strip()]
                
                # Parse rounds (semicolon-separated)
                rounds = None
                if job_data.get('rounds'):
                    rounds_str = job_data.get('rounds').strip().strip('"\'')
                    rounds = [r.strip() for r in rounds_str.split(';') if r.strip()]
                
                # Parse deadline
                deadline = None
                if job_data.get('deadline'):
                    try:
                        deadline_str = str(job_data.get('deadline')).strip()
                        # Try parsing various date formats
                        for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%m/%d/%Y']:
                            try:
                                deadline = datetime.strptime(deadline_str, fmt)
                                break
                            except:
                                continue
                    except:
                        pass
                
                # Parse is_active
                is_active = True
                if job_data.get('is_active'):
                    is_active_str = str(job_data.get('is_active')).strip().lower()
                    is_active = is_active_str in ['true', '1', 'yes', 'y']
                
                # Update existing job or create new one
                if existing_job:
                    # Update existing job
                    existing_job.title = job_data['title']
                    existing_job.company = job_data['company']
                    existing_job.role = job_data['role']
                    existing_job.description = job_data.get('description') or None
                    existing_job.location = job_data.get('location') or None
                    existing_job.ctc = job_data.get('ctc') or None
                    existing_job.job_type = job_data.get('job_type', 'On-Campus')
                    existing_job.eligibility_type = eligibility_type
                    existing_job.apply_link = job_data.get('apply_link') or None
                    existing_job.eligible_branches = eligible_branches
                    existing_job.eligible_user_ids = None  # Not used for College Admin jobs
                    existing_job.eligible_years = eligible_years
                    existing_job.requirements = requirements
                    existing_job.rounds = rounds
                    existing_job.deadline = deadline
                    existing_job.is_active = is_active
                    # Don't change college_id or created_by for updates
                    new_job = existing_job
                    action = "updated"
                else:
                    # Create new job with appropriate college_id
                    new_job = Job(
                        title=job_data['title'],
                        company=job_data['company'],
                        role=job_data['role'],
                        description=job_data.get('description') or None,
                        location=job_data.get('location') or None,
                        ctc=job_data.get('ctc') or None,
                        job_type=job_data.get('job_type', 'On-Campus'),
                        eligibility_type=eligibility_type,
                        eligible_branches=eligible_branches,
                        eligible_user_ids=None,  # Not used for College Admin jobs
                        eligible_years=eligible_years,
                        requirements=requirements,
                        rounds=rounds,
                        deadline=deadline,
                        apply_link=job_data.get('apply_link') or None,
                        is_active=is_active,
                        college_id=college_id,  # Set based on user role (None for super admin, college_id for admin/faculty)
                        created_by=current_user.id
                    )
                    db.add(new_job)
                    db.flush()  # Flush to get new_job.id
                    
                    # Automatically create Round 0 (Step 0) for tracking all applicants
                    from app.models.job_round import JobRound
                    round_0 = JobRound(
                        job_id=new_job.id,
                        name="Applied",
                        order=0,
                        description="Default round for tracking all students who applied for this job",
                        is_active=True
                    )
                    db.add(round_0)
                    
                    action = "created"
                
                # If super admin created the job, also add it to job aggregation table
                
                results["success"].append({
                    "index": idx,
                    "title": job_data.get('title', 'N/A'),
                    "id": new_job.id,
                    "action": action
                })
                results["total"] += 1
                
            except Exception as e:
                # Try to get title from normalized_data if available, otherwise from original job_data
                title = 'N/A'
                if 'normalized_data' in locals():
                    title = normalized_data.get('title', 'N/A')
                elif isinstance(job_data, dict):
                    # Try both original and lowercase keys
                    title = job_data.get('title') or job_data.get('Title') or job_data.get('TITLE', 'N/A')
                
                results["failed"].append({
                    "index": idx,
                    "title": title if title else 'N/A',
                    "error": str(e)
                })
        
        # Commit all successful jobs at once
        if results["success"]:
            try:
                db.commit()
            except Exception as commit_error:
                db.rollback()
                import traceback
                print(f"Commit error: {commit_error}")
                print(traceback.format_exc())
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail=f"Failed to commit jobs: {str(commit_error)}"
                )
        
        return {
            "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
            "success_count": len(results["success"]),
            "failed_count": len(results["failed"]),
            "success": results["success"],
            "failed": results["failed"]
        }
    except HTTPException:
        # Re-raise HTTP exceptions as-is (they already have proper status codes)
        db.rollback()
        raise
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Bulk upload error: {e}")
        print(error_trace)
        db.rollback()
        
        # Provide more detailed error message in debug mode
        error_detail = str(e)
        if settings.DEBUG:
            error_detail = f"{str(e)}\n\nTraceback:\n{error_trace}"
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Bulk upload failed: {error_detail}"
        )

