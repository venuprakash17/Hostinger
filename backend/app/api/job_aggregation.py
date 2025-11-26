"""Job Aggregation API endpoints"""
from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from app.core.database import get_db
from app.models.job_aggregation import JobAggregation
from app.models.job import Job
from app.models.user import User, UserRole, RoleEnum
from app.schemas.job_aggregation import (
    JobAggregationCreate, JobAggregationUpdate, JobAggregationResponse,
    JobAggregationSyncRequest
)
from app.api.auth import get_current_user, get_current_admin
from app.models.user import UserRole, RoleEnum
from datetime import datetime
import httpx
import asyncio
import csv
import io
import json
from openpyxl import load_workbook

router = APIRouter(prefix="/job-aggregation", tags=["job-aggregation"])


async def fetch_linkedin_jobs(keywords: List[str], location: Optional[str] = None, max_results: int = 50) -> List[dict]:
    """Fetch jobs from LinkedIn (mock implementation - replace with actual API)"""
    # Note: LinkedIn requires API access. This is a mock implementation.
    # In production, use LinkedIn Jobs API or web scraping with proper permissions.
    
    # Mock data for demonstration
    mock_jobs = [
        {
            "external_id": f"linkedin_{i}",
            "source": "linkedin",
            "title": f"Software Engineer - {keywords[0] if keywords else 'Tech'}",
            "company": f"Tech Company {i}",
            "role": "Software Engineer",
            "description": f"Looking for a skilled software engineer with experience in {', '.join(keywords[:3])}",
            "location": location or "Remote",
            "ctc": f"₹{(8 + i * 2)} LPA",
            "job_type": "Full-time",
            "source_url": f"https://linkedin.com/jobs/view/{i}",
            "posted_date": datetime.utcnow(),
        }
        for i in range(min(max_results, 10))
    ]
    
    return mock_jobs


async def fetch_indeed_jobs(keywords: List[str], location: Optional[str] = None, max_results: int = 50) -> List[dict]:
    """Fetch jobs from Indeed (mock implementation - replace with actual API)"""
    # Note: Indeed requires API access. This is a mock implementation.
    # In production, use Indeed API or web scraping with proper permissions.
    
    # Mock data for demonstration
    mock_jobs = [
        {
            "external_id": f"indeed_{i}",
            "source": "indeed",
            "title": f"Developer - {keywords[0] if keywords else 'Software'}",
            "company": f"Company {i}",
            "role": "Developer",
            "description": f"Seeking developer proficient in {', '.join(keywords[:3])}",
            "location": location or "Hyderabad",
            "ctc": f"₹{(6 + i * 1.5)} LPA",
            "job_type": "Full-time",
            "source_url": f"https://indeed.com/viewjob?jk={i}",
            "posted_date": datetime.utcnow(),
        }
        for i in range(min(max_results, 10))
    ]
    
    return mock_jobs


async def sync_jobs_from_source(source: str, keywords: List[str], location: Optional[str], 
                                max_results: int, college_id: Optional[int], db: Session):
    """Sync jobs from a specific source"""
    try:
        if source == "linkedin":
            jobs_data = await fetch_linkedin_jobs(keywords, location, max_results)
        elif source == "indeed":
            jobs_data = await fetch_indeed_jobs(keywords, location, max_results)
        else:
            return  # Unsupported source
        
        synced_count = 0
        for job_data in jobs_data:
            # Check if job already exists
            existing = db.query(JobAggregation).filter(
                JobAggregation.source == source,
                JobAggregation.external_id == job_data.get("external_id")
            ).first()
            
            if not existing:
                # Create new aggregation
                new_job = JobAggregation(
                    source=source,
                    external_id=job_data.get("external_id"),
                    source_url=job_data.get("source_url"),
                    title=job_data.get("title"),
                    company=job_data.get("company"),
                    role=job_data.get("role"),
                    description=job_data.get("description"),
                    location=job_data.get("location"),
                    ctc=job_data.get("ctc"),
                    job_type=job_data.get("job_type"),
                    posted_date=job_data.get("posted_date"),
                    college_id=college_id,
                    last_synced_at=datetime.utcnow()
                )
                db.add(new_job)
                synced_count += 1
            else:
                # Update existing
                existing.last_synced_at = datetime.utcnow()
                if job_data.get("title"):
                    existing.title = job_data["title"]
                if job_data.get("description"):
                    existing.description = job_data["description"]
        
        db.commit()
        return synced_count
    except Exception as e:
        db.rollback()
        raise e


@router.post("/sync", response_model=dict)
async def sync_jobs(
    request: JobAggregationSyncRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Sync jobs from external sources (Admin only)"""
    # Get admin's college
    admin_role = db.query(UserRole).filter(
        UserRole.user_id == current_user.id,
        UserRole.role == RoleEnum.ADMIN
    ).first()
    
    college_id = request.college_id or (admin_role.college_id if admin_role else None)
    
    # Sync from each source
    results = {}
    total_synced = 0
    
    for source in request.sources:
        try:
            count = await sync_jobs_from_source(
                source,
                request.keywords or ["software engineer", "developer"],
                request.location,
                request.max_results,
                college_id,
                db
            )
            results[source] = count
            total_synced += count
        except Exception as e:
            results[source] = f"Error: {str(e)}"
    
    return {
        "message": f"Synced {total_synced} jobs from {len(request.sources)} sources",
        "results": results,
        "total_synced": total_synced
    }


@router.get("/", response_model=List[JobAggregationResponse])
async def list_aggregated_jobs(
    source: Optional[str] = None,
    is_imported: Optional[bool] = None,
    college_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List aggregated jobs"""
    query = db.query(JobAggregation)
    
    if source:
        query = query.filter(JobAggregation.source == source)
    
    if is_imported is not None:
        query = query.filter(JobAggregation.is_imported == is_imported)
    
    if college_id:
        query = query.filter(JobAggregation.college_id == college_id)
    else:
        # Show only global jobs or user's college jobs
        user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
        role_names = [role.role for role in user_roles]
        
        if RoleEnum.ADMIN in role_names or RoleEnum.SUPER_ADMIN in role_names:
            # Admins can see all
            pass
        else:
            # Students see only global (college_id is None) or their college
            from app.models.profile import Profile
            profile = db.query(Profile).filter(Profile.user_id == current_user.id).first()
            if profile and profile.college_id:
                query = query.filter(
                    (JobAggregation.college_id == profile.college_id) |
                    (JobAggregation.college_id.is_(None))
                )
            else:
                query = query.filter(JobAggregation.college_id.is_(None))
    
    query = query.filter(JobAggregation.is_active == True)
    jobs = query.order_by(JobAggregation.posted_date.desc()).offset(skip).limit(limit).all()
    
    return jobs


@router.post("/{aggregation_id}/import", response_model=dict)
async def import_to_jobs(
    aggregation_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Import an aggregated job to the main jobs table (Admin only)"""
    aggregation = db.query(JobAggregation).filter(JobAggregation.id == aggregation_id).first()
    
    if not aggregation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job aggregation not found"
        )
    
    if aggregation.is_imported:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Job has already been imported"
        )
    
    # Get admin's college
    admin_role = db.query(UserRole).filter(
        UserRole.user_id == current_user.id,
        UserRole.role == RoleEnum.ADMIN
    ).first()
    
    if not admin_role or not admin_role.college_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Admin must be associated with a college"
        )
    
    # Create job in main jobs table
    new_job = Job(
        title=aggregation.title,
        company=aggregation.company,
        role=aggregation.role,
        description=aggregation.description,
        location=aggregation.location,
        ctc=aggregation.ctc,
        job_type=aggregation.job_type or "Off-Campus",
        eligibility_type="all_students",
        posted_date=aggregation.posted_date or datetime.utcnow(),
        college_id=admin_role.college_id,
        created_by=current_user.id
    )
    
    db.add(new_job)
    aggregation.is_imported = True
    db.commit()
    db.refresh(new_job)
    
    return {
        "message": "Job imported successfully",
        "job_id": new_job.id,
        "aggregation_id": aggregation.id
    }


@router.put("/{aggregation_id}", response_model=JobAggregationResponse)
async def update_aggregation(
    aggregation_id: int,
    update_data: JobAggregationUpdate,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update a job aggregation (Admin only)"""
    aggregation = db.query(JobAggregation).filter(JobAggregation.id == aggregation_id).first()
    
    if not aggregation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job aggregation not found"
        )
    
    update_dict = update_data.model_dump(exclude_unset=True)
    for field, value in update_dict.items():
        setattr(aggregation, field, value)
    
    db.commit()
    db.refresh(aggregation)
    
    return aggregation


@router.delete("/{aggregation_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_aggregation(
    aggregation_id: int,
    current_user: User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Delete a job aggregation (Admin only)"""
    aggregation = db.query(JobAggregation).filter(JobAggregation.id == aggregation_id).first()
    
    if not aggregation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Job aggregation not found"
        )
    
    db.delete(aggregation)
    db.commit()
    
    return None


def get_current_super_admin(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> User:
    """Verify user is super admin"""
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only super admins can perform this action"
        )
    
    return current_user


@router.post("/bulk-upload", response_model=dict)
async def bulk_upload_jobs(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Bulk upload jobs from Excel or CSV file (Super Admin only)
    
    Supported formats: .xlsx, .xls, .csv
    
    Excel/CSV Format (headers required):
    source,title,company,role,description,location,ctc,job_type,experience_required,skills_required,qualifications,posted_date,expiry_date,source_url,external_id
    
    Required fields: source, title, company, role
    Optional fields: description, location, ctc, job_type, experience_required, skills_required (comma-separated), qualifications, posted_date, expiry_date, source_url, external_id
    
    Example:
    source,title,company,role,description,location,ctc,job_type,experience_required,skills_required,qualifications,posted_date,expiry_date,source_url,external_id
    linkedin,Software Engineer,Google,Software Engineer,Looking for experienced developer,Hyderabad,₹15 LPA,Full-time,2-5 years,"Python,JavaScript,React",Bachelor's degree,2024-01-15,2024-12-31,https://linkedin.com/jobs/123,linkedin_123
    """
    file_ext = file.filename.lower().split('.')[-1]
    
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    data = []
    
    # Parse based on file type
    if file_ext == 'csv':
        csv_content = contents.decode('utf-8')
        reader = csv.DictReader(io.StringIO(csv_content))
        data = list(reader)
    else:
        # Excel file
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            job_dict = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    value = row[idx]
                    if value is not None:
                        job_dict[header.strip()] = str(value).strip()
                    else:
                        job_dict[header.strip()] = ""
            if job_dict:  # Only add if dict has data
                data.append(job_dict)
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    for idx, job_data in enumerate(data, start=1):
        try:
            # Validate required fields
            if not job_data.get('source') or not job_data.get('title') or not job_data.get('company') or not job_data.get('role'):
                raise ValueError("Missing required fields: source, title, company, role")
            
            source = job_data['source'].lower().strip()
            if source not in ['linkedin', 'indeed', 'naukri', 'glassdoor', 'monster', 'other']:
                source = 'other'
            
            # Parse skills_required (comma-separated string to list)
            skills_required = None
            if job_data.get('skills_required'):
                skills_list = [s.strip() for s in str(job_data['skills_required']).split(',') if s.strip()]
                if skills_list:
                    skills_required = skills_list
            
            # Parse dates
            posted_date = None
            if job_data.get('posted_date'):
                try:
                    # Try parsing various date formats
                    date_str = str(job_data['posted_date']).strip()
                    for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            posted_date = datetime.strptime(date_str, fmt)
                            break
                        except:
                            continue
                except:
                    pass
            
            expiry_date = None
            if job_data.get('expiry_date'):
                try:
                    date_str = str(job_data['expiry_date']).strip()
                    for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            expiry_date = datetime.strptime(date_str, fmt)
                            break
                        except:
                            continue
                except:
                    pass
            
            # Check if job already exists (by external_id or title+company+source)
            existing = None
            if job_data.get('external_id'):
                existing = db.query(JobAggregation).filter(
                    JobAggregation.source == source,
                    JobAggregation.external_id == job_data['external_id']
                ).first()
            
            if not existing:
                existing = db.query(JobAggregation).filter(
                    JobAggregation.source == source,
                    JobAggregation.title == job_data['title'],
                    JobAggregation.company == job_data['company']
                ).first()
            
            if existing:
                # Update existing
                existing.title = job_data['title']
                existing.company = job_data['company']
                existing.role = job_data['role']
                existing.description = job_data.get('description') or existing.description
                existing.location = job_data.get('location') or existing.location
                existing.ctc = job_data.get('ctc') or existing.ctc
                existing.job_type = job_data.get('job_type') or existing.job_type
                existing.experience_required = job_data.get('experience_required') or existing.experience_required
                existing.skills_required = skills_required or existing.skills_required
                existing.qualifications = job_data.get('qualifications') or existing.qualifications
                existing.posted_date = posted_date or existing.posted_date
                existing.expiry_date = expiry_date or existing.expiry_date
                existing.source_url = job_data.get('source_url') or existing.source_url
                existing.external_id = job_data.get('external_id') or existing.external_id
                existing.last_synced_at = datetime.utcnow()
                
                results["success"].append({
                    "index": idx,
                    "title": job_data['title'],
                    "id": existing.id,
                    "action": "updated"
                })
            else:
                # Create new
                new_job = JobAggregation(
                    source=source,
                    external_id=job_data.get('external_id'),
                    source_url=job_data.get('source_url'),
                    title=job_data['title'],
                    company=job_data['company'],
                    role=job_data['role'],
                    description=job_data.get('description'),
                    location=job_data.get('location'),
                    ctc=job_data.get('ctc'),
                    job_type=job_data.get('job_type'),
                    experience_required=job_data.get('experience_required'),
                    skills_required=skills_required,
                    qualifications=job_data.get('qualifications'),
                    posted_date=posted_date or datetime.utcnow(),
                    expiry_date=expiry_date,
                    is_active=True,
                    is_imported=False,
                    last_synced_at=datetime.utcnow()
                )
                db.add(new_job)
                db.flush()
                
                results["success"].append({
                    "index": idx,
                    "title": job_data['title'],
                    "id": new_job.id,
                    "action": "created"
                })
            
            results["total"] += 1
            
        except Exception as e:
            results["failed"].append({
                "index": idx,
                "title": job_data.get('title', 'N/A'),
                "error": str(e)
            })
    
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.get("/template")
async def download_job_template(
    current_user: User = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Download Excel template for job bulk upload (Super Admin only)"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs Template"
    
    # Headers
    headers = [
        "source", "title", "company", "role", "description", "location", "ctc",
        "job_type", "experience_required", "skills_required", "qualifications",
        "posted_date", "expiry_date", "source_url", "external_id"
    ]
    ws.append(headers)
    
    # Example row
    example_row = [
        "linkedin",
        "Software Engineer",
        "Google",
        "Software Engineer",
        "Looking for experienced software engineer with 2-5 years of experience",
        "Hyderabad",
        "₹15 LPA",
        "Full-time",
        "2-5 years",
        "Python,JavaScript,React,Node.js",
        "Bachelor's degree in Computer Science or related field",
        "2024-01-15",
        "2024-12-31",
        "https://linkedin.com/jobs/view/123456",
        "linkedin_123456"
    ]
    ws.append(example_row)
    
    # Style headers
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=job_upload_template.xlsx"}
    )

