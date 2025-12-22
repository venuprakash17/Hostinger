"""Bulk upload API endpoints"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy.orm import Session, load_only
from sqlalchemy import text
from typing import List, Optional
import csv
import io
import json
import hashlib
import time
from datetime import datetime
from openpyxl import load_workbook
from app.core.database import get_db
from app.models.user import User, UserRole, RoleEnum
from app.models.profile import Profile
from app.models.quiz import Quiz, CodingProblem
from app.models.company_training import RoundContent as RoundContentModel, Round as RoundModel, RoundType
from app.models.academic import (
    Subject, Section, Department, FacultySectionAssignment,
    SubjectAssignment, Semester, AcademicYear
)
from app.api.auth import get_current_user
from app.core.security import get_password_hash
from app.models.audit_log import AuditLog
from app.models.college import College
from app.models.institution import Institution

router = APIRouter(prefix="/bulk-upload", tags=["bulk-upload"])

# Import and include debug router
from app.api.bulk_upload_debug import router as debug_router
router.include_router(debug_router)


def get_current_super_admin(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
) -> User:
    """Verify user is super admin"""
    from app.models.user import UserRole, RoleEnum
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    if RoleEnum.SUPER_ADMIN not in role_names:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only super admins can perform bulk uploads"
        )
    
    return current_user


def get_current_admin_or_super(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Verify user is admin or super admin, return (user, is_super_admin)"""
    from app.models.user import UserRole, RoleEnum
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    
    is_super_admin = RoleEnum.SUPER_ADMIN in role_names
    is_admin = RoleEnum.ADMIN in role_names
    
    if not (is_super_admin or is_admin):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admins or super admins can perform this action"
        )
    
    return current_user, is_super_admin


def get_admin_college_id(
    current_user: User,
    db: Session
) -> Optional[int]:
    """Get college_id for admin/HOD/faculty user - checks role first, then profile"""
    from app.models.user import UserRole, RoleEnum
    from app.models.profile import Profile
    
    # Try to get from admin role
    admin_role = db.query(UserRole).filter(
        UserRole.user_id == current_user.id,
        UserRole.role == RoleEnum.ADMIN
    ).first()
    
    if admin_role and admin_role.college_id:
        return admin_role.college_id
    
    # Try to get from HOD role
    hod_role = db.query(UserRole).filter(
        UserRole.user_id == current_user.id,
        UserRole.role == RoleEnum.HOD
    ).first()
    
    if hod_role and hod_role.college_id:
        return hod_role.college_id
    
    # Try to get from faculty role
    faculty_role = db.query(UserRole).filter(
        UserRole.user_id == current_user.id,
        UserRole.role == RoleEnum.FACULTY
    ).first()
    
    if faculty_role and faculty_role.college_id:
        return faculty_role.college_id
    
    # Fallback: try to get from profile
    profile = db.query(Profile).filter(Profile.user_id == current_user.id).first()
    if profile and profile.college_id:
        return profile.college_id
    
    return None


@router.post("/students")
async def bulk_upload_students(
    file: UploadFile = File(...),
    college_id: Optional[int] = Query(None, description="College ID - optional, auto-detected from user context if not provided"),
    current_user_tuple = Depends(get_current_admin_or_super),
    db: Session = Depends(get_db)
):
    """Bulk upload students from CSV or Excel file.
    
    College admin can upload students for their college (college_id auto-detected).
    Super admin can upload for any college (college_id required if not provided).
    
    Supported formats: .csv, .xlsx, .xls
    
    Format (headers required):
    email,password,full_name,branch_id,section,roll_number,present_year
    
    Note: college_id is auto-detected from user context. Super admins can optionally provide it.
    
    Fields:
    - email: Student email (required)
    - password: Password (optional, defaults to roll_number or user.id)
    - full_name: Student full name
    - branch_id: Branch ID (required - e.g., "CSE001", "ECE001")
      Note: Department will be automatically resolved from branch_id
    - section: Section name (e.g., "A", "B", "C") - required
      Note: Sections will be created by college admin later and will automatically link by name
    - roll_number: Student roll number
    - present_year: Academic year (1, 2, 3, 4, or 5) - used for faculty utilization
    
    Example:
    email,password,full_name,branch_id,section,roll_number,present_year
    student1@example.com,Password123,John Doe,CSE001,A,20CS001,1
    student2@example.com,Password123,Jane Smith,CSE001,B,20CS002,1
    student3@example.com,Password123,Jane Doe,ECE001,A,20EC001,2
    
    Note: Section names (A, B, C, etc.) will automatically link when college admin creates sections
    with matching names in the department.
    """
    current_user, is_super_admin = current_user_tuple
    
    # Auto-detect college_id if not provided
    if not college_id:
        if is_super_admin:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="College ID is required for super admin uploads. Provide it as query parameter: ?college_id=X"
            )
        else:
            # Auto-detect from user context
            auto_college_id = get_admin_college_id(current_user, db)
            if not auto_college_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Could not determine college ID. Admin must be associated with a college."
                )
            college_id = auto_college_id
    
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    # Read file content
    contents = await file.read()
    
    # Parse based on file type
    if file_ext == 'csv':
        csv_content = contents.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))
    else:
        # Excel file
        excel_file = io.BytesIO(contents)
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Get headers from first row and normalize to lowercase
        headers = [cell.value for cell in sheet[1]]
        if not headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must have headers in the first row"
            )
        
        # Normalize headers to lowercase for case-insensitive matching
        normalized_headers = [str(h).strip().lower() if h else None for h in headers]
        
        # Convert Excel rows to CSV-like dict format
        csv_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for idx, header in enumerate(normalized_headers):
                if header:
                    # Safely access row[idx] - handle cases where row has fewer columns than headers
                    if idx < len(row):
                        row_dict[header] = str(row[idx]).strip() if row[idx] is not None else ""
                    else:
                        row_dict[header] = ""
            csv_rows.append(row_dict)
        
        csv_reader = csv_rows if csv_rows else []
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    # Handle both CSV reader (iterator) and Excel rows (list)
    rows_to_process = list(csv_reader) if isinstance(csv_reader, list) else csv_reader
    
    # Track emails processed in this upload to catch duplicates within the same file
    emails_in_upload = set()
    
    for row_num, row in enumerate(rows_to_process, start=2):  # Start at 2 (row 1 is header)
        try:
            email = row.get('email', '').strip().lower()
            password = row.get('password', '').strip()
            full_name = row.get('full_name', '').strip()
            branch_id_str = row.get('branch_id', '').strip()
            section = row.get('section', '').strip()
            roll_number = row.get('roll_number', '').strip()
            present_year = row.get('present_year', '').strip()
            
            # College ID is auto-detected or provided via query parameter
            final_college_id = college_id
            
            if not email:
                results["failed"].append({
                    "row": row_num,
                    "email": email or "N/A",
                    "error": "Email is required"
                })
                continue
            
            # Check for duplicate email within the same upload file
            if email in emails_in_upload:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": f"Duplicate email in upload file (already processed in this upload)"
                })
                continue
            
            # Check if user already exists in database
            existing = db.query(User).filter(User.email == email).first()
            if existing:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": "User already exists in database"
                })
                continue
            
            # Mark this email as processed
            emails_in_upload.add(email)
            
            # Resolve department_id from branch_id only
            from app.models.academic import Department
            department_id = None
            department_name = None
            
            if not branch_id_str:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": "Branch ID is required (e.g., 'CSE001', 'ECE001')"
                })
                continue
            
            # Find department by branch_id
            dept = db.query(Department).filter(
                Department.branch_id == branch_id_str,
                Department.college_id == final_college_id,
                Department.is_active == True
            ).first()
            if dept:
                department_id = dept.id
                department_name = dept.name
            else:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": f"Branch ID '{branch_id_str}' not found for this college"
                })
                continue
            
            # Resolve section_id from section name (if section exists)
            # Auto-create section if it doesn't exist
            from app.models.academic import Section
            section_id = None
            section_name = section if section else None
            
            if section and department_id:
                # Try to find section by name within the department
                sect = db.query(Section).filter(
                    Section.name == section,
                    Section.department_id == department_id,
                    Section.college_id == final_college_id,
                    Section.is_active == True
                ).first()
                
                if sect:
                    section_id = sect.id
                    section_name = sect.name
                else:
                    # Auto-create section if it doesn't exist
                    # Use present_year to determine section year
                    from app.core.year_utils import parse_year
                    year_int = None
                    if present_year:
                        year_str = parse_year(present_year)
                        if year_str:
                            try:
                                year_int = int(year_str)
                            except (ValueError, TypeError):
                                pass
                    
                    # Create new section
                    new_section = Section(
                        name=section.strip(),
                        college_id=final_college_id,
                        department_id=department_id,
                        year=year_int,
                        is_active=True
                    )
                    db.add(new_section)
                    db.flush()  # Get section.id
                    section_id = new_section.id
                    section_name = new_section.name
                    print(f"[Bulk Upload] Auto-created section '{section_name}' for department {department_id}, year {year_int}")
            
            # Generate password: use provided password, or roll_number in caps, or user_id in caps, or user.id in caps
            if password:
                final_password = password
            elif roll_number:
                final_password = roll_number.upper()
            else:
                # Will use user.id after creation
                final_password = "TEMP_PASSWORD_PLACEHOLDER"
            
            # Normalize present_year: convert "1st", "2nd", "3rd" to "1", "2", "3" for storage
            from app.core.year_utils import parse_year
            numeric_year = parse_year(present_year) if present_year else None
            
            # Create user
            user = User(
                email=email,
                password_hash=get_password_hash(final_password),
                is_active="true",
                is_verified="true"
            )
            db.add(user)
            db.flush()  # Get user.id
            
            # If password was placeholder, update it with user.id in caps
            if final_password == "TEMP_PASSWORD_PLACEHOLDER":
                user.password_hash = get_password_hash(str(user.id).upper())
            
            # Create profile with department_id and section_id
            profile = Profile(
                user_id=user.id,
                email=email,
                full_name=full_name if full_name else None,
                college_id=final_college_id,  # Required from query parameter
                department=department_name,  # Department name from branch_id
                department_id=department_id,  # Link to Department
                section=section_name or section,  # Keep name for backward compatibility
                section_id=section_id,  # Link to Section
                roll_number=roll_number if roll_number else None,
                present_year=numeric_year  # Store year for faculty utilization
            )
            db.add(profile)
            
            # Create student role
            role = UserRole(
                user_id=user.id,
                role=RoleEnum.STUDENT,
                college_id=final_college_id
            )
            db.add(role)
            
            results["success"].append({
                "row": row_num,
                "email": email,
                "name": full_name
            })
            results["total"] += 1
            
        except Exception as e:
            results["failed"].append({
                "row": row_num,
                "email": row.get('email', 'N/A'),
                "error": str(e)
            })
    
    # Commit all successful inserts
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.get("/template/institution-students")
async def download_institution_student_template(
    format: str = Query("csv", description="File format: csv or xlsx"),
    current_user: User = Depends(get_current_super_admin)
):
    """Download CSV or Excel template for institution student bulk upload
    
    Format: email,password,full_name
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    if format.lower() == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Institution Students Template"
        
        # Headers
        headers = ["email", "password", "full_name"]
        ws.append(headers)
        
        # Example rows
        examples = [
            ["student1@institution.com", "Password123", "John Doe"],
            ["student2@institution.com", "Password123", "Jane Smith"],
            ["student3@institution.com", "", "Jane Doe"]  # Password optional
        ]
        for example in examples:
            ws.append(example)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=institution_student_upload_template.xlsx",
            }
        )
    else:
        # CSV format
        template = "email,password,full_name\n"
        template += "student1@institution.com,Password123,John Doe\n"
        template += "student2@institution.com,Password123,Jane Smith\n"
        template += "student3@institution.com,,Jane Doe\n"
        
        return StreamingResponse(
            iter([template.encode('utf-8')]),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=institution_student_upload_template.csv",
                "Content-Length": str(len(template.encode('utf-8')))
            }
        )


@router.post("/institution-students")
async def bulk_upload_institution_students(
    file: UploadFile = File(...),
    institution_id: Optional[int] = Query(None, description="Institution ID - required for super admin uploads"),
    current_user: User = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Bulk upload institution students from CSV or Excel file.
    
    Super admin can upload students for any institution (institution_id required).
    
    Supported formats: .csv, .xlsx, .xls
    
    Format (headers required):
    email,password,full_name
    
    Fields:
    - email: Student email (required)
    - password: Password (optional, defaults to email or user.id)
    - full_name: Student full name (optional)
    
    Example:
    email,password,full_name
    student1@institution.com,Password123,John Doe
    student2@institution.com,Password123,Jane Smith
    student3@institution.com,,Jane Doe
    """
    if not institution_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Institution ID is required. Provide it as query parameter: ?institution_id=X"
        )
    
    # Verify institution exists
    institution = db.query(Institution).filter(Institution.id == institution_id).first()
    if not institution:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Institution with ID {institution_id} not found"
        )
    
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    # Read file content
    contents = await file.read()
    
    # Parse based on file type
    if file_ext == 'csv':
        csv_content = contents.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))
    else:
        # Excel file
        excel_file = io.BytesIO(contents)
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Get headers from first row and normalize to lowercase
        headers = [cell.value for cell in sheet[1]]
        if not headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must have headers in the first row"
            )
        
        # Normalize headers to lowercase for case-insensitive matching
        normalized_headers = [str(h).strip().lower() if h else None for h in headers]
        
        # Convert Excel rows to CSV-like dict format
        csv_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for idx, header in enumerate(normalized_headers):
                if header:
                    if idx < len(row):
                        row_dict[header] = str(row[idx]).strip() if row[idx] is not None else ""
                    else:
                        row_dict[header] = ""
            csv_rows.append(row_dict)
        
        csv_reader = csv_rows if csv_rows else []
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    # Handle both CSV reader (iterator) and Excel rows (list)
    rows_to_process = list(csv_reader) if isinstance(csv_reader, list) else csv_reader
    
    # Track emails processed in this upload to catch duplicates within the same file
    emails_in_upload = set()
    
    for row_num, row in enumerate(rows_to_process, start=2):  # Start at 2 (row 1 is header)
        try:
            email = row.get('email', '').strip().lower()
            password = row.get('password', '').strip()
            full_name = row.get('full_name', '').strip()
            
            if not email:
                results["failed"].append({
                    "row": row_num,
                    "email": email or "N/A",
                    "error": "Email is required"
                })
                continue
            
            # Check for duplicate email within the same upload file
            if email in emails_in_upload:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": f"Duplicate email in upload file (already processed in this upload)"
                })
                continue
            
            # Check if user already exists in database
            existing = db.query(User).filter(User.email == email).first()
            if existing:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": "User already exists in database"
                })
                continue
            
            # Mark this email as processed
            emails_in_upload.add(email)
            
            # Generate password: use provided password, or email, or user.id
            if password:
                final_password = password
            elif email:
                final_password = email.split('@')[0]  # Use part before @ as default
            else:
                final_password = "TEMP_PASSWORD_PLACEHOLDER"
            
            # Create user
            user = User(
                email=email,
                password_hash=get_password_hash(final_password),
                is_active="true",
                is_verified="true"
            )
            db.add(user)
            db.flush()  # Get user.id
            
            # If password was placeholder, update it with user.id
            if final_password == "TEMP_PASSWORD_PLACEHOLDER":
                user.password_hash = get_password_hash(str(user.id).upper())
            
            # Create profile with institution_id
            profile = Profile(
                user_id=user.id,
                email=email,
                full_name=full_name if full_name else None,
                institution_id=institution_id,
                college_id=None  # Institution students don't have college_id
            )
            db.add(profile)
            
            # Create institution student role
            role = UserRole(
                user_id=user.id,
                role=RoleEnum.INSTITUTION_STUDENT,
                institution_id=institution_id,
                college_id=None  # Institution students don't have college_id
            )
            db.add(role)
            
            results["success"].append({
                "row": row_num,
                "email": email,
                "name": full_name
            })
            results["total"] += 1
            
        except Exception as e:
            results["failed"].append({
                "row": row_num,
                "email": row.get('email', 'N/A'),
                "error": str(e)
            })
    
    # Commit all successful inserts
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


def get_current_content_creator(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Verify user can create content (super admin, admin, HOD, or faculty)"""
    user_roles = db.query(UserRole).filter(UserRole.user_id == current_user.id).all()
    role_names = [role.role for role in user_roles]
    profile = db.query(Profile).filter(Profile.user_id == current_user.id).first()
    
    is_super_admin = RoleEnum.SUPER_ADMIN in role_names
    is_admin = RoleEnum.ADMIN in role_names
    is_hod = RoleEnum.HOD in role_names
    is_faculty = RoleEnum.FACULTY in role_names
    
    if not (is_super_admin or is_admin or is_hod or is_faculty):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admins, faculty, HOD, or super admins can upload content"
        )
    
    college_id = None
    if user_roles:
        college_id = user_roles[0].college_id
    if not college_id and profile:
        college_id = profile.college_id
    
    return current_user, {
        "is_super_admin": is_super_admin,
        "is_admin": is_admin,
        "is_hod": is_hod,
        "is_faculty": is_faculty,
        "college_id": college_id,
        "profile": profile
    }


@router.post("/quizzes")
async def bulk_upload_quizzes(
    file: UploadFile = File(...),
    current_user_tuple = Depends(get_current_content_creator),
    db: Session = Depends(get_db)
):
    """Bulk upload quizzes from Excel, CSV, or JSON file.
    
    Supported formats: .xlsx, .xls, .csv, .json
    
    Excel/CSV Format (headers required):
    title,description,subject,duration_minutes,total_marks,scope_type,section_id,year,expiry_date,is_active
    
    JSON Format:
    [
      {
        "title": "Quiz Title",
        "description": "Description",
        "subject": "Subject",
        "duration_minutes": 30,
        "total_marks": 100,
        "scope_type": "section",
        "section_id": 1,
        "year": "1st",
        "expiry_date": "2024-12-31T23:59:59",
        "questions": [...],
        "is_active": true
      }
    ]
    """
    current_user, user_info = current_user_tuple
    file_ext = file.filename.lower().split('.')[-1]
    
    if file_ext not in ['json', 'csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only JSON, CSV, and Excel files (.json, .csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    data = []
    
    # Parse based on file type
    if file_ext == 'json':
        data = json.loads(contents.decode('utf-8'))
        if not isinstance(data, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="JSON must be an array of quiz objects"
            )
    else:
        # Excel or CSV
        if file_ext in ['xlsx', 'xls']:
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                quiz_dict = dict(zip(headers, row))
                data.append(quiz_dict)
        else:  # CSV
            csv_content = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_content))
            data = list(reader)
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    for idx, quiz_data in enumerate(data, start=1):
        try:
            # Convert string values to appropriate types for Excel/CSV
            if isinstance(quiz_data.get('duration_minutes'), str):
                quiz_data['duration_minutes'] = int(quiz_data['duration_minutes']) if quiz_data.get('duration_minutes') else 30
            if isinstance(quiz_data.get('total_marks'), str):
                quiz_data['total_marks'] = int(quiz_data['total_marks']) if quiz_data.get('total_marks') else 100
            if isinstance(quiz_data.get('section_id'), str):
                quiz_data['section_id'] = int(quiz_data['section_id']) if quiz_data.get('section_id') else None
            
            # Validate required fields
            if not quiz_data.get('title'):
                results["failed"].append({
                    "index": idx,
                    "title": quiz_data.get('title', 'N/A'),
                    "error": "Title is required"
                })
                continue
            
            # Determine scope based on role
            scope_type = quiz_data.get('scope_type', 'section')
            if user_info["is_super_admin"]:
                scope_type = "svnapro"
            elif user_info["is_admin"]:
                scope_type = quiz_data.get('scope_type', 'college')
            elif user_info["is_hod"]:
                scope_type = quiz_data.get('scope_type', 'department')
            else:
                scope_type = "section"
            
            # Set scope fields
            college_id = user_info["college_id"]
            department = user_info["profile"].department if user_info["profile"] else None
            section_id = quiz_data.get('section_id') if scope_type == "section" else None
            year = quiz_data.get('year') if scope_type in ["department", "college"] else None
            
            # Parse expiry_date if provided
            expiry_date = None
            if quiz_data.get('expiry_date'):
                try:
                    expiry_date = datetime.fromisoformat(str(quiz_data['expiry_date']).replace('Z', '+00:00'))
                except:
                    pass
            
            quiz = Quiz(
                title=quiz_data['title'],
                description=quiz_data.get('description'),
                subject=quiz_data.get('subject'),
                duration_minutes=quiz_data.get('duration_minutes', 30),
                total_marks=quiz_data.get('total_marks', 100),
                questions=quiz_data.get('questions', []),
                is_active=quiz_data.get('is_active', True) if isinstance(quiz_data.get('is_active'), bool) else (str(quiz_data.get('is_active', 'true')).lower() == 'true'),
                created_by=current_user.id,
                scope_type=scope_type,
                college_id=college_id,
                department=department,
                section_id=section_id,
                year=year,
                expiry_date=expiry_date
            )
            db.add(quiz)
            db.flush()
            
            results["success"].append({
                "index": idx,
                "title": quiz_data['title'],
                "id": quiz.id
            })
            results["total"] += 1
            
        except Exception as e:
            results["failed"].append({
                "index": idx,
                "title": quiz_data.get('title', 'N/A'),
                "error": str(e)
            })
    
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.post("/coding-problems")
async def bulk_upload_coding_problems(
    file: UploadFile = File(...),
    current_user_tuple = Depends(get_current_content_creator),
    db: Session = Depends(get_db)
):
    """Bulk upload coding problems from Excel, CSV, or JSON file (All Staff: Super Admin, Admin, Faculty, HOD).
    
    Supported formats: .xlsx, .xls, .csv, .json
    
    Excel/CSV Format (headers required):
    title,description,input_format,output_format,difficulty,tags,constraints,sample_input,sample_output,
    section,year,allowed_languages,restricted_languages,recommended_languages,
    starter_code_python,starter_code_c,starter_code_cpp,starter_code_java,starter_code_javascript,
    time_limit,memory_limit,test_cases,expiry_date,is_active
    
    Note: section and year are optional. If both empty, visible to whole college.
    - section: Section name (e.g., "A", "B") - optional
    - year: Year (1, 2, 3, or 4) - optional
    
    JSON Format:
    [
      {
        "title": "Problem Title",
        "description": "Problem description",
        "input_format": "Input format description",
        "output_format": "Output format description",
        "difficulty": "Easy|Medium|Hard",
        "tags": ["tag1", "tag2"],
        "constraints": "Constraints text",
        "sample_input": "Input example",
        "sample_output": "Output example",
        "year": 1,
        "allowed_languages": ["python", "c", "cpp"],
        "restricted_languages": [],
        "recommended_languages": ["c"],
        "starter_code_python": "def solution():\\n    pass",
        "time_limit": 5,
        "memory_limit": 256,
        "test_cases": [{"stdin": "1", "expected_output": "1", "is_public": true}],
        "expiry_date": "2024-12-31T23:59:59",
        "is_active": true
      }
    ]
    """
    # Get user info for scope determination
    current_user_obj, user_info = current_user_tuple
    
    file_ext = file.filename.lower().split('.')[-1]
    
    if file_ext not in ['json', 'csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only JSON, CSV, and Excel files (.json, .csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    data = []
    
    # Parse based on file type
    if file_ext == 'json':
        data = json.loads(contents.decode('utf-8'))
        if not isinstance(data, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="JSON must be an array of coding problem objects"
            )
    else:
        # Excel or CSV
        if file_ext in ['xlsx', 'xls']:
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                problem_dict = dict(zip(headers, row))
                data.append(problem_dict)
        else:  # CSV
            csv_content = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_content))
            data = list(reader)
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    def parse_year(value):
        """Parse year to integer (1-4)"""
        if isinstance(value, int):
            return value if 1 <= value <= 4 else 1
        if isinstance(value, str):
            value = value.lower().strip()
            if value.startswith('1'): return 1
            if value.startswith('2'): return 2
            if value.startswith('3'): return 3
            if value.startswith('4'): return 4
            try:
                num = int(value)
                return num if 1 <= num <= 4 else 1
            except:
                return 1
        return 1
    
    def parse_language_array(value):
        """Parse language array from string (CSV or JSON)"""
        if isinstance(value, list):
            return value
        if isinstance(value, str):
            # Try JSON first
            try:
                parsed = json.loads(value)
                if isinstance(parsed, list):
                    return parsed
            except:
                pass
            # Try CSV
            return [lang.strip() for lang in value.split(',') if lang.strip()]
        return []
    
    def parse_test_cases(value):
        """Parse test cases from string (JSON) or return as-is"""
        if isinstance(value, list):
            return value
        if isinstance(value, str):
            try:
                return json.loads(value)
            except:
                return []
        return []
    
    # Get user info for scope determination
    current_user_obj, user_info = current_user_tuple
    
    for idx, problem_data in enumerate(data, start=1):
        # Start a savepoint for each row to allow rollback on error
        savepoint = None
        try:
            # Validate required fields
            if not problem_data.get('title') or not problem_data.get('description'):
                results["failed"].append({
                    "index": idx,
                    "title": problem_data.get('title', 'N/A'),
                    "error": "Title and description are required"
                })
                continue
            
            # Create savepoint for this row (if transaction supports it)
            savepoint = None
            try:
                # Only create savepoint if we're in a transaction
                if hasattr(db, 'in_transaction') and db.in_transaction():
                    savepoint = db.begin_nested()
                elif hasattr(db, 'connection') and db.connection:
                    savepoint = db.begin_nested()
            except Exception:
                # If savepoints not supported, continue without it
                savepoint = None
            
            # Parse year (optional, must be 1-4 if provided)
            year_raw = problem_data.get('year')
            year = None
            if year_raw:
                year = parse_year(year_raw)
                if year not in [1, 2, 3, 4]:
                    results["failed"].append({
                        "index": idx,
                        "title": problem_data.get('title', 'N/A'),
                        "error": f"Year must be 1, 2, 3, or 4. Got: {year_raw}"
                    })
                    continue
            
            # Determine scope based on user role and section/year fields
            scope_type = "svnapro"
            college_id = None
            department = None
            section_id = None  # Not used for coding problems (year-based restriction only)
            year_value = None
            year_str_value = None
            
            # Get year from problem_data (section removed - year-based restriction only)
            year_value = year  # Use parsed year (can be None)
            
            # Determine scope - year-based restriction only (no section restriction)
            if not user_info.get("is_super_admin"):
                # Staff must create college-level problems
                college_id = user_info.get("college_id")
                if user_info.get("profile") and user_info["profile"].department:
                    department = user_info["profile"].department
                
                # Simplified Visibility Logic (year-based only):
                # - Year provided → department-wide for that year (scope_type = "department")
                # - No year → whole college (scope_type = "college")
                
                if year_value:
                    # Year provided → department-wide for that year
                    scope_type = "department"
                    year_str_value = f"{year_value}st" if year_value == 1 else f"{year_value}nd" if year_value == 2 else f"{year_value}rd" if year_value == 3 else f"{year_value}th"
                else:
                    # No year → whole college
                    scope_type = "college"
                
                # section_id not used for coding problems (year-based restriction only)
            else:
                # Super admin can create SvnaPro problems
                scope_type = "svnapro"
            
            # Convert year to year_str if not already set
            if not year_str_value and year_value:
                year_str_value = f"{year_value}st" if year_value == 1 else f"{year_value}nd" if year_value == 2 else f"{year_value}rd" if year_value == 3 else f"{year_value}th"
            
            # Parse language arrays
            allowed_languages = parse_language_array(problem_data.get('allowed_languages', '["python","c","cpp","java","javascript"]'))
            restricted_languages = parse_language_array(problem_data.get('restricted_languages', '[]'))
            recommended_languages = parse_language_array(problem_data.get('recommended_languages', '[]'))
            
            # Parse tags
            tags = []
            if isinstance(problem_data.get('tags'), str):
                tags = [tag.strip() for tag in problem_data['tags'].split(',') if tag.strip()]
            elif isinstance(problem_data.get('tags'), list):
                tags = problem_data['tags']
            
            # Parse test cases
            test_cases = parse_test_cases(problem_data.get('test_cases', '[]'))
            
            # Parse expiry_date if provided
            expiry_date = None
            if problem_data.get('expiry_date'):
                try:
                    expiry_date = datetime.fromisoformat(str(problem_data['expiry_date']).replace('Z', '+00:00'))
                except:
                    pass
            
            # Parse numeric fields
            time_limit = int(problem_data.get('time_limit', 5))
            memory_limit = int(problem_data.get('memory_limit', 256))
            
            # Check for duplicates: same title AND description (case-insensitive, normalized)
            title = problem_data['title'].strip()
            description = problem_data['description'].strip()
            
            # Normalize for comparison (lowercase, remove extra whitespace)
            title_normalized = ' '.join(title.lower().split())
            description_normalized = ' '.join(description.lower().split())
            
            # Check for duplicates using raw SQL ONLY to completely avoid ORM column issues
            # Use raw SQL to only query columns that exist (exclude academic_year_id and section)
            existing_problem_id = None
            try:
                # First, try to get all problems using raw SQL (only existing columns)
                raw_sql = text("""
                    SELECT id, title, description, year, year_str, problem_code, created_at
                    FROM coding_problems
                """)
                result = db.execute(raw_sql)
                all_problems = result.fetchall()
                
                # Check for normalized match
                for row in all_problems:
                    prob_title = row[1] if row[1] else ''
                    prob_desc = row[2] if row[2] else ''
                    prob_title_norm = ' '.join(prob_title.lower().split())
                    prob_desc_norm = ' '.join(prob_desc.lower().split())
                    if prob_title_norm == title_normalized and prob_desc_norm == description_normalized:
                        # Found match - store ID only (will use raw SQL for update)
                        existing_problem_id = row[0]
                        break
                
                # If not found, try exact match as fallback
                if not existing_problem_id:
                    raw_sql_exact = text("""
                        SELECT id FROM coding_problems
                        WHERE title = :title AND description = :description
                        ORDER BY created_at DESC
                        LIMIT 1
                    """)
                    result_exact = db.execute(
                        raw_sql_exact,
                        {"title": title, "description": description}
                    )
                    row_exact = result_exact.fetchone()
                    if row_exact:
                        existing_problem_id = row_exact[0]
            except Exception as e:
                # If raw SQL fails, rollback and continue without duplicate check
                db.rollback()
                existing_problem_id = None
            
            # Generate unique problem_code for analytics tracking
            problem_code_base = f"{title}_{description}".lower().replace(" ", "_")[:50]
            problem_code = f"CP_{hashlib.md5(f'{title}_{description}_{time.time()}'.encode()).hexdigest()[:12].upper()}"
            
            if existing_problem_id:
                # Duplicate found - update existing problem using raw SQL to avoid ORM column issues
                try:
                    # Get existing problem_code first (if any)
                    get_code_sql = text("SELECT problem_code FROM coding_problems WHERE id = :id")
                    code_result = db.execute(get_code_sql, {"id": existing_problem_id})
                    code_row = code_result.fetchone()
                    existing_code = code_row[0] if code_row and code_row[0] else problem_code
                    
                    # Prepare update values - convert complex types to JSON strings
                    import json as json_module
                    tags_json = json_module.dumps(tags) if tags else None
                    allowed_langs_json = json_module.dumps(allowed_languages) if allowed_languages else None
                    restricted_langs_json = json_module.dumps(restricted_languages) if restricted_languages else None
                    recommended_langs_json = json_module.dumps(recommended_languages) if recommended_languages else None
                    test_cases_json = json_module.dumps(test_cases) if test_cases else None
                    
                    # Build update SQL dynamically to handle NULL JSON values
                    update_fields = []
                    update_params = {"id": existing_problem_id}
                    
                    # Add fields only if they have values (COALESCE pattern)
                    if problem_data.get('input_format'):
                        update_fields.append("input_format = :input_format")
                        update_params["input_format"] = problem_data.get('input_format')
                    if problem_data.get('output_format'):
                        update_fields.append("output_format = :output_format")
                        update_params["output_format"] = problem_data.get('output_format')
                    if problem_data.get('difficulty'):
                        update_fields.append("difficulty = :difficulty")
                        update_params["difficulty"] = problem_data.get('difficulty')
                    if tags_json:
                        update_fields.append("tags = :tags::json")
                        update_params["tags"] = tags_json
                    if problem_data.get('constraints'):
                        update_fields.append("constraints = :constraints")
                        update_params["constraints"] = problem_data.get('constraints')
                    if problem_data.get('sample_input'):
                        update_fields.append("sample_input = :sample_input")
                        update_params["sample_input"] = problem_data.get('sample_input')
                    if problem_data.get('sample_output'):
                        update_fields.append("sample_output = :sample_output")
                        update_params["sample_output"] = problem_data.get('sample_output')
                    if year_value or year:
                        update_fields.append("year = :year")
                        update_params["year"] = year_value if year_value else year
                    if year_str_value:
                        update_fields.append("year_str = :year_str")
                        update_params["year_str"] = year_str_value
                    
                    # Always update these fields
                    update_fields.append("scope_type = :scope_type")
                    update_params["scope_type"] = scope_type
                    if college_id:
                        update_fields.append("college_id = :college_id")
                        update_params["college_id"] = college_id
                    if department:
                        update_fields.append("department = :department")
                        update_params["department"] = department
                    
                    if allowed_langs_json:
                        update_fields.append("allowed_languages = :allowed_languages::json")
                        update_params["allowed_languages"] = allowed_langs_json
                    if restricted_langs_json:
                        update_fields.append("restricted_languages = :restricted_languages::json")
                        update_params["restricted_languages"] = restricted_langs_json
                    if recommended_langs_json:
                        update_fields.append("recommended_languages = :recommended_languages::json")
                        update_params["recommended_languages"] = recommended_langs_json
                    
                    if problem_data.get('starter_code_python'):
                        update_fields.append("starter_code_python = :starter_code_python")
                        update_params["starter_code_python"] = problem_data.get('starter_code_python')
                    if problem_data.get('starter_code_c'):
                        update_fields.append("starter_code_c = :starter_code_c")
                        update_params["starter_code_c"] = problem_data.get('starter_code_c')
                    if problem_data.get('starter_code_cpp'):
                        update_fields.append("starter_code_cpp = :starter_code_cpp")
                        update_params["starter_code_cpp"] = problem_data.get('starter_code_cpp')
                    if problem_data.get('starter_code_java'):
                        update_fields.append("starter_code_java = :starter_code_java")
                        update_params["starter_code_java"] = problem_data.get('starter_code_java')
                    if problem_data.get('starter_code_javascript'):
                        update_fields.append("starter_code_javascript = :starter_code_javascript")
                        update_params["starter_code_javascript"] = problem_data.get('starter_code_javascript')
                    
                    update_fields.append("time_limit = :time_limit")
                    update_params["time_limit"] = time_limit
                    update_fields.append("memory_limit = :memory_limit")
                    update_params["memory_limit"] = memory_limit
                    
                    if test_cases_json:
                        update_fields.append("test_cases = :test_cases::json")
                        update_params["test_cases"] = test_cases_json
                    
                    is_active_val = problem_data.get('is_active', True) if isinstance(problem_data.get('is_active'), bool) else (str(problem_data.get('is_active', 'true')).lower() == 'true')
                    update_fields.append("is_active = :is_active")
                    update_params["is_active"] = is_active_val
                    
                    if expiry_date:
                        update_fields.append("expiry_date = :expiry_date")
                        update_params["expiry_date"] = expiry_date
                    
                    update_fields.append("problem_code = COALESCE(problem_code, :problem_code)")
                    update_params["problem_code"] = existing_code
                    update_fields.append("updated_at = NOW()")
                    
                    # Build and execute update SQL
                    update_sql = text(f"UPDATE coding_problems SET {', '.join(update_fields)} WHERE id = :id")
                    db.execute(update_sql, update_params)
                    db.flush()
                    
                    results["success"].append({
                        "index": idx,
                        "title": problem_data['title'],
                        "id": existing_problem_id,
                        "problem_code": existing_code,
                        "action": "updated",  # Indicates this was a duplicate that was updated
                        "duplicate_of": existing_problem_id
                    })
                    results["total"] += 1
                except Exception as update_error:
                    # If update fails, rollback and treat as new
                    db.rollback()
                    existing_problem_id = None
            else:
                # New problem - create it
                problem = CodingProblem(
                    title=title,
                    description=description,
                    input_format=problem_data.get('input_format'),
                    output_format=problem_data.get('output_format'),
                    difficulty=problem_data.get('difficulty'),
                    tags=tags,
                    constraints=problem_data.get('constraints'),
                    sample_input=problem_data.get('sample_input'),
                    sample_output=problem_data.get('sample_output'),
                    year=year_value if year_value else year,
                    allowed_languages=allowed_languages if allowed_languages else ["python", "c", "cpp", "java", "javascript"],
                    restricted_languages=restricted_languages,
                    recommended_languages=recommended_languages,
                    starter_code_python=problem_data.get('starter_code_python'),
                    starter_code_c=problem_data.get('starter_code_c'),
                    starter_code_cpp=problem_data.get('starter_code_cpp'),
                    starter_code_java=problem_data.get('starter_code_java'),
                    starter_code_javascript=problem_data.get('starter_code_javascript'),
                    time_limit=time_limit,
                    memory_limit=memory_limit,
                    test_cases=test_cases,
                    is_active=problem_data.get('is_active', True) if isinstance(problem_data.get('is_active'), bool) else (str(problem_data.get('is_active', 'true')).lower() == 'true'),
                    created_by=current_user_obj.id,
                    expiry_date=expiry_date,
                    problem_code=problem_code,  # Unique code for analytics
                    scope_type=scope_type,
                    college_id=college_id,
                    department=department,
                    section_id=None,  # Not used - year-based restriction only (no section restrictions)
                    year_str=year_str_value
                )
                db.add(problem)
                db.flush()
                
                # Store ID immediately after flush, before any potential rollback
                # This prevents "transaction is closed" errors
                problem_id = problem.id
                problem_code_value = getattr(problem, 'problem_code', None) or problem_code
                
                results["success"].append({
                    "index": idx,
                    "title": problem_data['title'],
                    "id": problem_id,
                    "problem_code": problem_code_value,
                    "action": "created"  # New problem created
                })
                results["total"] += 1
            
            # Commit savepoint if it exists
            if savepoint:
                try:
                    savepoint.commit()
                except Exception as savepoint_error:
                    # If savepoint commit fails, rollback it but don't fail the row
                    try:
                        savepoint.rollback()
                    except Exception:
                        pass
                    # The main transaction should still be valid since we stored the ID before this
            
        except Exception as e:
            # Rollback savepoint on error
            if savepoint:
                try:
                    savepoint.rollback()
                except Exception:
                    pass
            # Also rollback the main transaction to clear any aborted state
            try:
                db.rollback()
            except Exception:
                pass
            
            results["failed"].append({
                "index": idx,
                "title": problem_data.get('title', 'N/A'),
                "error": str(e)
            })
    
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.get("/template/students")
async def download_student_template(
    format: str = "csv",  # csv or xlsx
    college_id: Optional[int] = None,  # Optional college_id for context
    current_user: User = Depends(get_current_super_admin)
):
    """Download CSV or Excel template for student bulk upload
    
    Required fields:
    - email: Student email (required)
    - password: Password (optional, defaults to roll_number or user.id)
    - full_name: Student full name
    - branch_id: Branch ID (required - e.g., "CSE001", "ECE001")
    - section: Section name (e.g., "A", "B", "C") - will auto-link when created by college admin
    - roll_number: Student roll number (can be used for login)
    - present_year: Academic year (1, 2, 3, 4, or 5)
    - college_id: College ID (optional if provided in query param - will default to query param value)
    """
    from fastapi.responses import StreamingResponse
    
    if format.lower() == "xlsx":
        # Excel format
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Headers - branch_id is required, section is specified by name only (A, B, C, etc.)
        # Section will auto-link when created by college admin
        headers = ["email", "password", "full_name", "branch_id", "section", "roll_number", "present_year"]
        ws.append(headers)
        
        # Sample data rows - branch_id is required, section by name only
        ws.append(["student1@example.com", "Password123", "John Doe", "CSE001", "A", "20CS001", "1"])
        ws.append(["student2@example.com", "Password123", "Jane Smith", "CSE001", "B", "20CS002", "1"])
        ws.append(["student3@example.com", "Password123", "Jane Doe", "ECE001", "A", "20EC001", "2"])
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.getvalue()
        excel_buffer.close()
        
        return StreamingResponse(
            iter([excel_content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=student_upload_template.xlsx",
                "Content-Length": str(len(excel_content))
            }
        )
    else:
        # CSV format (default) - branch_id is required, section by name only
        template = "email,password,full_name,branch_id,section,roll_number,present_year\n"
        template += "student1@example.com,Password123,John Doe,CSE001,A,20CS001,1\n"
        template += "student2@example.com,Password123,Jane Smith,CSE001,B,20CS002,1\n"
        template += "student3@example.com,Password123,Jane Doe,ECE001,A,20EC001,2\n"
        
        return StreamingResponse(
            iter([template.encode('utf-8')]),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=student_upload_template.csv",
                "Content-Length": str(len(template.encode('utf-8')))
            }
        )


@router.get("/templates/quiz")
async def download_quiz_template(
    format: str = "xlsx",  # xlsx, csv, or json
    current_user_tuple = Depends(get_current_content_creator)
):
    """Download template for quiz bulk upload in Excel, CSV, or JSON format"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    if format == "json":
        template = [
            {
                "title": "Sample Quiz",
                "description": "Quiz description",
                "subject": "Computer Science",
                "duration_minutes": 30,
                "total_marks": 100,
                "scope_type": "section",
                "section_id": 1,
                "year": "1st",
                "expiry_date": "2024-12-31T23:59:59",
                "is_active": True
            }
        ]
        return Response(
            content=json.dumps(template, indent=2),
            media_type="application/json",
            headers={
                "Content-Disposition": "attachment; filename=quiz_upload_template.json"
            }
        )
    elif format == "csv":
        csv_content = "title,description,subject,duration_minutes,total_marks,scope_type,section_id,year,expiry_date,is_active\n"
        csv_content += "Sample Quiz,Quiz description,Computer Science,30,100,section,1,1st,2024-12-31T23:59:59,true\n"
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=quiz_upload_template.csv"
            }
        )
    else:  # xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Quiz Template"
        
        # Headers
        headers = ["title", "description", "subject", "duration_minutes", "total_marks", "scope_type", "section_id", "year", "expiry_date", "is_active"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Sample row
        sample_data = ["Sample Quiz", "Quiz description", "Computer Science", 30, 100, "section", 1, "1st", "2024-12-31T23:59:59", True]
        for col_idx, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=quiz_upload_template.xlsx"
            }
        )


@router.get("/templates/coding-problem")
async def download_coding_problem_template(
    format: str = "xlsx",  # xlsx, csv, or json
    current_user_tuple = Depends(get_current_content_creator)
):
    """Download template for coding problem bulk upload in Excel, CSV, or JSON format"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    if format == "json":
        template = [
            {
                "title": "Sample Problem: Sum of Two Numbers",
                "description": "Write a function to add two numbers and return the result.",
                "input_format": "Two integers a and b",
                "output_format": "Return the sum of a and b",
                "constraints": "1 ≤ a, b ≤ 1000",
                "sample_input": "5\n10",
                "sample_output": "15",
                "difficulty": "Easy",
                "tags": "array,math",
                "year": 1,
                "allowed_languages": "python,c,cpp,java,javascript",
                "restricted_languages": "",
                "recommended_languages": "",
                "starter_code_python": "def add(a, b):\n    # Your code here\n    return a + b",
                "starter_code_c": "#include <stdio.h>\n\nint main() {\n    int a, b;\n    scanf(\"%d %d\", &a, &b);\n    printf(\"%d\\n\", a + b);\n    return 0;\n}",
                "starter_code_cpp": "#include <iostream>\nusing namespace std;\n\nint main() {\n    int a, b;\n    cin >> a >> b;\n    cout << a + b << endl;\n    return 0;\n}",
                "starter_code_java": "import java.util.Scanner;\n\npublic class Solution {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int a = sc.nextInt();\n        int b = sc.nextInt();\n        System.out.println(a + b);\n    }\n}",
                "starter_code_javascript": "function add(a, b) {\n    // Your code here\n    return a + b;\n}",
                "time_limit": 5,
                "memory_limit": 256,
                "test_cases": "[{\"stdin\": \"5\\n10\", \"expected_output\": \"15\", \"is_public\": true}, {\"stdin\": \"100\\n200\", \"expected_output\": \"300\", \"is_public\": false}]",
                "expiry_date": "2024-12-31T23:59:59",
                "is_active": True
            }
        ]
        return Response(
            content=json.dumps(template, indent=2),
            media_type="application/json",
            headers={
                "Content-Disposition": "attachment; filename=coding_problem_upload_template.json"
            }
        )
    elif format == "csv":
        csv_content = "title,description,input_format,output_format,constraints,sample_input,sample_output,difficulty,tags,year,allowed_languages,restricted_languages,recommended_languages,starter_code_python,starter_code_c,starter_code_cpp,starter_code_java,starter_code_javascript,time_limit,memory_limit,test_cases,expiry_date,is_active\n"
        csv_content += "Sample Problem: Sum of Two Numbers,\"Write a function to add two numbers\",\"Two integers a and b\",\"Return the sum\",\"1 ≤ a, b ≤ 1000\",\"5\\n10\",15,Easy,\"array,math\",1,\"python,c,cpp,java,javascript\",\"\",\"\",\"def add(a, b):\\n    return a + b\",\"#include <stdio.h>\\nint main() { int a,b; scanf(\\\"%d %d\\\", &a, &b); printf(\\\"%d\\n\\\", a+b); return 0; }\",\"#include <iostream>\\nusing namespace std;\\nint main() { int a,b; cin>>a>>b; cout<<a+b<<endl; return 0; }\",\"import java.util.Scanner;\\npublic class Solution { public static void main(String[] args) { Scanner sc = new Scanner(System.in); int a=sc.nextInt(); int b=sc.nextInt(); System.out.println(a+b); } }\",\"function add(a, b) { return a + b; }\",5,256,\"[{\\\"stdin\\\": \\\"5\\\\n10\\\", \\\"expected_output\\\": \\\"15\\\", \\\"is_public\\\": true}]\",2024-12-31T23:59:59,true\n"
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=coding_problem_upload_template.csv"
            }
        )
    else:  # xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Coding Problem Template"
        
        # Headers - All required fields (section removed - use year only for restrictions)
        headers = [
            "title", "description", "input_format", "output_format", "constraints",
            "sample_input", "sample_output", "difficulty", "tags", "year",
            "allowed_languages", "restricted_languages", "recommended_languages",
            "starter_code_python", "starter_code_c", "starter_code_cpp", "starter_code_java", "starter_code_javascript",
            "time_limit", "memory_limit", "test_cases", "expiry_date", "is_active"
        ]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Sample row with all fields
        sample_data = [
            "Sample Problem: Sum of Two Numbers",
            "Write a function to add two numbers and return the result.",
            "Two integers a and b",
            "Return the sum of a and b",
            "1 ≤ a, b ≤ 1000",
            "5\n10",
            "15",
            "Easy",
            "array,math",
            1,  # year (1-4, leave empty for all years)
            "python,c,cpp,java,javascript",
            "",
            "",
            "def add(a, b):\n    # Your code here\n    return a + b",
            "#include <stdio.h>\n\nint main() {\n    int a, b;\n    scanf(\"%d %d\", &a, &b);\n    printf(\"%d\\n\", a + b);\n    return 0;\n}",
            "#include <iostream>\nusing namespace std;\n\nint main() {\n    int a, b;\n    cin >> a >> b;\n    cout << a + b << endl;\n    return 0;\n}",
            "import java.util.Scanner;\n\npublic class Solution {\n    public static void main(String[] args) {\n        Scanner sc = new Scanner(System.in);\n        int a = sc.nextInt();\n        int b = sc.nextInt();\n        System.out.println(a + b);\n    }\n}",
            "function add(a, b) {\n    // Your code here\n    return a + b;\n}",
            5,
            256,
            '[{"stdin": "5\\n10", "expected_output": "15", "is_public": true}, {"stdin": "100\\n200", "expected_output": "300", "is_public": false}]',
            "2024-12-31T23:59:59",
            True
        ]
        for col_idx, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add instructions sheet
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            ["CODING PROBLEM TEMPLATE - INSTRUCTIONS"],
            [""],
            ["REQUIRED FIELDS:"],
            ["• title - Problem title (required)"],
            ["• description - Problem description (required)"],
            ["• allowed_languages - Comma-separated: python,c,cpp,java,javascript (required)"],
            ["• time_limit - Time limit in seconds (required)"],
            ["• memory_limit - Memory limit in MB (required)"],
            [""],
            ["OPTIONAL FIELDS:"],
            ["• year - Year (1-4) - Leave empty for all years"],
            ["  Note: Year-based restriction - problems are visible to students in that year and below"],
            ["  - Year 1: Visible to 1st year students only"],
            ["  - Year 2: Visible to 1st and 2nd year students"],
            ["  - Year 3: Visible to 1st, 2nd, and 3rd year students"],
            ["  - Year 4: Visible to all students"],
            ["  - Empty: Visible to all students (no year restriction)"],
            ["• input_format - Describe input format"],
            ["• output_format - Describe output format"],
            ["• constraints - Problem constraints"],
            ["• sample_input - Sample input"],
            ["• sample_output - Sample output"],
            ["• difficulty - Easy/Medium/Hard"],
            ["• tags - Comma-separated tags"],
            ["• restricted_languages - Languages that MUST be used (comma-separated)"],
            ["• recommended_languages - Recommended languages (comma-separated)"],
            ["• starter_code_python - Python starter code"],
            ["• starter_code_c - C starter code"],
            ["• starter_code_cpp - C++ starter code"],
            ["• starter_code_java - Java starter code"],
            ["• starter_code_javascript - JavaScript starter code"],
            ["• test_cases - JSON array: [{\"stdin\": \"input\", \"expected_output\": \"output\", \"is_public\": true}]"],
            ["• expiry_date - ISO format: 2024-12-31T23:59:59"],
            ["• is_active - true/false"],
            [""],
            ["TEST CASES FORMAT:"],
            ["JSON array with objects containing:"],
            ["  - stdin: Input string"],
            ["  - expected_output: Expected output string"],
            ["  - is_public: true/false (whether students can see this test)"],
            [""],
            ["EXAMPLE:"],
            ['[{"stdin": "5\\n10", "expected_output": "15", "is_public": true}]'],
            [""],
            ["LANGUAGE ARRAYS:"],
            ["• Use comma-separated values: python,c,cpp,java,javascript"],
            ["• Or JSON array format: [\"python\", \"c\", \"cpp\"]"],
        ]
        for row_idx, instruction in enumerate(instructions, 1):
            ws2.cell(row=row_idx, column=1, value=instruction[0])
            if row_idx == 1:
                ws2.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
        
        ws2.column_dimensions['A'].width = 80
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=coding_problem_upload_template.xlsx"
            }
        )


@router.post("/hod-faculty")
async def bulk_upload_hod_faculty(
    file: UploadFile = File(...),
    college_id: Optional[int] = Query(None, description="College ID - optional, auto-detected from user context if not provided"),
    current_user_tuple = Depends(get_current_admin_or_super),
    db: Session = Depends(get_db)
):
    """Bulk upload HOD and Faculty from CSV or Excel file.
    
    College admin can upload HOD and Faculty for their college (college_id auto-detected).
    Super admin can upload for any college (college_id required if not provided).
    
    Supported formats: .csv, .xlsx, .xls
    
    Format (headers required):
    email,full_name,role,branch_id,user_id
    
    Note: college_id is auto-detected from user context. Super admins can optionally provide it.
    
    Example:
    email,full_name,role,branch_id,user_id
    hod.cs@sbit.edu,Dr. John Doe,hod,CSE001,HOD001
    faculty.cs@sbit.edu,Dr. Jane Smith,faculty,CSE001,FAC001
    
    Note: password will be auto-generated from user_id (in caps) or user.id (in caps)
    """
    current_user, is_super_admin = current_user_tuple
    
    # Auto-detect college_id if not provided
    if not college_id:
        if is_super_admin:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="College ID is required for super admin uploads. Provide it as query parameter: ?college_id=X"
            )
        else:
            # Auto-detect from user context
            auto_college_id = get_admin_college_id(current_user, db)
            if not auto_college_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Could not determine college ID. Admin must be associated with a college."
                )
            college_id = auto_college_id
    
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    # Read file content
    contents = await file.read()
    
    # Parse based on file type
    if file_ext == 'csv':
        csv_content = contents.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))
    else:
        # Excel file
        excel_file = io.BytesIO(contents)
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Get headers from first row and normalize to lowercase
        headers = [cell.value for cell in sheet[1]]
        if not headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must have headers in the first row"
            )
        
        # Normalize headers to lowercase for case-insensitive matching
        normalized_headers = [str(h).strip().lower() if h else None for h in headers]
        
        # Convert Excel rows to CSV-like dict format
        csv_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for idx, header in enumerate(normalized_headers):
                if header:
                    # Safely access row[idx] - handle cases where row has fewer columns than headers
                    if idx < len(row):
                        row_dict[header] = str(row[idx]).strip() if row[idx] is not None else ""
                    else:
                        row_dict[header] = ""
            csv_rows.append(row_dict)
        
        csv_reader = csv_rows if csv_rows else []
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    # Handle both CSV reader (iterator) and Excel rows (list)
    rows_to_process = list(csv_reader) if isinstance(csv_reader, list) else csv_reader
    
    for row_num, row in enumerate(rows_to_process, start=2):  # Start at 2 (row 1 is header)
        try:
            email = row.get('email', '').strip().lower()
            full_name = row.get('full_name', '').strip()
            branch_id_str = row.get('branch_id', '').strip()
            role = row.get('role', '').strip().lower()
            user_id = row.get('user_id', '').strip()
            password = row.get('password', '').strip()
            
            # College ID is required from query parameter and applied to all staff
            final_college_id = college_id
            
            if not email:
                results["failed"].append({
                    "row": row_num,
                    "email": email or "N/A",
                    "error": "Email is required"
                })
                continue
            
            if role not in ['hod', 'faculty']:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": f"Role must be 'hod' or 'faculty', got '{role}'"
                })
                continue
            
            # College ID is required from query parameter
            
            # Department is required for faculty and HOD - resolved from branch_id only
            if role in ['faculty', 'hod']:
                if not branch_id_str:
                    results["failed"].append({
                        "row": row_num,
                        "email": email,
                        "error": "Branch ID is required for faculty and HOD roles (e.g., 'CSE001', 'ECE001')"
                    })
                    continue
            
            # Resolve department_id from branch_id only
            from app.models.academic import Department
            department_id = None
            department_name = None
            
            if role in ['faculty', 'hod']:
                # Find department by branch_id
                dept = db.query(Department).filter(
                    Department.branch_id == branch_id_str,
                    Department.college_id == final_college_id,
                    Department.is_active == True
                ).first()
                if dept:
                    department_id = dept.id
                    department_name = dept.name
                else:
                    results["failed"].append({
                        "row": row_num,
                        "email": email,
                        "error": f"Branch ID '{branch_id_str}' not found for this college"
                })
                continue
            
            # Check if user already exists
            existing = db.query(User).filter(User.email == email).first()
            if existing:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": "User already exists"
                })
                continue
            
            # Determine password: use provided password, or user_id, or user.id as fallback
            if password:
                # Use provided password from upload file
                final_password = password
            elif user_id:
                # Use user_id as password (uppercase)
                final_password = str(user_id).upper()
            else:
                # Will use user.id after creation
                final_password = "TEMP_PASSWORD_PLACEHOLDER"
            
            # Create user
            user = User(
                email=email,
                password_hash=get_password_hash(final_password),
                is_active="true",
                is_verified="true"
            )
            db.add(user)
            db.flush()  # Get user.id
            
            # If password was placeholder, update it with user.id in caps
            if final_password == "TEMP_PASSWORD_PLACEHOLDER":
                user.password_hash = get_password_hash(str(user.id).upper())
            
            # Create profile with department_id
            profile = Profile(
                user_id=user.id,
                email=email,
                full_name=full_name if full_name else None,
                college_id=final_college_id,
                department=department_name,  # Department name from branch_id
                department_id=department_id,  # Link to Department
            )
            db.add(profile)
            
            # Handle HOD assignment - ensure only one HOD per department
            if role == 'hod' and department_id:
                hod_department = db.query(Department).filter(Department.id == department_id).first()
                if hod_department:
                    # If department already has an HOD, remove the old HOD's role
                    if hod_department.hod_id:
                        old_hod_id = hod_department.hod_id
                        # Remove old HOD role
                        old_hod_role = db.query(UserRole).filter(
                            UserRole.user_id == old_hod_id,
                            UserRole.role == RoleEnum.HOD
                        ).first()
                        if old_hod_role:
                            db.delete(old_hod_role)
                        # Clear department's hod_id
                        hod_department.hod_id = None
                    
                    # Assign new HOD to department
                    hod_department.hod_id = user.id
                    db.add(hod_department)
                    db.flush()  # Flush to ensure department update is saved
            
            # Create role (after HOD assignment to ensure department is set)
            role_enum = RoleEnum.HOD if role == 'hod' else RoleEnum.FACULTY
            user_role = UserRole(
                user_id=user.id,
                role=role_enum,
                college_id=final_college_id
            )
            db.add(user_role)
            
            results["success"].append({
                "row": row_num,
                "email": email,
                "name": full_name,
                "role": role
            })
            results["total"] += 1
            
        except Exception as e:
            results["failed"].append({
                "row": row_num,
                "email": row.get('email', 'N/A'),
                "error": str(e)
            })
    
    # Commit all successful inserts
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.post("/staff")
async def bulk_upload_staff(
    file: UploadFile = File(...),
    college_id: Optional[int] = Query(None, description="College ID - optional, auto-detected from user context if not provided"),
    current_user_tuple = Depends(get_current_admin_or_super),
    db: Session = Depends(get_db)
):
    """Bulk upload Faculty, HOD, and College Admin from CSV or Excel file.
    
    College admin can upload staff for their college (college_id auto-detected).
    Super admin can upload for any college (college_id required if not provided).
    
    Supported formats: .csv, .xlsx, .xls
    
    Format (headers required):
    email,full_name,role,branch_id,user_id,password
    
    Note: college_id is auto-detected from user context. Super admins can optionally provide it.
    
    Fields:
    - email: User email address (required)
    - full_name: Full name
    - role: 'faculty', 'hod', or 'admin' (admin requires college_id)
    - branch_id: Branch ID (recommended - e.g., "CSE001", "ECE001") - most user-friendly
    - user_id: Staff ID (optional, used for password if password not provided)
    - password: Password (optional, will use user_id or user.id if not provided)
    
    Note:
    - Each department can have only one HOD. If uploading a new HOD for a department that already has one, the old HOD will be replaced.
    - Faculty and HOD must be assigned to a department.
    - College admin does not require a department.
    
    Example:
    email,full_name,role,branch_id,user_id,password
    faculty.cs@example.com,Dr. Jane Smith,faculty,CSE001,FAC001,MyPassword123
    hod.cs@example.com,Dr. John Doe,hod,CSE001,HOD001,
    admin@example.com,Admin User,admin,,ADMIN001,AdminPass123
    """
    from app.models.academic import Department
    
    current_user, is_super_admin = current_user_tuple
    
    # ============================================================================
    # STEP 1: Auto-detect college_id if not provided
    # ============================================================================
    if not college_id:
        if is_super_admin:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="College ID is required for super admin uploads. Provide it as query parameter: ?college_id=X"
            )
        else:
            auto_college_id = get_admin_college_id(current_user, db)
            if not auto_college_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Could not determine college ID. Admin must be associated with a college."
                )
            college_id = auto_college_id
    
    # ============================================================================
    # STEP 2: Validate file type
    # ============================================================================
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    # ============================================================================
    # STEP 3: Read and parse file
    # ============================================================================
    contents = await file.read()
    rows_to_process = []
    skipped_count = 0
    
    if file_ext == 'csv':
        # Parse CSV file
        csv_content = contents.decode('utf-8')
        csv_reader_raw = csv.DictReader(io.StringIO(csv_content))
        
        for row in csv_reader_raw:
            # Normalize all keys to lowercase
            normalized_row = {}
            for k, v in row.items():
                if k:
                    normalized_row[str(k).lower().strip()] = str(v).strip() if v is not None else ""
            
            # Check if row has any data
            has_data = any(str(v).strip() for v in normalized_row.values() if v and str(v).strip())
            if has_data:
                rows_to_process.append(normalized_row)
            else:
                skipped_count += 1
        
        print(f"[Bulk Upload Staff] CSV: Found {len(rows_to_process)} rows with data, Skipped {skipped_count} empty rows")
        
    else:
        # Parse Excel file
        excel_file = io.BytesIO(contents)
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        
        # Find the correct sheet
        if "Staff" in workbook.sheetnames:
            sheet = workbook["Staff"]
            print(f"[Bulk Upload Staff] Using 'Staff' sheet")
        else:
            sheet = workbook.active
            print(f"[Bulk Upload Staff] Using active sheet: '{sheet.title}'")
        
        # Get and normalize headers
        headers_raw = [cell.value for cell in sheet[1]]
        if not headers_raw or all(h is None or str(h).strip() == "" for h in headers_raw):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must have headers in the first row"
            )
        
        headers = [str(h).strip().lower() if h else "" for h in headers_raw]
        print(f"[Bulk Upload Staff] Excel Headers: {headers}")
        
        # Parse all rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for idx, header in enumerate(headers):
                if header:
                    if idx < len(row):
                        cell_value = row[idx]
                        row_dict[header] = str(cell_value).strip() if cell_value is not None else ""
                    else:
                        row_dict[header] = ""
            
            # Check if row has any data (explicit check for None, empty strings, whitespace)
            has_data = False
            for v in row_dict.values():
                if v is not None and str(v).strip():
                    has_data = True
                    break
            
            if has_data:
                rows_to_process.append(row_dict)
                # Debug logging for first few rows
                if row_idx <= 5:
                    print(f"[Bulk Upload Staff] Row {row_idx}: email='{row_dict.get('email', '')}', role='{row_dict.get('role', '')}', branch='{row_dict.get('branch_id', '')}'")
            else:
                skipped_count += 1
        
        print(f"[Bulk Upload Staff] Excel: Found {len(rows_to_process)} rows with data, Skipped {skipped_count} empty rows")
    
    # ============================================================================
    # STEP 4: Validate we have rows to process
    # ============================================================================
    if not rows_to_process:
        return {
            "message": "No data rows found in file. Please ensure the file has data rows (not just headers).",
            "success_count": 0,
            "failed_count": 0,
            "success": [],
            "failed": [{"row": 0, "email": "N/A", "error": "No data rows found in file"}]
        }
    
    print(f"[Bulk Upload Staff] Processing {len(rows_to_process)} rows for college_id={college_id}")
    
    # ============================================================================
    # STEP 5: Process each row
    # ============================================================================
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    for idx, row in enumerate(rows_to_process):
        row_num = idx + 2  # Excel row number (row 1 is header)
        
        try:
            # Extract and normalize fields
            email = str(row.get('email', '')).strip().lower()
            full_name = str(row.get('full_name', '')).strip()
            role = str(row.get('role', '')).strip().lower()
            branch_id_str = str(row.get('branch_id', '')).strip()
            user_id = str(row.get('user_id', '')).strip()
            password = str(row.get('password', '')).strip()
            
            # Validate required fields
            if not email:
                results["failed"].append({
                    "row": row_num,
                    "email": "N/A",
                    "error": "Email is required"
                })
                continue
            
            if role not in ['faculty', 'hod', 'admin']:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": f"Role must be 'faculty', 'hod', or 'admin', got '{role}'"
                })
                continue
            
            # Resolve department for faculty/HOD
            department_id = None
            department_name = None
            
            if role in ['faculty', 'hod']:
                if not branch_id_str:
                    results["failed"].append({
                        "row": row_num,
                        "email": email,
                        "error": "Branch ID is required for faculty and HOD roles (e.g., 'CSE001', 'ECE001')"
                    })
                    continue
                
                dept = db.query(Department).filter(
                    Department.branch_id == branch_id_str,
                    Department.college_id == college_id,
                    Department.is_active == True
                ).first()
                
                if not dept:
                    results["failed"].append({
                        "row": row_num,
                        "email": email,
                        "error": f"Branch ID '{branch_id_str}' not found for this college"
                    })
                    continue
                
                department_id = dept.id
                department_name = dept.name
            
            # Check if user already exists
            existing = db.query(User).filter(User.email == email).first()
            if existing:
                results["failed"].append({
                    "row": row_num,
                    "email": email,
                    "error": "User already exists"
                })
                continue
            
            # Determine password
            if password:
                final_password = password
            elif user_id:
                final_password = str(user_id).upper()
            else:
                final_password = "TEMP_PASSWORD_PLACEHOLDER"
            
            # Create user
            user = User(
                email=email,
                password_hash=get_password_hash(final_password),
                is_active="true",
                is_verified="true"
            )
            db.add(user)
            db.flush()
            
            # Update password if placeholder was used
            if final_password == "TEMP_PASSWORD_PLACEHOLDER":
                user.password_hash = get_password_hash(str(user.id).upper())
            
            # Create profile
            profile = Profile(
                user_id=user.id,
                email=email,
                full_name=full_name if full_name else None,
                college_id=college_id,
                department=department_name,
                department_id=department_id,
            )
            db.add(profile)
            
            # Handle HOD assignment (only one HOD per department)
            if role == 'hod' and department_id:
                hod_department = db.query(Department).filter(Department.id == department_id).first()
                if hod_department:
                    # Remove old HOD if exists
                    if hod_department.hod_id:
                        old_hod_role = db.query(UserRole).filter(
                            UserRole.user_id == hod_department.hod_id,
                            UserRole.role == RoleEnum.HOD
                        ).first()
                        if old_hod_role:
                            db.delete(old_hod_role)
                        hod_department.hod_id = None
                    
                    # Assign new HOD
                    hod_department.hod_id = user.id
                    db.add(hod_department)
                    db.flush()
            
            # Create user role
            role_enum_map = {
                'faculty': RoleEnum.FACULTY,
                'hod': RoleEnum.HOD,
                'admin': RoleEnum.ADMIN
            }
            user_role = UserRole(
                user_id=user.id,
                role=role_enum_map[role],
                college_id=college_id
            )
            db.add(user_role)
            
            results["success"].append({
                "row": row_num,
                "email": email,
                "name": full_name,
                "role": role
            })
            results["total"] += 1
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            email_val = row.get('email', 'N/A') if isinstance(row, dict) else 'N/A'
            print(f"[Bulk Upload Staff] ERROR Row {row_num} (email: {email_val}): {str(e)}")
            print(f"[Bulk Upload Staff] Traceback: {error_details}")
            results["failed"].append({
                "row": row_num,
                "email": email_val,
                "error": str(e)
            })
    
    # ============================================================================
    # STEP 6: Commit all successful inserts
    # ============================================================================
    if results["success"]:
        try:
            db.commit()
            print(f"[Bulk Upload Staff] ✅ Successfully committed {len(results['success'])} users")
        except Exception as e:
            db.rollback()
            import traceback
            error_details = traceback.format_exc()
            print(f"[Bulk Upload Staff] ❌ Commit failed: {str(e)}")
            print(f"[Bulk Upload Staff] Traceback: {error_details}")
            return {
                "message": f"Failed to save users: {str(e)}",
                "success_count": 0,
                "failed_count": len(results["failed"]) + len(results["success"]),
                "success": [],
                "failed": results["failed"] + [{"row": 0, "email": "N/A", "error": f"Database commit failed: {str(e)}"}]
            }
    else:
        print(f"[Bulk Upload Staff] ⚠️  No successful rows to commit. Failed: {len(results['failed'])}")
    
    # ============================================================================
    # STEP 7: Return results
    # ============================================================================
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.get("/template/hod-faculty")
async def download_hod_faculty_template(
    format: str = "csv",  # csv or xlsx
    current_user_tuple = Depends(get_current_admin_or_super)
):
    """Download CSV or Excel template for HOD/Faculty bulk upload"""
    from fastapi.responses import StreamingResponse
    
    if format.lower() == "xlsx":
        # Excel format
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "HOD Faculty"
        
        # Headers - branch_id is required for faculty and HOD
        # Note: college_id is provided via query parameter, not in the file
        headers = ["email", "full_name", "role", "branch_id", "user_id", "password"]
        ws.append(headers)
        
        # Sample data rows - branch_id is required for faculty and HOD
        # Password is optional - if not provided, will use user_id or user.id
        ws.append(["hod.cs@sbit.edu", "Dr. John Doe", "hod", "CSE001", "HOD001", ""])
        ws.append(["faculty.cs@sbit.edu", "Dr. Jane Smith", "faculty", "CSE001", "FAC001", "MyPassword123"])
        ws.append(["faculty2.cs@sbit.edu", "Dr. Jane Doe", "faculty", "ECE001", "FAC002", ""])
        ws.append(["admin@example.com", "Admin User", "admin", "", "ADMIN001", "AdminPass123"])
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.getvalue()
        excel_buffer.close()
        
        return StreamingResponse(
            iter([excel_content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=hod_faculty_upload_template.xlsx",
                "Content-Length": str(len(excel_content))
            }
        )
    else:
        # CSV format (default) - branch_id is required for faculty and HOD
        # Note: college_id is provided via query parameter, not in the file
        # Password is optional - if not provided, will use user_id or user.id
        template = "email,full_name,role,branch_id,user_id,password\n"
        template += "hod.cs@sbit.edu,Dr. John Doe,hod,CSE001,HOD001,\n"
        template += "faculty.cs@sbit.edu,Dr. Jane Smith,faculty,CSE001,FAC001,MyPassword123\n"
        template += "faculty2.cs@sbit.edu,Dr. Jane Doe,faculty,ECE001,FAC002,\n"
        template += "admin@example.com,Admin User,admin,,ADMIN001,AdminPass123\n"
        
        return StreamingResponse(
            iter([template.encode('utf-8')]),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=hod_faculty_upload_template.csv",
                "Content-Length": str(len(template.encode('utf-8')))
            }
        )


@router.get("/template/staff")
async def download_staff_template(
    format: str = "csv",  # csv or xlsx
    current_user: User = Depends(get_current_super_admin)
):
    """Download CSV or Excel template for Staff (Faculty/HOD/Admin) bulk upload (Super Admin only)
    
    Enhanced Excel format with:
    - Multiple sheets for different roles
    - Data validation dropdowns
    - Instructions sheet
    - Better formatting
    """
    from fastapi.responses import StreamingResponse
    
    if format.lower() == "xlsx":
        # Enhanced Excel format with multiple sheets
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Instructions Sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        ws_instructions.merge_cells('A1:D1')
        ws_instructions['A1'] = "Staff Bulk Upload Template - Instructions"
        ws_instructions['A1'].font = Font(bold=True, size=14)
        ws_instructions['A1'].alignment = Alignment(horizontal='center')
        
        instructions = [
            "",
            "This template allows you to bulk upload Faculty, HOD, and College Admin users.",
            "",
            "FIELD DESCRIPTIONS:",
            "• email: User email address (required, unique)",
            "• full_name: Full name of the user (required)",
            "• role: Must be 'faculty', 'hod', or 'admin' (required)",
            "• branch_id: Branch ID (required for faculty and HOD - e.g., 'CSE001', 'ECE001')",
            "  Note: Department will be automatically resolved from branch_id",
            "• college_id: College ID number (required for admin, optional for others)",
            "• user_id: Staff ID (optional, used for password if password not provided)",
            "• password: Password (optional)",
            "  - If provided, will be used as the user's password",
            "  - If not provided, will use user_id (uppercase) as password",
            "  - If neither password nor user_id provided, will use user.id (uppercase)",
            "  - Users can change their password after login via Profile page",
            "",
            "IMPORTANT NOTES:",
            "• Each department can have only ONE HOD. Uploading a new HOD will replace the existing one.",
            "• Faculty and HOD must be assigned to a department (via branch_id).",
            "• College admin does not require a branch_id.",
            "• Workflow: First create departments/branches, then add faculty/HOD, then add students.",
            "",
            "EXAMPLES:",
            "• Faculty: faculty.cs@example.com,Dr. Jane Smith,faculty,CSE001,1,FAC001",
            "• HOD: hod.cs@example.com,Dr. John Doe,hod,CSE001,1,HOD001",
            "• Admin: admin@example.com,Admin User,admin,,1,ADMIN001",
            "",
            "NOTES:",
            "• Password field is optional:",
            "  - If provided, will be used as the user's password",
            "  - If not provided, will use user_id (uppercase) as password",
            "  - If neither password nor user_id provided, will use user.id (uppercase)",
            "  - Users can change their password after login via Profile page",
            "• Use branch_id for all department linking",
            "• Leave fields empty if not applicable",
            "• Use the 'Staff' sheet below to enter your data",
        ]
        
        for idx, instruction in enumerate(instructions, start=2):
            ws_instructions[f'A{idx}'] = instruction
            ws_instructions[f'A{idx}'].alignment = Alignment(wrap_text=True)
        
        # Adjust column widths
        ws_instructions.column_dimensions['A'].width = 80
        
        # Staff Data Sheet
        ws_staff = wb.create_sheet("Staff", 1)
        
        # Headers with formatting - branch_id is required for faculty and HOD
        # Password is optional - if not provided, will use user_id or user.id
        headers = ["email", "full_name", "role", "branch_id", "user_id", "password"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_staff.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data validation for role column
        role_validation = DataValidation(type="list", formula1='"faculty,hod,admin"', allow_blank=False)
        role_validation.error = "Please select a valid role: faculty, hod, or admin"
        role_validation.errorTitle = "Invalid Role"
        ws_staff.add_data_validation(role_validation)
        role_validation.add(f"C2:C1000")  # Apply to role column
        
        # Sample data rows - branch_id is required for faculty and HOD
        sample_data = [
            ["faculty.cs@example.com", "Dr. Jane Smith", "faculty", "CSE001", "FAC001", "MyPassword123"],
            ["hod.cs@example.com", "Dr. John Doe", "hod", "CSE001", "HOD001", ""],
            ["faculty2.cs@example.com", "Dr. Jane Doe", "faculty", "ECE001", "FAC002", "FacultyPass456"],
            ["admin@example.com", "Admin User", "admin", "", "ADMIN001", "AdminPass123"]
        ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_staff.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 3:  # Role column
                    role_validation.add(cell)
        
        # Adjust column widths
        column_widths = [25, 20, 12, 20, 12, 12, 15, 15, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            ws_staff.column_dimensions[chr(64 + col_idx)].width = width
        
        # Freeze header row
        ws_staff.freeze_panes = "A2"
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.getvalue()
        excel_buffer.close()
        
        return StreamingResponse(
            iter([excel_content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=staff_upload_template_enhanced.xlsx",
                "Content-Length": str(len(excel_content))
            }
        )
    else:
        # CSV format (default) - branch_id is required for faculty and HOD
        # Note: college_id is provided via query parameter, not in the file
        # Password is optional - if not provided, will use user_id or user.id
        template = "email,full_name,role,branch_id,user_id,password\n"
        template += "faculty.cs@example.com,Dr. Jane Smith,faculty,CSE001,FAC001,MyPassword123\n"
        template += "hod.cs@example.com,Dr. John Doe,hod,CSE001,HOD001,\n"
        template += "faculty2.cs@example.com,Dr. Jane Doe,faculty,ECE001,FAC002,FacultyPass456\n"
        template += "admin@example.com,Admin User,admin,,ADMIN001,AdminPass123\n"
        
        return StreamingResponse(
            iter([template.encode('utf-8')]),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=staff_upload_template.csv",
                "Content-Length": str(len(template.encode('utf-8')))
            }
        )


@router.get("/template/coding-problems")
async def download_coding_problem_template(
    current_user: User = Depends(get_current_super_admin)
):
    """Download JSON template for coding problem bulk upload"""
    from fastapi.responses import Response
    
    template = {
        "problems": [
            {
                "title": "Two Sum",
                "description": "Given an array of integers, find two numbers that add up to a target.",
                "difficulty": "Easy",
                "tags": ["Array", "Hash Table"],
                "constraints": "1 <= nums.length <= 10^4",
                "sample_input": "nums = [2,7,11,15], target = 9",
                "sample_output": "[0,1]",
                "test_cases": {
                    "inputs": [{"nums": [2,7,11,15], "target": 9}],
                    "outputs": [[0,1]]
                },
                "is_active": True
            }
        ]
    }
    
    return Response(
        content=json.dumps(template, indent=2),
        media_type="application/json",
        headers={
            "Content-Disposition": "attachment; filename=coding_problem_upload_template.json"
        }
    )


@router.post("/departments")
async def bulk_upload_departments(
    file: UploadFile = File(...),
    college_id: Optional[int] = None,
    current_user_tuple = Depends(get_current_admin_or_super),
    db: Session = Depends(get_db)
):
    """Bulk upload departments/branches from CSV or Excel file.
    
    College admin can upload departments for their college (college_id auto-detected).
    Super admin can upload for any college (college_id required if not provided).
    
    Supported formats: .csv, .xlsx, .xls
    
    Format (headers required):
    name,code,branch_id,number_of_years,vertical
    
    Note: college_id is auto-detected from user context. Super admins can optionally provide it.
    Note: branch_id should be unique. If not provided, will be auto-generated from code.
    
    Example:
    name,code,branch_id,number_of_years,vertical
    Computer Science,CSE,CSE001,4,B.Tech
    Electronics,ECE,ECE001,4,B.Tech
    Mechanical Engineering,ME,ME001,4,B.E
    """
    from app.models.academic import Department
    
    current_user, is_super_admin = current_user_tuple
    
    # Auto-detect college_id if not provided
    if not college_id:
        if is_super_admin:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="College ID is required for super admin uploads. Provide it as query parameter: ?college_id=X"
            )
        else:
            # Auto-detect from user context
            auto_college_id = get_admin_college_id(current_user, db)
            if not auto_college_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Could not determine college ID. Admin must be associated with a college."
                )
            college_id = auto_college_id
    
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only CSV and Excel files (.csv, .xlsx, .xls) are supported"
        )
    
    # Read file content
    contents = await file.read()
    
    # Parse based on file type
    if file_ext == 'csv':
        csv_content = contents.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))
    else:
        # Excel file
        excel_file = io.BytesIO(contents)
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
        sheet = workbook.active
        
        # Get headers from first row and normalize to lowercase
        headers = [cell.value for cell in sheet[1]]
        if not headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must have headers in the first row"
            )
        
        # Normalize headers to lowercase for case-insensitive matching
        normalized_headers = [str(h).strip().lower() if h else None for h in headers]
        
        # Convert Excel rows to CSV-like dict format
        csv_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for idx, header in enumerate(normalized_headers):
                if header:
                    # Safely access row[idx] - handle cases where row has fewer columns than headers
                    if idx < len(row):
                        row_dict[header] = str(row[idx]).strip() if row[idx] is not None else ""
                    else:
                        row_dict[header] = ""
            csv_rows.append(row_dict)
        
        csv_reader = csv_rows if csv_rows else []
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    # Handle both CSV reader (iterator) and Excel rows (list)
    rows_to_process = list(csv_reader) if isinstance(csv_reader, list) else csv_reader
    
    for row_num, row in enumerate(rows_to_process, start=2):  # Start at 2 (row 1 is header)
        try:
            # Row keys are already normalized to lowercase from Excel parsing
            name = row.get('name', '').strip() if isinstance(row.get('name'), str) else str(row.get('name', '')).strip()
            code = row.get('code', '').strip() if isinstance(row.get('code'), str) else str(row.get('code', '')).strip()
            branch_id = row.get('branch_id', '').strip() if isinstance(row.get('branch_id'), str) else str(row.get('branch_id', '')).strip()
            row_college_id = row.get('college_id', '').strip() if isinstance(row.get('college_id'), str) else str(row.get('college_id', '')).strip()
            # Handle number_of_years - can be string or number
            number_of_years_val = row.get('number_of_years', '') or row.get('number of years', '') or row.get('years', '')
            if number_of_years_val is None:
                number_of_years_str = ''
            elif isinstance(number_of_years_val, (int, float)):
                number_of_years_str = str(int(number_of_years_val))
            else:
                number_of_years_str = str(number_of_years_val).strip()
            vertical = row.get('vertical', '').strip() if isinstance(row.get('vertical'), str) else str(row.get('vertical', '')).strip()
            
            # Use college_id from query param (already auto-detected if not provided)
            final_college_id = college_id
            
            if not name:
                results["failed"].append({
                    "row": row_num,
                    "name": name or "N/A",
                    "error": "Name is required"
                })
                continue
            
            if not final_college_id:
                results["failed"].append({
                    "row": row_num,
                    "name": name,
                    "error": "College ID is required (provide in query param or CSV)"
                })
                continue
            
            # Auto-generate branch_id from code if not provided
            if not branch_id and code:
                branch_id = code.upper()
            elif not branch_id:
                # Generate from name
                branch_id = name.upper().replace(' ', '_')[:20]
            
            # Check if branch_id already exists (only active departments)
            existing_dept_by_branch = db.query(Department).filter(
                Department.branch_id == branch_id,
                Department.is_active == True
            ).first()
            if existing_dept_by_branch:
                results["failed"].append({
                    "row": row_num,
                    "name": name,
                    "error": f"Branch ID '{branch_id}' already exists"
                })
                continue
            
            # Check if department with same name and college already exists (only active)
            existing_name = db.query(Department).filter(
                Department.name == name,
                Department.college_id == final_college_id,
                Department.is_active == True
            ).first()
            if existing_name:
                results["failed"].append({
                    "row": row_num,
                    "name": name,
                    "error": f"Department '{name}' already exists for this college"
                })
                continue
            
            # Check if there's an inactive department with same name/college - reactivate it instead
            inactive_dept = db.query(Department).filter(
                Department.name == name,
                Department.college_id == final_college_id,
                Department.is_active == False
            ).first()
            
            if inactive_dept:
                # Reactivate and update the inactive department
                inactive_dept.is_active = True
                if code:
                    inactive_dept.code = code
                if branch_id:
                    inactive_dept.branch_id = branch_id
                if number_of_years_str:
                    try:
                        inactive_dept.number_of_years = int(number_of_years_str)
                    except ValueError:
                        pass
                if vertical:
                    inactive_dept.vertical = vertical
                
                db.commit()
                results["success"].append({
                    "row": row_num,
                    "name": name,
                    "code": code,
                    "branch_id": branch_id or inactive_dept.branch_id,
                    "note": "Reactivated existing inactive department"
                })
                results["total"] += 1
                continue
            
            # Check if there's an inactive department with same branch_id - reactivate it
            inactive_dept_by_branch = db.query(Department).filter(
                Department.branch_id == branch_id,
                Department.is_active == False
            ).first()
            
            if inactive_dept_by_branch:
                # Reactivate and update the inactive department
                inactive_dept_by_branch.is_active = True
                inactive_dept_by_branch.name = name
                if code:
                    inactive_dept_by_branch.code = code
                if number_of_years_str:
                    try:
                        inactive_dept_by_branch.number_of_years = int(number_of_years_str)
                    except ValueError:
                        pass
                if vertical:
                    inactive_dept_by_branch.vertical = vertical
                inactive_dept_by_branch.college_id = final_college_id
                
                db.commit()
                results["success"].append({
                    "row": row_num,
                    "name": name,
                    "code": code,
                    "branch_id": branch_id,
                    "note": "Reactivated existing inactive department"
                })
                results["total"] += 1
                continue
            
            # Parse number_of_years
            number_of_years = None
            if number_of_years_str:
                try:
                    number_of_years = int(number_of_years_str)
                except ValueError:
                    pass
            
            # Create department
            department = Department(
                name=name,
                code=code if code else None,
                branch_id=branch_id,
                college_id=final_college_id,
                number_of_years=number_of_years,
                vertical=vertical if vertical else None,
                is_active=True
            )
            db.add(department)
            
            results["success"].append({
                "row": row_num,
                "name": name,
                "code": code,
                "branch_id": branch_id
            })
            results["total"] += 1
            
        except Exception as e:
            results["failed"].append({
                "row": row_num,
                "name": row.get('name', 'N/A'),
                "error": str(e)
            })
    
    # Commit all successful inserts
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed: {len(results['success'])} successful, {len(results['failed'])} failed",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.get("/template/departments")
async def download_department_template(
    format: str = "csv",  # csv or xlsx
    college_id: Optional[int] = None,  # Optional college_id for auto-linking
    current_user_tuple = Depends(get_current_admin_or_super)
):
    """Download CSV or Excel template for department/branch bulk upload
    
    If college_id is provided as query parameter, it will be auto-set and not included in template.
    Otherwise, college_id column will be included in template.
    """
    from fastapi.responses import StreamingResponse
    
    # If college_id is provided, don't include it in template (will be auto-set)
    include_college_id = college_id is None
    
    if format.lower() == "xlsx":
        # Excel format
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Departments"
        
        # Headers with formatting - exclude college_id if provided in query param
        headers = ["name", "code", "branch_id", "number_of_years", "vertical"]
        if include_college_id:
            headers.insert(3, "college_id")  # Insert college_id after branch_id
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sample data rows
        sample_data = [
            ["Computer Science", "CSE", "CSE001", "4", "B.Tech"],
            ["Electronics", "ECE", "ECE001", "4", "B.Tech"],
            ["Mechanical Engineering", "ME", "ME001", "4", "B.E"]
        ]
        if include_college_id:
            # Add college_id to sample data
            sample_data = [
                ["Computer Science", "CSE", "CSE001", "1", "4", "B.Tech"],
                ["Electronics", "ECE", "ECE001", "1", "4", "B.Tech"],
                ["Mechanical Engineering", "ME", "ME001", "1", "4", "B.E"]
            ]
        
        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        column_widths = [25, 12, 15, 15, 15]  # Without college_id
        if include_college_id:
            column_widths = [25, 12, 15, 12, 15, 15]  # With college_id
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_content = excel_buffer.getvalue()
        excel_buffer.close()
        
        return StreamingResponse(
            iter([excel_content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=department_upload_template.xlsx",
                "Content-Length": str(len(excel_content))
            }
        )
    else:
        # CSV format (default)
        if include_college_id:
            template = "name,code,branch_id,college_id,number_of_years,vertical\n"
            template += "Computer Science,CSE,CSE001,1,4,B.Tech\n"
            template += "Electronics,ECE,ECE001,1,4,B.Tech\n"
            template += "Mechanical Engineering,ME,ME001,1,4,B.E\n"
        else:
            template = "name,code,branch_id,number_of_years,vertical\n"
            template += "Computer Science,CSE,CSE001,4,B.Tech\n"
            template += "Electronics,ECE,ECE001,4,B.Tech\n"
            template += "Mechanical Engineering,ME,ME001,4,B.E\n"
        
        return StreamingResponse(
            iter([template.encode('utf-8')]),
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=department_upload_template.csv",
                "Content-Length": str(len(template.encode('utf-8')))
        }
    )


@router.post("/subjects")
async def bulk_upload_subjects(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Bulk upload subjects from Excel, CSV, or JSON file (Super Admin only).
    
    Supported formats: .xlsx, .xls, .csv, .json
    
    Excel/CSV Format (headers required):
    name,code,college_id,department_id,semester_id,year,credits
    
    JSON Format:
    [
      {
        "name": "Data Structures",
        "code": "CS301",
        "college_id": 1,
        "department_id": 1,
        "semester_id": 1,
        "year": "3rd",
        "credits": 4
      }
    ]
    """
    file_ext = file.filename.lower().split('.')[-1]
    
    if file_ext not in ['json', 'csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only JSON, CSV, and Excel files (.json, .csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    data = []
    
    # Parse based on file type
    if file_ext == 'json':
        data = json.loads(contents.decode('utf-8'))
        if not isinstance(data, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="JSON must be an array of subject objects"
            )
    else:
        # Excel or CSV
        if file_ext in ['xlsx', 'xls']:
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [str(cell.value).lower().strip() if cell.value else '' for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                subject_dict = dict(zip(headers, row))
                data.append(subject_dict)
        else:  # CSV
            csv_content = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_content))
            data = list(reader)
            # Normalize headers to lowercase
            data = [{k.lower().strip(): v for k, v in row.items()} for row in data]
    
    results = {
        "success": [],
        "failed": [],
        "total": 0
    }
    
    for idx, subject_data in enumerate(data, start=1):
        try:
            # Validate required fields
            if not subject_data.get('name'):
                results["failed"].append({
                    "index": idx,
                    "name": subject_data.get('name', 'N/A'),
                    "error": "Name is required"
                })
                continue
            
            name = str(subject_data.get('name')).strip()
            code = str(subject_data.get('code', '')).strip() if subject_data.get('code') else None
            college_id = None
            department_id = None
            semester_id = None
            year = None
            credits = None
            
            # Parse college_id
            if subject_data.get('college_id'):
                try:
                    college_id = int(subject_data.get('college_id'))
                except:
                    results["failed"].append({
                        "index": idx,
                        "name": name,
                        "error": f"Invalid college_id: {subject_data.get('college_id')}"
                    })
                    continue
            
            # Parse department_id
            if subject_data.get('department_id'):
                try:
                    department_id = int(subject_data.get('department_id'))
                except:
                    pass
            
            # Parse semester_id
            if subject_data.get('semester_id'):
                try:
                    semester_id = int(subject_data.get('semester_id'))
                except:
                    pass
            
            # Parse year
            if subject_data.get('year'):
                year = str(subject_data.get('year')).strip()
            
            # Parse credits
            if subject_data.get('credits'):
                try:
                    credits = int(subject_data.get('credits'))
                except:
                    pass
            
            # Validate college exists
            if college_id:
                college = db.query(College).filter(College.id == college_id).first()
                if not college:
                    results["failed"].append({
                        "index": idx,
                        "name": name,
                        "error": f"College with id {college_id} not found"
                    })
                    continue
            
            # Validate department exists if provided
            if department_id:
                dept = db.query(Department).filter(Department.id == department_id).first()
                if not dept:
                    results["failed"].append({
                        "index": idx,
                        "name": name,
                        "error": f"Department with id {department_id} not found"
                    })
                    continue
                if college_id and dept.college_id != college_id:
                    results["failed"].append({
                        "index": idx,
                        "name": name,
                        "error": f"Department {department_id} does not belong to college {college_id}"
                    })
                    continue
            
            # Check for duplicates
            existing = db.query(Subject).filter(
                Subject.name == name,
                Subject.college_id == college_id if college_id else True
            ).first()
            
            if existing:
                # Update existing
                existing.code = code or existing.code
                existing.department_id = department_id or existing.department_id
                existing.semester_id = semester_id or existing.semester_id
                existing.year = year or existing.year
                existing.credits = credits or existing.credits
                existing.updated_at = datetime.utcnow()
                
                db.flush()
                
                results["success"].append({
                    "index": idx,
                    "name": name,
                    "id": existing.id,
                    "action": "updated"
                })
            else:
                # Create new
                subject = Subject(
                    name=name,
                    code=code,
                    college_id=college_id,
                    department_id=department_id,
                    semester_id=semester_id,
                    year=year,
                    credits=credits,
                    is_active=True
                )
                db.add(subject)
                db.flush()
                
                results["success"].append({
                    "index": idx,
                    "name": name,
                    "id": subject.id,
                    "action": "created"
                })
            
            results["total"] += 1
            
        except Exception as e:
            results["failed"].append({
                "index": idx,
                "name": subject_data.get('name', 'N/A'),
                "error": str(e)
            })
    
    if results["success"]:
        db.commit()
    
    return {
        "message": f"Bulk upload completed. {results['total']} subjects processed.",
        "success_count": len(results["success"]),
        "failed_count": len(results["failed"]),
        "success": results["success"],
        "failed": results["failed"]
    }


@router.post("/academic-structure")
async def bulk_upload_academic_structure(
    file: UploadFile = File(...),
    academic_year_id: Optional[int] = Query(None, description="Academic year ID (optional, will use current if not provided)"),
    college_id: Optional[int] = Query(None, description="College ID (optional, auto-detected for admin/HOD)"),
    create_subjects_if_missing: bool = Query(True, description="Create subjects if they don't exist"),
    create_sections_if_missing: bool = Query(True, description="Create sections if they don't exist"),
    current_user_tuple = Depends(get_current_admin_or_super),
    db: Session = Depends(get_db)
):
    """
    Bulk upload academic structure: subjects + assignments in one go
    
    CSV/Excel format:
    - Subject Name, Subject Code, Department Code, Year, Semester, Section, Faculty Email, Faculty Staff ID, Academic Year
    
    This endpoint:
    1. Creates/finds subjects
    2. Creates/finds sections
    3. Creates subject assignments (faculty → subject → section)
    4. Creates faculty-section assignments
    """
    current_user, is_super_admin = current_user_tuple
    
    # Auto-detect college_id if not provided
    if not college_id:
        college_id = get_admin_college_id(current_user, db)
        if not college_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="college_id is required. Could not auto-detect from user roles."
            )
    
    # Get or use current academic year
    if not academic_year_id:
        current_academic_year = db.query(AcademicYear).filter(
            AcademicYear.is_current == True
        ).first()
        if not current_academic_year:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No current academic year found. Please create one first."
            )
        academic_year_id = current_academic_year.id
    else:
        academic_year = db.query(AcademicYear).filter(AcademicYear.id == academic_year_id).first()
        if not academic_year:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Academic year {academic_year_id} not found"
            )
    
    # Verify college access
    if not is_super_admin:
        admin_college_id = get_admin_college_id(current_user, db)
        if college_id != admin_college_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only upload academic structure for your own college"
            )
    
    # Read file
    contents = await file.read()
    file_extension = file.filename.split('.')[-1].lower() if '.' in file.filename else ''
    
    results = {
        "total_rows": 0,
        "successful": 0,
        "failed": 0,
        "created_subjects": 0,
        "created_sections": 0,
        "created_assignments": 0,
        "created_faculty_section_assignments": 0,
        "errors": []
    }
    
    rows = []
    
    try:
        if file_extension in ['csv']:
            # Parse CSV
            content_str = contents.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content_str))
            rows = list(csv_reader)
        elif file_extension in ['xlsx', 'xls']:
            # Parse Excel
            workbook = load_workbook(io.BytesIO(contents), data_only=True)
            sheet = workbook.active
            
            # Get headers (normalize to lowercase)
            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
            
            # Read rows
            for row in sheet.iter_rows(min_row=2, values_only=False):
                row_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        cell_value = row[idx].value
                        row_dict[header] = str(cell_value).strip() if cell_value else ''
                
                # Check if row has data
                if any(str(v).strip() for v in row_dict.values() if v):
                    rows.append(row_dict)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported file format. Please upload CSV or Excel file."
            )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error parsing file: {str(e)}"
        )
    
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data found in file"
        )
    
    results["total_rows"] = len(rows)
    
    # Normalize headers (case-insensitive)
    normalized_headers = {
        'subject name': 'subject_name',
        'subject_code': 'subject_code',
        'subject code': 'subject_code',
        'department code': 'department_code',
        'department_code': 'department_code',
        'year': 'year',
        'semester': 'semester',
        'section': 'section',
        'faculty email': 'faculty_email',
        'faculty_email': 'faculty_email',
        'faculty staff id': 'faculty_staff_id',
        'faculty_staff_id': 'faculty_staff_id',
        'academic year': 'academic_year',
        'academic_year': 'academic_year'
    }
    
    for idx, row in enumerate(rows, start=2):  # Start at 2 (row 1 is header)
        try:
            # Normalize row keys
            normalized_row = {}
            for key, value in row.items():
                normalized_key = normalized_headers.get(key.lower(), key.lower())
                normalized_row[normalized_key] = value.strip() if value else ''
            
            subject_name = normalized_row.get('subject_name', '').strip()
            subject_code = normalized_row.get('subject_code', '').strip()
            department_code = normalized_row.get('department_code', '').strip()
            year = normalized_row.get('year', '').strip()
            semester_name = normalized_row.get('semester', '').strip()
            section_name = normalized_row.get('section', '').strip()
            faculty_email = normalized_row.get('faculty_email', '').strip()
            faculty_staff_id = normalized_row.get('faculty_staff_id', '').strip()
            
            # Validation
            if not subject_name:
                results["errors"].append({
                    "row": idx,
                    "error": "Subject Name is required"
                })
                results["failed"] += 1
                continue
            
            if not department_code:
                results["errors"].append({
                    "row": idx,
                    "error": "Department Code is required"
                })
                results["failed"] += 1
                continue
            
            if not year:
                results["errors"].append({
                    "row": idx,
                    "error": "Year is required"
                })
                results["failed"] += 1
                continue
            
            if not semester_name:
                results["errors"].append({
                    "row": idx,
                    "error": "Semester is required"
                })
                results["failed"] += 1
                continue
            
            if not section_name:
                results["errors"].append({
                    "row": idx,
                    "error": "Section is required"
                })
                results["failed"] += 1
                continue
            
            if not faculty_email and not faculty_staff_id:
                results["errors"].append({
                    "row": idx,
                    "error": "Either Faculty Email or Faculty Staff ID is required"
                })
                results["failed"] += 1
                continue
            
            # Find department
            department = db.query(Department).filter(
                Department.code == department_code.upper(),
                Department.college_id == college_id
            ).first()
            
            if not department:
                results["errors"].append({
                    "row": idx,
                    "error": f"Department with code '{department_code}' not found"
                })
                results["failed"] += 1
                continue
            
            # Find or create subject
            subject = None
            if subject_code:
                subject = db.query(Subject).filter(
                    Subject.code == subject_code,
                    Subject.college_id == college_id
                ).first()
            
            if not subject:
                subject = db.query(Subject).filter(
                    Subject.name == subject_name,
                    Subject.college_id == college_id,
                    Subject.department_id == department.id
                ).first()
            
            if not subject:
                if create_subjects_if_missing:
                    subject = Subject(
                        name=subject_name,
                        code=subject_code or None,
                        college_id=college_id,
                        department_id=department.id,
                        year=year,
                        is_active=True
                    )
                    db.add(subject)
                    db.flush()
                    results["created_subjects"] += 1
                else:
                    results["errors"].append({
                        "row": idx,
                        "error": f"Subject '{subject_name}' not found and create_subjects_if_missing is False"
                    })
                    results["failed"] += 1
                    continue
            else:
                # Update subject if needed
                if subject_code and not subject.code:
                    subject.code = subject_code
                if not subject.year:
                    subject.year = year
            
            # Find semester
            semester = None
            if semester_name:
                # Try to find by name
                semester = db.query(Semester).filter(
                    Semester.name.ilike(f"%{semester_name}%"),
                    Semester.college_id == college_id,
                    Semester.academic_year_id == academic_year_id
                ).first()
                
                # Try to find by number if name doesn't work
                if not semester:
                    try:
                        sem_num = int(semester_name)
                        semester = db.query(Semester).filter(
                            Semester.number == sem_num,
                            Semester.college_id == college_id,
                            Semester.academic_year_id == academic_year_id
                        ).first()
                    except ValueError:
                        pass
            
            if not semester:
                results["errors"].append({
                    "row": idx,
                    "error": f"Semester '{semester_name}' not found for academic year {academic_year_id}"
                })
                results["failed"] += 1
                continue
            
            # Update subject semester if needed
            if not subject.semester_id:
                subject.semester_id = semester.id
            
            # Find or create section
            section = None
            # Try to find section by name, department, year, semester
            section = db.query(Section).filter(
                Section.name == section_name,
                Section.department_id == department.id,
                Section.college_id == college_id,
                Section.semester_id == semester.id
            ).first()
            
            if not section and create_sections_if_missing:
                # Create section
                # Determine year as integer from string
                year_int = None
                if year:
                    year_mapping = {"1st": 1, "2nd": 2, "3rd": 3, "4th": 4, "5th": 5}
                    year_int = year_mapping.get(year.lower(), None)
                
                section = Section(
                    name=section_name,
                    college_id=college_id,
                    department_id=department.id,
                    semester_id=semester.id,
                    year=year_int,
                    is_active=True
                )
                db.add(section)
                db.flush()
                results["created_sections"] += 1
            
            if not section:
                results["errors"].append({
                    "row": idx,
                    "error": f"Section '{section_name}' not found and create_sections_if_missing is False"
                })
                results["failed"] += 1
                continue
            
            # Find faculty
            faculty = None
            if faculty_email:
                faculty_user = db.query(User).filter(User.email == faculty_email).first()
                if faculty_user:
                    # Check if user has faculty role
                    faculty_role = db.query(UserRole).filter(
                        UserRole.user_id == faculty_user.id,
                        UserRole.role == RoleEnum.FACULTY,
                        UserRole.college_id == college_id
                    ).first()
                    if faculty_role:
                        faculty = faculty_user
            
            if not faculty and faculty_staff_id:
                faculty_profile = db.query(Profile).filter(
                    Profile.staff_id == faculty_staff_id,
                    Profile.college_id == college_id
                ).first()
                if faculty_profile:
                    faculty_role = db.query(UserRole).filter(
                        UserRole.user_id == faculty_profile.user_id,
                        UserRole.role == RoleEnum.FACULTY,
                        UserRole.college_id == college_id
                    ).first()
                    if faculty_role:
                        faculty = db.query(User).filter(User.id == faculty_profile.user_id).first()
            
            if not faculty:
                results["errors"].append({
                    "row": idx,
                    "error": f"Faculty not found (email: {faculty_email}, staff_id: {faculty_staff_id})"
                })
                results["failed"] += 1
                continue
            
            # Create subject assignment
            existing_assignment = db.query(SubjectAssignment).filter(
                SubjectAssignment.faculty_id == faculty.id,
                SubjectAssignment.subject_id == subject.id,
                SubjectAssignment.semester_id == semester.id,
                SubjectAssignment.section_id == section.id
            ).first()
            
            if not existing_assignment:
                assignment = SubjectAssignment(
                    faculty_id=faculty.id,
                    subject_id=subject.id,
                    semester_id=semester.id,
                    section=section_name,
                    section_id=section.id,
                    is_active=True,
                    assigned_by=current_user.id
                )
                db.add(assignment)
                db.flush()
                results["created_assignments"] += 1
            else:
                # Update if inactive
                if not existing_assignment.is_active:
                    existing_assignment.is_active = True
                    existing_assignment.assigned_by = current_user.id
                    existing_assignment.updated_at = datetime.utcnow()
                    results["created_assignments"] += 1
            
            # Create faculty-section assignment (if not exists)
            existing_faculty_section = db.query(FacultySectionAssignment).filter(
                FacultySectionAssignment.faculty_id == faculty.id,
                FacultySectionAssignment.section_id == section.id
            ).first()
            
            if not existing_faculty_section:
                faculty_section = FacultySectionAssignment(
                    faculty_id=faculty.id,
                    section_id=section.id,
                    is_active=True,
                    assigned_by=current_user.id
                )
                db.add(faculty_section)
                db.flush()
                results["created_faculty_section_assignments"] += 1
            else:
                # Update if inactive
                if not existing_faculty_section.is_active:
                    existing_faculty_section.is_active = True
                    existing_faculty_section.assigned_by = current_user.id
                    existing_faculty_section.updated_at = datetime.utcnow()
                    results["created_faculty_section_assignments"] += 1
            
            results["successful"] += 1
            
        except Exception as e:
            results["errors"].append({
                "row": idx,
                "error": str(e)
            })
            results["failed"] += 1
    
    if results["successful"] > 0:
        db.commit()
    
    return {
        "message": f"Academic structure upload completed. {results['successful']} rows processed successfully.",
        "total_rows": results["total_rows"],
        "successful": results["successful"],
        "failed": results["failed"],
        "created_subjects": results["created_subjects"],
        "created_sections": results["created_sections"],
        "created_assignments": results["created_assignments"],
        "created_faculty_section_assignments": results["created_faculty_section_assignments"],
        "errors": results["errors"][:50]  # Limit to first 50 errors
    }


@router.post("/questions")
async def bulk_upload_questions(
    file: UploadFile = File(...),
    current_user_tuple = Depends(get_current_content_creator),
    db: Session = Depends(get_db)
):
    """Bulk upload questions from Excel, CSV, or JSON file.
    
    Supported formats: .xlsx, .xls, .csv, .json
    
    Question Types:
    - mcq: Multiple Choice Question (requires option_a, option_b, option_c, option_d, correct_answer)
    - fill_blank: Fill in the blank (requires correct_answer_text)
    - true_false: True/False (requires is_true or correct_answer='A' for True, 'B' for False)
    
    Excel/CSV Format (headers required):
    question,question_type,option_a,option_b,option_c,option_d,correct_answer,correct_answer_text,is_true,marks,timer_seconds
    
    Example rows:
    MCQ: "What is 2+2?","mcq","2","3","4","5","C","",,1,60
    Fill Blank: "The capital of France is ____","fill_blank","","","","","","Paris",,1,30
    True/False: "Python is a programming language","true_false","True","False","","","A","",,1,20
    """
    current_user_obj, user_info = current_user_tuple
    
    file_ext = file.filename.lower().split('.')[-1]
    
    if file_ext not in ['json', 'csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only JSON, CSV, and Excel files (.json, .csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    data = []
    
    # Parse based on file type
    if file_ext == 'json':
        data = json.loads(contents.decode('utf-8'))
        if not isinstance(data, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="JSON must be an array of question objects"
            )
    else:
        # Excel or CSV
        if file_ext in ['xlsx', 'xls']:
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(str(v).strip() if v else "" for v in row):
                    continue
                question_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        question_dict[header] = row[idx]
                data.append(question_dict)
        else:  # CSV
            csv_content = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_content))
            # Normalize CSV headers to lowercase (case-insensitive)
            data = []
            for row in reader:
                normalized_row = {k.lower().strip(): v for k, v in row.items() if k}
                data.append(normalized_row)
    
    questions = []
    errors = []
    
    for idx, row in enumerate(data, start=1):
        try:
            # Keys are already normalized to lowercase from parsing above
            row_dict = row
            
            question_text = str(row_dict.get('question', '')).strip()
            if not question_text:
                errors.append(f"Row {idx}: Question text is required")
                continue
            
            question_type = str(row_dict.get('question_type', 'mcq')).lower().strip()
            if question_type not in ['mcq', 'fill_blank', 'true_false']:
                question_type = 'mcq'  # Default to MCQ
            
            marks = int(row_dict.get('marks', 1)) if row_dict.get('marks') else 1
            timer_seconds = int(row_dict.get('timer_seconds')) if row_dict.get('timer_seconds') else None
            
            question_obj = {
                "question": question_text,
                "question_type": question_type,
                "marks": marks,
            }
            
            if timer_seconds:
                question_obj["timer_seconds"] = timer_seconds
            
            # Handle different question types
            if question_type == 'mcq':
                option_a = str(row_dict.get('option_a', '')).strip()
                option_b = str(row_dict.get('option_b', '')).strip()
                option_c = str(row_dict.get('option_c', '')).strip()
                option_d = str(row_dict.get('option_d', '')).strip()
                correct_answer = str(row_dict.get('correct_answer', '')).upper().strip()
                
                if not all([option_a, option_b, option_c, option_d]):
                    errors.append(f"Row {idx}: MCQ requires all four options (option_a, option_b, option_c, option_d)")
                    continue
                
                if correct_answer not in ['A', 'B', 'C', 'D']:
                    errors.append(f"Row {idx}: MCQ correct_answer must be A, B, C, or D")
                    continue
                
                question_obj.update({
                    "option_a": option_a,
                    "option_b": option_b,
                    "option_c": option_c,
                    "option_d": option_d,
                    "correct_answer": correct_answer
                })
            
            elif question_type == 'fill_blank':
                correct_answer_text = str(row_dict.get('correct_answer_text', '')).strip()
                if not correct_answer_text:
                    errors.append(f"Row {idx}: fill_blank requires correct_answer_text")
                    continue
                
                question_obj["correct_answer_text"] = correct_answer_text
            
            elif question_type == 'true_false':
                # Support both is_true boolean and correct_answer (A=True, B=False)
                is_true = row_dict.get('is_true')
                correct_answer = str(row_dict.get('correct_answer', '')).upper().strip()
                
                if is_true is not None:
                    # Convert to boolean
                    if isinstance(is_true, str):
                        is_true = is_true.lower() in ['true', '1', 'yes', 't']
                    question_obj["is_true"] = bool(is_true)
                    question_obj["option_a"] = "True"
                    question_obj["option_b"] = "False"
                    question_obj["correct_answer"] = "A" if is_true else "B"
                elif correct_answer in ['A', 'B']:
                    question_obj["option_a"] = "True"
                    question_obj["option_b"] = "False"
                    question_obj["correct_answer"] = correct_answer
                    question_obj["is_true"] = (correct_answer == 'A')
                else:
                    errors.append(f"Row {idx}: true_false requires is_true (true/false) or correct_answer (A/B)")
                    continue
            
            questions.append(question_obj)
            
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    return {
        "message": f"Questions parsed: {len(questions)} valid, {len(errors)} errors",
        "questions": questions,
        "errors": errors[:50]  # Limit to first 50 errors
    }


@router.get("/templates/questions")
async def download_question_template(
    format: str = "xlsx",  # xlsx, csv, or json
    current_user_tuple = Depends(get_current_content_creator)
):
    """Download template for question bulk upload in Excel, CSV, or JSON format"""
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    if format == "json":
        template = [
            {
                "question": "What is 2+2?",
                "question_type": "mcq",
                "option_a": "2",
                "option_b": "3",
                "option_c": "4",
                "option_d": "5",
                "correct_answer": "C",
                "marks": 1,
                "timer_seconds": 60
            },
            {
                "question": "The capital of France is ____",
                "question_type": "fill_blank",
                "correct_answer_text": "Paris",
                "marks": 1,
                "timer_seconds": 30
            },
            {
                "question": "Python is a programming language",
                "question_type": "true_false",
                "is_true": True,
                "marks": 1,
                "timer_seconds": 20
            },
        ]
        return Response(
            content=json.dumps(template, indent=2),
            media_type="application/json",
            headers={
                "Content-Disposition": "attachment; filename=question_upload_template.json"
            }
        )
    elif format == "csv":
        csv_content = "question,question_type,option_a,option_b,option_c,option_d,correct_answer,correct_answer_text,is_true,marks,timer_seconds\n"
        csv_content += "What is 2+2?,mcq,2,3,4,5,C,,,1,60\n"
        csv_content += "The capital of France is ____,fill_blank,,,,,Paris,,1,30\n"
        csv_content += "Python is a programming language,true_false,True,False,,,A,,True,1,20\n"
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=question_upload_template.csv"
            }
        )
    else:  # xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Question Template"
        
        # Headers
        headers = ["question", "question_type", "option_a", "option_b", "option_c", "option_d", "correct_answer", "correct_answer_text", "is_true", "marks", "timer_seconds"]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Sample rows for different question types
        sample_rows = [
            ["What is 2+2?", "mcq", "2", "3", "4", "5", "C", "", "", 1, 60],
            ["The capital of France is ____", "fill_blank", "", "", "", "", "", "Paris", "", 1, 30],
            ["Python is a programming language", "true_false", "True", "False", "", "", "A", "", "True", 1, 20]
        ]
        
        for row_idx, row_data in enumerate(sample_rows, start=2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=question_upload_template.xlsx"
            }
        )


@router.post("/company-training/round-contents/quiz")
async def bulk_upload_company_training_quiz_questions(
    round_id: int = Query(..., description="Round ID to add questions to"),
    file: UploadFile = File(...),
    current_user_tuple = Depends(get_current_content_creator),
    db: Session = Depends(get_db)
):
    """Bulk upload quiz questions for company training round.
    
    Uses the same format as /bulk-upload/questions but creates RoundContent entries.
    Supports mcq, fill_blank, and true_false question types with timers.
    """
    from app.models.company_training import Round as RoundModel
    
    # Verify round exists
    round_obj = db.query(RoundModel).filter(RoundModel.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Round {round_id} not found"
        )
    
    if round_obj.round_type != RoundType.QUIZ:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Round {round_id} is not a quiz round"
        )
    
    # Parse file using same logic as bulk_upload_questions
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['json', 'csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only JSON, CSV, and Excel files (.json, .csv, .xlsx, .xls) are supported"
        )
    
    contents = await file.read()
    data = []
    
    if file_ext == 'json':
        data = json.loads(contents.decode('utf-8'))
        if not isinstance(data, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="JSON must be an array of question objects"
            )
    else:
        if file_ext in ['xlsx', 'xls']:
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(str(v).strip() if v else "" for v in row):
                    continue
                question_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        question_dict[header] = row[idx]
                data.append(question_dict)
        else:  # CSV
            csv_content = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_content))
            data = []
            for row in reader:
                normalized_row = {k.lower().strip(): v for k, v in row.items() if k}
                data.append(normalized_row)
    
    created = []
    errors = []
    current_user_obj, user_info = current_user_tuple
    
    # Get max order_index for this round
    max_order = db.query(db.func.max(RoundContentModel.order_index)).filter(
        RoundContentModel.round_id == round_id
    ).scalar() or -1
    
    for idx, row in enumerate(data, start=1):
        try:
            row_dict = row
            
            question_text = str(row_dict.get('question', '')).strip()
            if not question_text:
                errors.append(f"Row {idx}: Question text is required")
                continue
            
            question_type = str(row_dict.get('question_type', 'mcq')).lower().strip()
            if question_type not in ['mcq', 'fill_blank', 'true_false']:
                question_type = 'mcq'
            
            marks = int(row_dict.get('marks', 1)) if row_dict.get('marks') else 1
            timer_seconds = int(row_dict.get('timer_seconds')) if row_dict.get('timer_seconds') else None
            
            content_data = {
                "round_id": round_id,
                "quiz_question": question_text,
                "quiz_question_type": question_type,
                "quiz_marks": marks,
                "order_index": max_order + idx,
                "is_active": True,
            }
            
            if timer_seconds:
                content_data["quiz_timer_seconds"] = timer_seconds
            
            # Handle different question types
            if question_type == 'mcq':
                option_a = str(row_dict.get('option_a', '')).strip()
                option_b = str(row_dict.get('option_b', '')).strip()
                option_c = str(row_dict.get('option_c', '')).strip()
                option_d = str(row_dict.get('option_d', '')).strip()
                correct_answer = str(row_dict.get('correct_answer', '')).upper().strip()
                
                if not all([option_a, option_b, option_c, option_d]):
                    errors.append(f"Row {idx}: MCQ requires all four options")
                    continue
                
                if correct_answer not in ['A', 'B', 'C', 'D']:
                    errors.append(f"Row {idx}: MCQ correct_answer must be A, B, C, or D")
                    continue
                
                content_data.update({
                    "quiz_option_a": option_a,
                    "quiz_option_b": option_b,
                    "quiz_option_c": option_c,
                    "quiz_option_d": option_d,
                    "correct_answer": correct_answer
                })
            
            elif question_type == 'fill_blank':
                correct_answer_text = str(row_dict.get('correct_answer_text', '')).strip()
                if not correct_answer_text:
                    errors.append(f"Row {idx}: fill_blank requires correct_answer_text")
                    continue
                content_data["quiz_correct_answer_text"] = correct_answer_text
            
            elif question_type == 'true_false':
                is_true = row_dict.get('is_true')
                correct_answer = str(row_dict.get('correct_answer', '')).upper().strip()
                
                if is_true is not None:
                    if isinstance(is_true, str):
                        is_true = is_true.lower() in ['true', '1', 'yes', 't']
                    content_data["quiz_is_true"] = bool(is_true)
                    content_data["quiz_option_a"] = "True"
                    content_data["quiz_option_b"] = "False"
                    content_data["correct_answer"] = "A" if is_true else "B"
                elif correct_answer in ['A', 'B']:
                    content_data["quiz_option_a"] = "True"
                    content_data["quiz_option_b"] = "False"
                    content_data["correct_answer"] = correct_answer
                    content_data["quiz_is_true"] = (correct_answer == 'A')
                else:
                    errors.append(f"Row {idx}: true_false requires is_true or correct_answer")
                    continue
            
            content = RoundContentModel(**content_data)
            db.add(content)
            db.flush()
            created.append(content.id)
            
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    db.commit()
    
    return {
        "message": f"Created {len(created)} quiz questions, {len(errors)} errors",
        "created_count": len(created),
        "created_ids": created,
        "errors": errors[:50]
    }


@router.post("/company-training/round-contents/coding")
async def bulk_upload_company_training_coding_problems(
    round_id: int = Query(..., description="Round ID to add problems to"),
    file: UploadFile = File(...),
    current_user_tuple = Depends(get_current_content_creator),
    db: Session = Depends(get_db)
):
    """Bulk upload coding problems for company training round.
    
    Uses similar format to /bulk-upload/coding-problems but creates RoundContent entries.
    Supports exam timer mode.
    """
    from app.models.company_training import Round as RoundModel
    
    # Verify round exists
    round_obj = db.query(RoundModel).filter(RoundModel.id == round_id).first()
    if not round_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Round {round_id} not found"
        )
    
    if round_obj.round_type != RoundType.CODING:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Round {round_id} is not a coding round"
        )
    
    # Parse file (similar to coding problems bulk upload)
    file_ext = file.filename.lower().split('.')[-1]
    if file_ext not in ['json', 'csv', 'xlsx', 'xls']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only JSON, CSV, and Excel files are supported"
        )
    
    contents = await file.read()
    data = []
    
    if file_ext == 'json':
        data = json.loads(contents.decode('utf-8'))
        if not isinstance(data, list):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="JSON must be an array of problem objects"
            )
    else:
        if file_ext in ['xlsx', 'xls']:
            wb = load_workbook(io.BytesIO(contents))
            ws = wb.active
            headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(str(v).strip() if v else "" for v in row):
                    continue
                problem_dict = {}
                for idx, header in enumerate(headers):
                    if idx < len(row):
                        problem_dict[header] = row[idx]
                data.append(problem_dict)
        else:  # CSV
            csv_content = contents.decode('utf-8')
            reader = csv.DictReader(io.StringIO(csv_content))
            data = []
            for row in reader:
                normalized_row = {k.lower().strip(): v for k, v in row.items() if k}
                data.append(normalized_row)
    
    created = []
    errors = []
    current_user_obj, user_info = current_user_tuple
    
    # Get max order_index
    max_order = db.query(db.func.max(RoundContentModel.order_index)).filter(
        RoundContentModel.round_id == round_id
    ).scalar() or -1
    
    for idx, row in enumerate(data, start=1):
        try:
            row_dict = row
            
            title = str(row_dict.get('title', '')).strip()
            if not title:
                errors.append(f"Row {idx}: Title is required")
                continue
            
            description = str(row_dict.get('description', '')).strip()
            difficulty = str(row_dict.get('difficulty', 'medium')).lower().strip()
            if difficulty not in ['easy', 'medium', 'hard']:
                difficulty = 'medium'
            
            content_data = {
                "round_id": round_id,
                "coding_title": title,
                "coding_description": description,
                "coding_difficulty": difficulty,
                "coding_input_format": str(row_dict.get('input_format', '')).strip() or None,
                "coding_output_format": str(row_dict.get('output_format', '')).strip() or None,
                "coding_constraints": str(row_dict.get('constraints', '')).strip() or None,
                "coding_sample_input": str(row_dict.get('sample_input', '')).strip() or None,
                "coding_sample_output": str(row_dict.get('sample_output', '')).strip() or None,
                "coding_time_limit": int(row_dict.get('time_limit')) if row_dict.get('time_limit') else None,
                "coding_memory_limit": int(row_dict.get('memory_limit')) if row_dict.get('memory_limit') else None,
                "coding_starter_code_python": str(row_dict.get('starter_code_python', '')).strip() or None,
                "coding_starter_code_c": str(row_dict.get('starter_code_c', '')).strip() or None,
                "coding_starter_code_cpp": str(row_dict.get('starter_code_cpp', '')).strip() or None,
                "coding_starter_code_java": str(row_dict.get('starter_code_java', '')).strip() or None,
                "coding_starter_code_javascript": str(row_dict.get('starter_code_javascript', '')).strip() or None,
                "order_index": max_order + idx,
                "is_active": True,
            }
            
            # Exam timer support
            exam_timer_enabled = row_dict.get('exam_timer_enabled', '').lower().strip()
            if exam_timer_enabled in ['true', '1', 'yes', 't']:
                content_data["coding_exam_timer_enabled"] = True
                exam_duration = int(row_dict.get('exam_duration_minutes')) if row_dict.get('exam_duration_minutes') else None
                if exam_duration:
                    content_data["coding_exam_duration_minutes"] = exam_duration
            
            # Parse test cases (JSON string or array)
            test_cases_str = row_dict.get('test_cases', '')
            if test_cases_str:
                try:
                    if isinstance(test_cases_str, str):
                        test_cases = json.loads(test_cases_str)
                    else:
                        test_cases = test_cases_str
                    if isinstance(test_cases, list):
                        content_data["coding_test_cases"] = test_cases
                except:
                    errors.append(f"Row {idx}: Invalid test_cases JSON format")
                    continue
            
            content = RoundContentModel(**content_data)
            db.add(content)
            db.flush()
            created.append(content.id)
            
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")
    
    db.commit()
    
    return {
        "message": f"Created {len(created)} coding problems, {len(errors)} errors",
        "created_count": len(created),
        "created_ids": created,
        "errors": errors[:50]
    }

